# controllers/dashboard_1LaporanRekapController.py
from flask import render_template, request, send_file
from io import BytesIO
from sqlalchemy import func
from datetime import datetime, timedelta
import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.drawing.image import Image as XLImage


from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)

from reportlab.lib.pagesizes import (
    A4,
    landscape
)

from reportlab.lib.styles import (
    getSampleStyleSheet,
    ParagraphStyle
)

from reportlab.lib import colors
from app import db
from app.models.absensiModel import Absensi
from app.models.pegawaiModel import Pegawai
from app.models.kalenderModel import MfKalender
from app.models.dinasLuarModel import DinasLuar
from app.models.unitKerjaModel import MfUnitKerja
from app.models.timeRecorderModel import TimeRecorder
from app.models.tunjanganModel import MfTunjangan
from app.models.potModel import MfPot
from app.models.classModel import MfClass
from app.models.jabatanModel import MfJabatan
from app.utils.pegawaiHelper import search_operational_pegawai
from app.models.eselonModel import MfEselon
from app.models.golonganModel import MfGolongan
from app.models.lemburModel import Lembur
from app.models.logActivityModel import LogActivity
from app.utils.pegawaiHelper import is_pegawai_aktif_periode
from app.utils.rekapAbsensiHelper import generate_rekap_absensi_all_data
from app.utils.rekapAbsensiMatrixHelper import generate_rekap_absensi_matrix

from app.utils.rekapAbsensiReportHelper import (
    generate_rekap_absensi_report
)

from app.utils.pegawaiSortHelper import sort_pegawai_rows

from app.utils.absensiNormalisasiHelper import (
    merge_absensi_dinas_luar
)

def laporan_cetak_daftar_lembur_umum():
    """Render halaman Cetak Daftar Lembur Umum."""
    unit_kerja_list = MfUnitKerja.query.order_by(
        MfUnitKerja.URUT_REPORT.asc(),
        MfUnitKerja.NAMA_UNIT_KERJA.asc()
    ).all()
    return render_template(
        'pages/dashboard_1/Laporan Cetak Daftar Lembur Umum.html',
        unit_kerja_list=unit_kerja_list
    )


def export_rekap_daftar_lembur_umum():
    """Export Rekap Daftar Lembur Umum (Tab 1) - matriks + perhitungan uang."""
    unit_list = request.form.getlist('unit_kerja[]')
    bulan_str = request.form.get('bulan', '')
    bendahara = request.form.get('bendahara', '')
    pppk = request.form.get('pppk', '')
    pembuat = request.form.get('pembuat', '')
    
    if not unit_list or not bulan_str:
        return {'error': 'Unit atau bulan kosong'}, 400
    
    try:
        unit_ids = [int(u) for u in unit_list]
    except ValueError:
        return {'error': 'Unit Kerja ID harus berupa angka'}, 400
    
    tahun, bulan = map(int, bulan_str.split('-'))
    tgl_awal = datetime(tahun, bulan, 1)
    if bulan == 12:
        tgl_akhir = datetime(tahun + 1, 1, 1) - timedelta(days=1)
    else:
        tgl_akhir = datetime(tahun, bulan + 1, 1) - timedelta(days=1)
    
    # Ambil data lembur join pegawai via NIP
    rows = (
        db.session.query(Lembur, Pegawai)
        .join(Pegawai, Lembur.FINGER_ID == Pegawai.FINGER_ID)
        .filter(Lembur.TGL_KERJA.between(tgl_awal, tgl_akhir))
        .filter(Pegawai.UNIT_KERJA_ID.in_(unit_ids))
        .order_by(Pegawai.NAMA, Lembur.TGL_KERJA)
        .all()
    )
    
    if not rows:
        return {'error': 'Data tidak ada'}, 400
    
    # Ambil pegawai distinct
    pegawai_list = list(set(p for _, p in rows))
    pegawai_list.sort(key=lambda x: x.NAMA)
    
    # Ambil kalender
    kalender_rows = (
        MfKalender.query
        .filter(MfKalender.TGL_KERJA.between(tgl_awal, tgl_akhir))
        .order_by(MfKalender.TGL_KERJA.asc())
        .all()
    )
    
    # Ambil tunjangan
    tunjangan_list = (
        MfTunjangan.query
        .filter(MfTunjangan.ACTIVITY == 'Piket siaga')
        .filter(MfTunjangan.JENIS_TUNJANGAN.in_(['U.Makan', 'U.Lembur']))
        .filter(MfTunjangan.TGL_MULAI <= tgl_akhir.date())
        .all()
    )
    
    n_tgl = (tgl_akhir - tgl_awal).days + 1
    
    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Lembur"
    ws.sheet_properties.tabColor = "FF7B00"
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    coltgl = 5 + n_tgl - 1
    col_total = coltgl + 9
    
    # Logo
    try:
        img = XLImage('static/img/LogoSAR.png')
        img.width, img.height = 50, 50
        ws.add_image(img, 'A1')
    except:
        pass
    
    thin = Side(style='thin')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    # Judul
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=col_total)
    ws.cell(row=2, column=4, value='DAFTAR LEMBUR UMUM').font = Font(bold=True, size=14)
    ws.cell(row=2, column=4).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=col_total)
    ws.cell(row=3, column=4, value=f"Periode {tgl_awal:%d.%m.%Y} s/d {tgl_akhir:%d.%m.%Y}")
    ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
    
    # Header
    ws.merge_cells('B5:B7')
    ws.cell(row=5, column=2, value='No').border = border
    ws.cell(row=5, column=2).alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('C5:C7')
    ws.cell(row=5, column=3, value='Nama').border = border
    ws.cell(row=5, column=3).alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('D5:D7')
    ws.cell(row=5, column=4, value='Gol').border = border
    ws.cell(row=5, column=4).alignment = Alignment(horizontal='center', vertical='center')
    
    # Header: Jumlah Jam Kegiatan Lembur Pada Tanggal
    ws.merge_cells(start_row=5, start_column=5, end_row=5, end_column=coltgl)
    ws.cell(row=5, column=5, value='Jumlah Jam Kegiatan Lembur Pada Tanggal').border = border
    ws.cell(row=5, column=5).alignment = Alignment(horizontal='center')
    
    # Isi tanggal
    for i, d in enumerate(range(n_tgl)):
        col = 5 + i
        hari = (tgl_awal + timedelta(days=d))
        cell = ws.cell(row=6, column=col, value=hari.day)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        if hari.weekday() >= 5:
            cell.font = Font(color='FF0000')
    
    # Header: Jumlah Jam, Makan, Uang
    for col, val in {
        coltgl+1: 'Jumlah Jam\nHari Kerja', coltgl+2: 'Jumlah Jam\nHari Libur',
        coltgl+3: 'Jumlah\nMakan Libur', coltgl+4: 'Lembur', coltgl+5: 'Makan Libur',
        coltgl+6: 'Jumlah Dari\nKolom', coltgl+7: 'Potongan\nPPH', coltgl+8: 'Jumlah\nBersih', coltgl+9: 'Tanda\nTangan'
    }.items():
        ws.merge_cells(start_row=5, start_column=col, end_row=7, end_column=col)
        ws.cell(row=5, column=col, value=val).border = border
        ws.cell(row=5, column=col).font = Font(bold=True)
        ws.cell(row=5, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Lebar kolom
    ws.column_dimensions['B'].width = 5

    # ============================================================
    # AUTO WIDTH KOLOM NAMA
    # Mengikuti panjang nama pegawai.
    # Maksimum dibatasi agar Excel tetap proporsional.
    # ============================================================

    max_nama = max(
        [
            len(str(p.NAMA or ''))
            for p in pegawai_list
        ],
        default=20
    )

    ws.column_dimensions['C'].width = min(
        max(max_nama + 5, 30),
        45
    )
    ws.column_dimensions['D'].width = 5
    for i in range(5, coltgl + 1):
        col_letter = chr(64 + i) if i <= 26 else 'A'
        ws.column_dimensions[col_letter].width = 5
    for i in range(coltgl + 1, col_total + 1):
        col_letter = chr(64 + i) if i <= 26 else 'A'
        ws.column_dimensions[col_letter].width = 12
    
    # Isi data
    row = 8
    no = 1
    
    for peg in pegawai_list:
        lembur_peg = [(l, p) for l, p in rows if p.NIP == peg.NIP]
        
        # Merge 2 baris per pegawai
        for r in [row, row + 1]:
            for c in range(2, col_total + 1):
                ws.cell(row=r, column=c).border = border
        
        ws.cell(row=row, column=2, value=no)
        ws.cell(row=row, column=3, value=peg.NAMA)
        ws.cell(row=row + 1, column=3, value=peg.NIP)
        ws.cell(row=row, column=4, value=peg.GOL_ID)
        
        # Isi jam lembur per tanggal
        tot_hk = tot_hl = tot_makan = 0
        for i, d in enumerate(range(n_tgl)):
            tgl_iter = tgl_awal + timedelta(days=d)
            is_libur = tgl_iter.weekday() >= 5
            
            jam_lembur = 0
            for l, _ in lembur_peg:
                if l.TGL_KERJA and l.TGL_KERJA.date() == tgl_iter.date():
                    jam_in = l.JAM_BAKU_IN if l.JAM_BAKU_IN else l.JAM_IN
                    jam_out = l.JAM_BAKU_OUT if l.JAM_BAKU_OUT else l.JAM_OUT
                    
                    if jam_in and jam_out:
                        delta = (jam_out - jam_in).total_seconds() / 3600
                        if delta > 0:
                            jam_lembur = delta
                    break
            
            if jam_lembur > 0:
                col = 5 + i
                ws.cell(row=row, column=col, value=round(jam_lembur, 1))
            
            if is_libur:
                tot_hl += jam_lembur
            else:
                tot_hk += jam_lembur
            
            if jam_lembur > 0:
                tot_makan += 1
        
        ws.cell(row=row, column=coltgl + 1, value=round(tot_hk, 1))
        ws.cell(row=row, column=coltgl + 2, value=round(tot_hl, 1))
        ws.cell(row=row, column=coltgl + 3, value=tot_makan)
        
        # Hitung uang
        uang_lembur = 0
        uang_makan = 0
        for t in tunjangan_list:
            if t.JENIS_TUNJANGAN == 'U.Lembur':
                uang_lembur = t.NOMINAL or 0
            elif t.JENIS_TUNJANGAN == 'U.Makan':
                uang_makan = t.NOMINAL or 0
        
        total_lembur = (tot_hk + tot_hl) * uang_lembur
        total_makan = tot_makan * uang_makan
        total_bersih = total_lembur + total_makan
        
        ws.cell(row=row, column=coltgl + 4, value=total_lembur).number_format = '#,##0'
        ws.cell(row=row, column=coltgl + 5, value=total_makan).number_format = '#,##0'
        ws.cell(row=row, column=coltgl + 6, value=total_bersih).number_format = '#,##0'
        ws.cell(row=row, column=coltgl + 7, value=0).number_format = '#,##0'
        ws.cell(row=row, column=coltgl + 8, value=total_bersih).number_format = '#,##0'
        ws.cell(row=row, column=coltgl + 9, value=no)
        
        no += 1
        row += 2
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
        download_name=f"Daftar_Lembur_{tgl_awal:%Y%m}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def export_detail_jam_lembur_umum():
    """Export Detail Jam Lembur Umum (Tab 2) - jam in/out per hari."""
    unit_list = request.form.getlist('unit_kerja[]')
    bulan_str = request.form.get('bulan', '')
    
    if not unit_list or not bulan_str:
        return {'error': 'Unit atau bulan kosong'}, 400
    
    try:
        unit_ids = [int(u) for u in unit_list]
    except ValueError:
        return {'error': 'Unit Kerja ID harus berupa angka'}, 400
    
    tahun, bulan = map(int, bulan_str.split('-'))
    tgl_awal = datetime(tahun, bulan, 1)
    if bulan == 12:
        tgl_akhir = datetime(tahun + 1, 1, 1) - timedelta(days=1)
    else:
        tgl_akhir = datetime(tahun, bulan + 1, 1) - timedelta(days=1)
    
    rows = (
        db.session.query(Lembur, Pegawai)
        .join(Pegawai, Lembur.FINGER_ID == Pegawai.FINGER_ID)
        .filter(Lembur.TGL_KERJA.between(tgl_awal, tgl_akhir))
        .filter(Pegawai.UNIT_KERJA_ID.in_(unit_ids))
        .order_by(Pegawai.NAMA, Lembur.TGL_KERJA)
        .all()
    )
    
    if not rows:
        return {'error': 'Data tidak ada'}, 400
    
    n_tgl = (tgl_akhir - tgl_awal).days + 1
    pegawai_list = sorted(set(p for _, p in rows), key=lambda x: x.NAMA)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Detail Lembur"
    ws.sheet_properties.tabColor = "FF7B00"
    ws.page_setup.orientation = 'landscape'
    
    try:
        img = XLImage('static/img/LogoSAR.png')
        img.width, img.height = 50, 50
        ws.add_image(img, 'A1')
    except:
        pass
    
    thin = Side(style='thin')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    col_total = 5 + n_tgl - 1
    
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=col_total)
    ws.cell(row=2, column=4, value='DAFTAR DETAIL LEMBUR UMUM').font = Font(bold=True, size=12)
    ws.cell(row=2, column=4).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=col_total)
    ws.cell(row=3, column=4, value=f"Periode {tgl_awal:%d.%m.%Y} s/d {tgl_akhir:%d.%m.%Y}")
    ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
    
    # Header
    for col, val in {2: 'No', 3: 'Nama', 4: 'Gol', 5: 'Jumlah Jam Kegiatan Lembur Pada Tanggal'}.items():
        if col == 5:
            ws.merge_cells(start_row=5, start_column=5, end_row=5, end_column=col_total)
        ws.cell(row=5, column=col, value=val).border = border
        ws.cell(row=5, column=col).font = Font(bold=True)
        ws.cell(row=5, column=col).alignment = Alignment(horizontal='center', vertical='center')
    
    for i, d in enumerate(range(n_tgl)):
        col = 5 + i
        ws.cell(row=6, column=col, value=(tgl_awal + timedelta(days=d)).day).border = border
        ws.cell(row=6, column=col).alignment = Alignment(horizontal='center')
    
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 5
    for i in range(5, col_total + 1):
        col_letter = chr(64 + i) if i <= 26 else 'A'
        ws.column_dimensions[col_letter].width = 8
    
    row = 7
    no = 1
    for peg in pegawai_list:
        for c in range(2, col_total + 1):
            ws.cell(row=row, column=c).border = border
        
        ws.cell(row=row, column=2, value=no)
        ws.cell(row=row, column=3, value=f"{peg.NAMA}\n{peg.NIP}")
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=4, value=peg.GOL_ID)
        
        lembur_peg = [(l, p) for l, p in rows if p.NIP == peg.NIP]
        for i, d in enumerate(range(n_tgl)):
            tgl_iter = tgl_awal + timedelta(days=d)
            col = 5 + i
            
            for l, _ in lembur_peg:
                if l.TGL_KERJA and l.TGL_KERJA.date() == tgl_iter.date():
                    jam_in = l.JAM_IN.strftime('%H:%M') if l.JAM_IN else '-'
                    jam_out = l.JAM_OUT.strftime('%H:%M') if l.JAM_OUT else '-'
                    ws.cell(row=row, column=col, value=f"{jam_in}-{jam_out}")
                    break
        
        no += 1
        row += 1
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
        download_name=f"Detail_Lembur_{tgl_awal:%Y%m}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def laporan_rekap_absensi_all():
    """
    Render halaman Laporan Rekap Absensi All.
    Unit Kerja dropdown diisi dari tabel MF_UNIT_KERJA (server-side render),
    bukan hardcode di HTML.
    """
    unit_kerja_list = MfUnitKerja.query.order_by(
        MfUnitKerja.URUT_REPORT.asc(),
        MfUnitKerja.NAMA_UNIT_KERJA.asc()
    ).all()

    return render_template(
        'pages/dashboard_1/Laporan Rekap Absensi All.html',
        unit_kerja_list=unit_kerja_list
    )


def preview_rekap_absensi_all():
    """
    Preview Rekap Absensi All.

    Sumber data:
    generate_rekap_absensi_all_data()

    Dipakai oleh:
    VIEW DATA
    """

    unit_list = request.form.getlist(
        'unit_kerja[]'
    )

    tgl_awal_str = request.form.get(
        'tgl_awal'
    )

    tgl_akhir_str = request.form.get(
        'tgl_akhir'
    )


    if not unit_list or not tgl_awal_str or not tgl_akhir_str:
        return {
            'error': 'Unit atau periode belum lengkap'
        }, 400


    try:
        unit_ids = [
            int(x)
            for x in unit_list
        ]

        tgl_awal = datetime.strptime(
            tgl_awal_str,
            '%Y-%m-%d'
        )

        tgl_akhir = datetime.strptime(
            tgl_akhir_str,
            '%Y-%m-%d'
        )

    except Exception as e:
        return {
            'error': str(e)
        },400


    data = generate_rekap_absensi_all_data(
        unit_ids,
        tgl_awal,
        tgl_akhir
    )


    report = generate_rekap_absensi_report(
        data
    )


    return {
        'success': True,
        'periode': {
            'awal': tgl_awal.strftime('%Y-%m-%d'),
            'akhir': tgl_akhir.strftime('%Y-%m-%d')
        },
        'total_kalender': len(
            data['kalender']
        ),
        'total_pegawai': len(
            data['pegawai']
        ),
        'total_absensi': len(
            data['absensi']
        ),

        'total_report': len(
            report
        ),

        'report': report,

        'pegawai': [
            {
                'nama': p.NAMA,
                'nip': p.NIP,
                'eselon': p.ESELON,
                'class_id': p.CLASS_ID
            }
            for p in data['pegawai'][:100]
        ]
    }



def export_rekap_absensi_all():
    unit_list = request.form.getlist('unit_kerja[]')
    tgl_awal_str = request.form.get('tgl_awal')
    tgl_akhir_str = request.form.get('tgl_akhir')
    
    if not unit_list or not tgl_awal_str or not tgl_akhir_str:
        return {'error': 'Unit kosong atau format tanggal salah'}, 400
    
    # Konversi unit_list ke integer
    try:
        unit_ids = [int(u) for u in unit_list]
    except ValueError:
        return {'error': 'Unit Kerja ID harus berupa angka'}, 400
    
    tgl_awal = datetime.strptime(tgl_awal_str, '%Y-%m-%d')
    tgl_akhir = datetime.strptime(tgl_akhir_str, '%Y-%m-%d')
    tampilkan_ket = request.form.get('kolom_keterangan') == 'tampilkan'
    nama_eselon3 = request.form.get('nama_eselon3', '')
    pangkat_eselon3 = request.form.get('pangkat_eselon3', '')
    petugas1 = request.form.get('petugas1', '')
    petugas2 = request.form.get('petugas2', '')

    # Cek tanggal server
    tgl_server = datetime.now()
    if tgl_server.date() < tgl_awal.date():
        return {'error': 'Tgl server lebih kecil dari tanggal awal periode'}, 400
    if tgl_server.date() < tgl_akhir.date():
        tgl_akhir = tgl_server

    # ================================================================
    # HRIS 2013 BUSINESS FLOW
    # ================================================================
    # KALENDER = sumber penentu hari kerja/libur.
    # PEGAWAI  = sumber daftar pegawai yang harus muncul di laporan.
    # ABSENSI  = sumber transaksi aktual.
    #
    # ABSENSI -> PEGAWAI menggunakan FingerID.
    #
    # Penting:
    # Jangan menjadikan ABSENSI sebagai sumber daftar pegawai.
    # Pegawai tanpa record ABSENSI tetap harus dapat muncul di laporan.
    # ================================================================

    kalender_rows = (
        MfKalender.query
        .filter(MfKalender.TGL_KERJA >= tgl_awal)
        .filter(MfKalender.TGL_KERJA <= tgl_akhir)
        .order_by(MfKalender.TGL_KERJA.asc())
        .all()
    )

    if not kalender_rows:
        return {'error': 'Data KALENDER tidak tersedia untuk periode tersebut'}, 400

    hari_kerja = {
        k.TGL_KERJA.date()
        for k in kalender_rows
        if (k.IS_LIBUR or 'N').upper() != 'Y'
    }

    if not hari_kerja:
        return {'error': 'Tidak ada hari kerja dalam periode tersebut'}, 400

    # ------------------------------------------------
    # 1. Ambil PEGAWAI aktif pada periode laporan.
    # ------------------------------------------------
    pegawai_rows = (
        Pegawai.query
        .filter(Pegawai.UNIT_KERJA_ID.in_(unit_ids))
        .filter(Pegawai.TGL_MASUK <= tgl_akhir)
        .filter(
            db.or_(
                Pegawai.IS_KELUAR == 'N',
                db.and_(
                    Pegawai.IS_KELUAR == 'Y',
                    Pegawai.TGL_KELUAR >= tgl_awal
                )
            )
        )
        .order_by(Pegawai.NAMA)
        .all()
    )

    if not pegawai_rows:
        return {'error': 'Tidak ada pegawai aktif pada unit/periode tersebut'}, 400

    # ------------------------------------------------
    # 2. Ambil ABSENSI aktual.
    #
    # Connector resmi:
    # ABSENSI.FINGER_ID -> PEGAWAI.FINGER_ID
    # ------------------------------------------------
    q = (
        db.session.query(Absensi, Pegawai)
        .join(Pegawai, Absensi.FINGER_ID == Pegawai.FINGER_ID)
        .filter(Absensi.TGL_KERJA >= tgl_awal)
        .filter(Absensi.TGL_KERJA <= tgl_akhir)
        .filter(Pegawai.UNIT_KERJA_ID.in_(unit_ids))
    )

    rows = q.all()

    # ------------------------------------------------
    # SIAGA diproses oleh:
    #
    # generate_rekap_absensi_all_data()
    # +
    # generate_rekap_absensi_report()
    #
    # Jangan query langsung di sini.
    # ------------------------------------------------


    # ------------------------------------------------
    # 3. Index ABSENSI berdasarkan NIP.
    #
    # NIP hanya digunakan sebagai identifier laporan
    # setelah connector FingerID berhasil dilakukan.
    # ------------------------------------------------
    absensi_by_nip = {}

    for absensi, pegawai in rows:
        nip = pegawai.NIP

        if nip not in absensi_by_nip:
            absensi_by_nip[nip] = []

        absensi_by_nip[nip].append(absensi)

    # ------------------------------------------------
    # 4. Gunakan NORMALISASI ABSENSI HRIS Reborn.
    #
    # Sumber tunggal:
    #
    # PEGAWAI
    # ABSENSI
    # DINAS_LUAR
    # SIAGA
    #
    # VIEW / EXCEL / PDF harus memakai sumber sama.
    # ------------------------------------------------


    data_normalisasi = generate_rekap_absensi_all_data(
        unit_ids,
        tgl_awal,
        tgl_akhir
    )


    hasil = generate_rekap_absensi_report(
        data_normalisasi
    )



    df_hasil = pd.DataFrame(hasil)

    # 4. (Opsional) keterangan dinas luar/cuti kalau tampilkan_ket True
    if tampilkan_ket:
        dl_rows = (
            DinasLuar.query
            .filter(DinasLuar.TGL_AWAL_DINAS_LUAR <= tgl_akhir)
            .filter(DinasLuar.TGL_AKHIR_DINAS_LUAR >= tgl_awal)
            .all()
        )
        # susun jadi dict {nip: "1. ... \n2. ..."} sesuai format lama

    # 5. Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar"
    ws.sheet_properties.tabColor = "FF7B00"

    try:
        img = XLImage('static/img/LogoSAR.png')
        img.width, img.height = 50, 50
        ws.add_image(img, 'A1')
    except FileNotFoundError:
        pass

    ws.merge_cells('D2:AL2')
    ws['D2'] = 'Laporan Rekap Daftar Hadir Pegawai'
    ws['D2'].font = Font(bold=True)
    ws['D2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('D3:AL3')
    ws['D3'] = f"Periode {tgl_awal:%d.%m.%Y} s/d {tgl_akhir:%d.%m.%Y}"
    ws['D3'].alignment = Alignment(horizontal='center')

    ws.merge_cells('D4:AL4')
    ws['D4'] = f"Unit : {', '.join(unit_list)}"
    ws['D4'].alignment = Alignment(horizontal='center')

    thin = Side(style='thin')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    header_row = 5
    headers = ['No', 'Nama', 'TLM1', 'TLM2', 'TLM3', 'TLM4', 'Cuti', 'Sakit', 'Alpa', 'Siaga']  # lengkapi sesuai kebutuhan
    for col, h in enumerate(headers, start=2):
        c = ws.cell(row=header_row, column=col, value=h)
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')

    row = header_row + 1
    for i, r in enumerate(df_hasil.to_dict('records'), start=1):
        ws.cell(row=row, column=2, value=i).border = border
        ws.cell(row=row, column=3, value=f"{r['nip']}\n{r['nama']}").border = border
        ws.cell(row=row, column=4, value=r['tlm1'] or None).border = border
        ws.cell(row=row, column=5, value=r['tlm2'] or None).border = border
        ws.cell(row=row, column=6, value=r['tlm3'] or None).border = border
        ws.cell(row=row, column=7, value=r['tlm4'] or None).border = border
        ws.cell(row=row, column=8, value=r['cuti'] or None).border = border
        ws.cell(row=row, column=9, value=r['sakit'] or None).border = border
        ws.cell(row=row, column=10, value=r['alpa'] or None).border = border
        ws.cell(row=row, column=11, value=r['siaga'] or None).border = border
        row += 1

    row += 2
    ws.cell(row=row, column=4, value='Mengetahui,')
    ws.cell(row=row+1, column=4, value='Pejabat Eselon III')
    ws.cell(row=row+4, column=4, value=nama_eselon3).font = Font(underline='single')
    ws.cell(row=row+5, column=4, value=pangkat_eselon3)

    ws.cell(row=row, column=32, value=f"Surabaya, {tgl_akhir:%d %B %Y}")
    ws.cell(row=row+1, column=32, value='Petugas Pengelola Daftar Hadir')
    ws.cell(row=row+3, column=32, value=f"1. {petugas1}")
    ws.cell(row=row+5, column=32, value=f"2. {petugas2}")

    # 6. Stream sebagai download, bukan simpan ke disk
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Laporan_Rekap_Daftar_Hadir_Peg_{tgl_awal:%Y%m%d}_{tgl_akhir:%Y%m%d}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )



def export_rekap_absensi_all_pdf():
    """
    Export Rekap Absensi All ke PDF.

    Sumber data tunggal:

    PEGAWAI
    ABSENSI
    DINAS_LUAR
    SIAGA

    melalui:

    generate_rekap_absensi_all_data()
    generate_rekap_absensi_report()
    """


    unit_list = request.form.getlist(
        'unit_kerja[]'
    )

    tgl_awal_str = request.form.get(
        'tgl_awal'
    )

    tgl_akhir_str = request.form.get(
        'tgl_akhir'
    )


    if not unit_list or not tgl_awal_str or not tgl_akhir_str:

        return {
            'error':
                'Unit atau tanggal kosong'
        },400



    unit_ids = [
        int(x)
        for x in unit_list
    ]


    tgl_awal = datetime.strptime(
        tgl_awal_str,
        '%Y-%m-%d'
    )


    tgl_akhir = datetime.strptime(
        tgl_akhir_str,
        '%Y-%m-%d'
    )



    data = generate_rekap_absensi_all_data(
        unit_ids,
        tgl_awal,
        tgl_akhir
    )


    report = generate_rekap_absensi_report(
        data
    )



    buffer = BytesIO()



    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=25,
        leftMargin=25,
        topMargin=30,
        bottomMargin=30
    )



    elements = []


    styles = getSampleStyleSheet()


    title_style = ParagraphStyle(
        'RekapAbsensiTitle',
        parent=styles['Heading1'],
        fontSize=14,
        alignment=1,
        spaceAfter=15
    )


    elements.append(
        Paragraph(
            "Laporan Rekap Absensi Pegawai",
            title_style
        )
    )


    elements.append(
        Paragraph(
            f"Periode {tgl_awal:%d-%m-%Y} s/d {tgl_akhir:%d-%m-%Y}",
            styles['Normal']
        )
    )


    elements.append(
        Spacer(
            1,
            15
        )
    )



    table_data = [

        [
            'No',
            'NIP',
            'Nama',
            'TLM1',
            'TLM2',
            'PSW1',
            'PSW2',
            'Cuti',
            'Sakit',
            'Alpa',
            'DL',
            'DL OP',
            'DL SD',
            'Siaga'
        ]

    ]



    for no, row in enumerate(
        report,
        start=1
    ):

        table_data.append(

            [

                str(no),

                row.get('nip',''),

                row.get('nama',''),

                str(row.get('tlm1',0)),

                str(row.get('tlm2',0)),

                str(row.get('psw1',0)),

                str(row.get('psw2',0)),

                str(row.get('cuti',0)),

                str(row.get('sakit',0)),

                str(row.get('alpa',0)),

                str(row.get('dl',0)),

                str(row.get('dl_op',0)),

                str(row.get('dl_sd',0)),

                str(row.get('siaga',0))

            ]

        )




    table = Table(
        table_data,
        repeatRows=1
    )


    table.setStyle(

        TableStyle(

            [

                (
                    'BACKGROUND',
                    (0,0),
                    (-1,0),
                    colors.HexColor('#EB6831')
                ),

                (
                    'TEXTCOLOR',
                    (0,0),
                    (-1,0),
                    colors.white
                ),

                (
                    'FONTNAME',
                    (0,0),
                    (-1,0),
                    'Helvetica-Bold'
                ),

                (
                    'FONTSIZE',
                    (0,0),
                    (-1,-1),
                    7
                ),

                (
                    'GRID',
                    (0,0),
                    (-1,-1),
                    0.5,
                    colors.grey
                ),

                (
                    'VALIGN',
                    (0,0),
                    (-1,-1),
                    'MIDDLE'
                )

            ]

        )

    )


    elements.append(
        table
    )


    doc.build(
        elements
    )


    buffer.seek(0)



    filename = (
        f"Rekap_Absensi_"
        f"{tgl_awal:%Y%m%d}_"
        f"{tgl_akhir:%Y%m%d}.pdf"
    )



    return send_file(

        buffer,

        mimetype='application/pdf',

        as_attachment=True,

        download_name=filename

    )




def laporan_rekap_absensi_individu():
    """
    Render halaman Laporan Rekap Absensi Individu.
    """
    return render_template('pages/dashboard_1/Laporan Rekap Absensi Individu.html')

def export_rekap_absensi_individu():
    """
    Export Laporan Rekap Absensi Individu (per pegawai, detail per hari).
    Mirip dengan FillRekapAbsensiPerson di VB.NET.
    """
    nip_list = request.form.getlist('nip[]')  # ['198501232009122002', ...]
    tgl_awal_str = request.form.get('tgl_awal')
    tgl_akhir_str = request.form.get('tgl_akhir')
    
    if not nip_list or not tgl_awal_str or not tgl_akhir_str:
        return {'error': 'NIP atau tanggal kosong'}, 400
    
    tgl_awal = datetime.strptime(tgl_awal_str, '%Y-%m-%d')
    tgl_akhir = datetime.strptime(tgl_akhir_str, '%Y-%m-%d')
    
    # Cek tanggal server
    tgl_server = datetime.now()
    if tgl_server.date() < tgl_awal.date():
        return {'error': 'Tgl server lebih kecil dari tanggal awal periode'}, 400
    if tgl_server.date() < tgl_akhir.date():
        tgl_akhir = tgl_server
    
    # 1. Ambil data kalender (hari kerja saja)
    # HRIS 2013:
    # KALENDER adalah sumber penentu hari kerja/libur.
    # Jangan menebak hari kerja dari weekday() jika data kalender tersedia.
    kalender_rows = (
        MfKalender.query
        .filter(MfKalender.TGL_KERJA >= tgl_awal)
        .filter(MfKalender.TGL_KERJA <= tgl_akhir)
        .order_by(MfKalender.TGL_KERJA.asc())
        .all()
    )

    if not kalender_rows:
        return {'error': 'Data KALENDER tidak tersedia untuk periode tersebut'}, 400

    kalender_hari_kerja = [
        k for k in kalender_rows
        if (k.IS_LIBUR or 'N').upper() != 'Y'
    ]
    
    if not kalender_hari_kerja:
        return {'error': 'Tidak ada hari kerja dalam periode tersebut'}, 400
    
    # 2. Ambil data absensi untuk NIP yang dipilih.
    # HRIS 2013 menghubungkan ABSENSI ke PEGAWAI melalui FingerID.
    absensi_rows = (
        db.session.query(Absensi, Pegawai)
        .join(Pegawai, Absensi.FINGER_ID == Pegawai.FINGER_ID)
        .filter(Absensi.TGL_KERJA.between(tgl_awal, tgl_akhir))
        .filter(Pegawai.NIP.in_(nip_list))
        .order_by(Pegawai.NAMA, Absensi.TGL_KERJA)
        .all()
    )
    
    # 3. Ambil data pegawai
    pegawai_rows = (
        Pegawai.query
        .filter(Pegawai.NIP.in_(nip_list))
        .order_by(Pegawai.NAMA)
        .all()
    )
    
    if not pegawai_rows:
        return {'error': 'Pegawai tidak ditemukan'}, 400
    
    # 4. Build data per pegawai per tanggal
    # Struktur: {nip: {nama: ..., unit_kerja: ..., rows: [{tgl, tlm, kategori_tlm, psw, kategori_psw, cuti, dl, sakit, sakit_a, alpa, alpa_a, ket}]}}
    absensi_dict = {}
    for a, p in absensi_rows:
        tgl_key = a.TGL_KERJA.strftime('%Y-%m-%d') if a.TGL_KERJA else None
        if p.NIP not in absensi_dict:
            absensi_dict[p.NIP] = {}
        absensi_dict[p.NIP][tgl_key] = a
    
    # 5. Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Individu"
    ws.sheet_properties.tabColor = "FF7B00"
    
    # Setup printer
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    # Header
    ws.merge_cells('D2:N2')
    ws['D2'] = 'Laporan Rekap Daftar Hadir Pegawai'
    ws['D2'].font = Font(bold=True, size=12)
    ws['D2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('D3:N3')
    ws['D3'] = f"Periode {tgl_awal:%d.%m.%Y} s/d {tgl_akhir:%d.%m.%Y}"
    ws['D3'].alignment = Alignment(horizontal='center')
    
    thin = Side(style='thin')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    # Kolom header
    headers = ['No.', 'Tanggal', 'TLM\n(menit)', 'Kategori\nTLM', 'PSW\n(menit)', 
               'Kategori\nPSW', 'Cuti\n(hari)', 'Dinas\nLuar', 'Sakit\nDokter', 
               'Sakit\ntnp dr', 'Tdk Hadir\ndgn Izin', 'Tdk Hadir\nTanpa Ket', 'Keterangan']
    
    header_row = 5
    for col, h in enumerate(headers, start=2):
        c = ws.cell(row=header_row, column=col, value=h)
        c.border = border
        c.font = Font(bold=True, size=9)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Set lebar kolom
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 9
    ws.column_dimensions['E'].width = 9
    ws.column_dimensions['F'].width = 9
    ws.column_dimensions['G'].width = 9
    ws.column_dimensions['H'].width = 9
    ws.column_dimensions['I'].width = 9
    ws.column_dimensions['J'].width = 9
    ws.column_dimensions['K'].width = 9
    ws.column_dimensions['L'].width = 9
    ws.column_dimensions['M'].width = 9
    ws.column_dimensions['N'].width = 25
    
    row = header_row + 1
    
    for pegawai in pegawai_rows:
        # Baris nama pegawai (merge)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
        ws.cell(row=row, column=2, value=pegawai.NAMA).font = Font(bold=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')
        for col in range(2, 15):
            ws.cell(row=row, column=col).border = border
        row += 1
        
        # Detail per hari
        no = 0
        for kl in kalender_hari_kerja:
            no += 1
            tgl_str = kl.TGL_KERJA.strftime('%Y-%m-%d') if kl.TGL_KERJA else ''
            
            # Border
            for col in range(2, 15):
                ws.cell(row=row, column=col).border = border
            
            ws.cell(row=row, column=2, value=no).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=row, column=3, value=kl.TGL_KERJA.strftime('%d-%m-%Y') if kl.TGL_KERJA else '').alignment = Alignment(horizontal='left', vertical='center')
            
            # Cek absensi untuk tanggal ini
            absensi = absensi_dict.get(pegawai.NIP, {}).get(tgl_str)
            
            if absensi:
                ws.cell(row=row, column=4, value=absensi.AWAL_TLM or 0).alignment = Alignment(horizontal='right', vertical='center')
                ws.cell(row=row, column=5, value=absensi.TINGKAT_TLM or '').alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=6, value=absensi.TOTAL_PSW or 0).alignment = Alignment(horizontal='right', vertical='center')
                ws.cell(row=row, column=7, value=absensi.TINGKAT_PSW or '').alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=14, value=absensi.KET_IN or '').alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Kategori transaksi
                transaksi = (absensi.TRANSAKSI_IN or '').upper()
                pendukung = (absensi.PENDUKUNG_IN or '').upper()
                
                if transaksi == 'CUTI':
                    ws.cell(row=row, column=8, value=1)
                elif transaksi == 'DINASLUAR':
                    ws.cell(row=row, column=9, value=1)
                elif transaksi == 'SAKIT':
                    if pendukung == 'Y':
                        ws.cell(row=row, column=10, value=1)
                    else:
                        ws.cell(row=row, column=11, value=1)
                elif transaksi == 'ALPA':
                    if pendukung == 'Y':
                        ws.cell(row=row, column=12, value=1)
                    else:
                        ws.cell(row=row, column=13, value=1)
                elif transaksi == 'IJIN':
                    if pendukung == 'Y':
                        ws.cell(row=row, column=12, value=1)
                    else:
                        ws.cell(row=row, column=13, value=1)
            else:
                # Tidak ada record = Alpa tanpa keterangan
                ws.cell(row=row, column=13, value=1)
            
            # Alignment untuk kolom angka
            for col in [8, 9, 10, 11, 12, 13]:
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right', vertical='center')
            
            row += 1
        
        row += 1  # Spasi antar pegawai
    
    # Keterangan di bawah
    row += 1
    ws.cell(row=row, column=2, value='Keterangan:').font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=2, value='TLM')
    ws.cell(row=row, column=4, value=': Terlambat Masuk')
    row += 1
    ws.cell(row=row, column=2, value='PSW')
    ws.cell(row=row, column=4, value=': Pulang Sebelum Waktu')
    row += 1
    ws.cell(row=row, column=2, value='TLM (-)')
    ws.cell(row=row, column=4, value=': Datang Lebih Awal')
    
    # Stream download
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Laporan_Rekap_Daftar_Hadir_Per_Pegawai_{tgl_awal:%Y%m%d}_{tgl_akhir:%Y%m%d}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def search_pegawai_by_name():
    """
    API untuk search pegawai berdasarkan nama
    untuk dropdown autocomplete.

    Standar HRIS Reborn:

        - Minimal 1 karakter
        - Hanya Pegawai Operasional
        - IS_KELUAR = N
        - Unit Kerja IS_USE = Y
        - Maksimal 15 kandidat
        - Nama jabatan berasal dari MF_JABATAN
    """

    keyword = request.args.get('keyword', '').strip()

    if not keyword:
        return {'data': []}

    # ============================================================
    # AUTOCOMPLETE PEGAWAI TERPUSAT
    #
    # Business Rule berada di:
    #
    #   app/utils/pegawaiHelper.py
    #
    # Jangan melakukan query Pegawai langsung di endpoint ini.
    # ============================================================

    pegawai_list = search_operational_pegawai(
        keyword,
        limit=15
    )

    # ============================================================
    # MASTER JABATAN
    #
    # Nama jabatan resmi berasal dari:
    #
    #   Pegawai.JABATAN_ID
    #          ↓
    #   MF_JABATAN.JABATAN_ID
    #          ↓
    #   MF_JABATAN.NAMA_JABATAN
    # ============================================================

    jabatan_ids = {
        peg.JABATAN_ID
        for peg in pegawai_list
        if peg.JABATAN_ID not in (None, 0)
    }

    jabatan_map = {}

    if jabatan_ids:
        jabatan_rows = (
            MfJabatan.query
            .filter(
                MfJabatan.JABATAN_ID.in_(jabatan_ids)
            )
            .all()
        )

        jabatan_map = {
            jab.JABATAN_ID: jab.NAMA_JABATAN
            for jab in jabatan_rows
        }

    return {
        'data': [
            {
                'nip': peg.NIP,
                'nama': peg.NAMA or '',
                'jabatan': jabatan_map.get(
                    peg.JABATAN_ID
                ),
            }
            for peg in pegawai_list
        ]
    }


def laporan_rekap_absensi_log_finger():
    """
    Render halaman Laporan Rekap Absensi Log Finger.
    Unit Kerja dropdown diisi dari tabel MF_UNIT_KERJA.
    """
    unit_kerja_list = MfUnitKerja.query.order_by(
        MfUnitKerja.URUT_REPORT.asc(),
        MfUnitKerja.NAMA_UNIT_KERJA.asc()
    ).all()

    return render_template(
        'pages/dashboard_1/Laporan Rekap Absensi Log Finger.html',
        unit_kerja_list=unit_kerja_list
    )

def laporan_rekap_absensi_log_finger():
    """
    Render halaman Laporan Rekap Absensi Log Finger.
    Unit Kerja dropdown diisi dari tabel MF_UNIT_KERJA.
    """
    unit_kerja_list = MfUnitKerja.query.order_by(
        MfUnitKerja.URUT_REPORT.asc(),
        MfUnitKerja.NAMA_UNIT_KERJA.asc()
    ).all()

    return render_template(
        'pages/dashboard_1/Laporan Rekap Absensi Log Finger.html',
        unit_kerja_list=unit_kerja_list
    )

def export_rekap_absensi_log_finger():
    """Export Laporan Log Finger via TIME_RECORDER -> ABSENSI -> PEGAWAI"""
    unit_list = request.form.getlist('unit_kerja[]')
    tgl_awal_str = request.form.get('tgl_awal')
    tgl_akhir_str = request.form.get('tgl_akhir')
    
    if not unit_list or not tgl_awal_str or not tgl_akhir_str:
        return {'error': 'Unit kosong atau format tanggal salah'}, 400
    
    try:
        unit_ids = [int(u) for u in unit_list]
    except ValueError:
        return {'error': 'Unit Kerja ID harus berupa angka'}, 400
    
    tgl_awal = datetime.strptime(tgl_awal_str, '%Y-%m-%d')
    tgl_akhir = datetime.strptime(tgl_akhir_str, '%Y-%m-%d') + timedelta(days=1)
    
    # HRIS 2013 menggunakan FingerID sebagai connector absensi.
    # TIME_RECORDER dan PEGAWAI sama-sama memiliki FingerID,
    # sehingga tidak perlu mengambil NIP dari ABSENSI.
    rows = (
        db.session.query(TimeRecorder, Pegawai, MfUnitKerja)
        .join(Pegawai, TimeRecorder.FINGER_ID == Pegawai.FINGER_ID)
        .join(MfUnitKerja, Pegawai.UNIT_KERJA_ID == MfUnitKerja.UNIT_KERJA_ID)
        .filter(TimeRecorder.WAKTU.between(tgl_awal, tgl_akhir))
        .filter(Pegawai.UNIT_KERJA_ID.in_(unit_ids))
        .order_by(Pegawai.NAMA, TimeRecorder.WAKTU)
        .all()
    )
    
    if not rows:
        return {'error': 'Record tidak ada'}, 400
    
    # Ambil nama unit untuk judul
    unit_names_list = MfUnitKerja.query.filter(MfUnitKerja.UNIT_KERJA_ID.in_(unit_ids)).all()
    unit_names = ', '.join([u.NAMA_UNIT_KERJA for u in unit_names_list])
    
    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Log Finger"
    ws.sheet_properties.tabColor = "FF7B00"
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    # Logo
    try:
        img = XLImage('static/img/LogoSAR.png')
        img.width, img.height = 50, 50
        ws.add_image(img, 'A1')
    except:
        pass
    
    # Judul
    ws.merge_cells('D2:F2')
    ws['D2'] = 'Rekap Log Finger'
    ws['D2'].font = Font(bold=True, size=12)
    ws['D2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('D3:F3')
    ws['D3'] = f"Periode {tgl_awal:%d.%m.%Y} s/d {(tgl_akhir - timedelta(days=1)):%d.%m.%Y}"
    ws['D3'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('D4:F4')
    ws['D4'] = f"Unit : {unit_names}"
    ws['D4'].alignment = Alignment(horizontal='center')
    
    thin = Side(style='thin')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    # Header kolom
    headers = ['No.', 'Tanggal', 'Jam', 'Status', 'Device']
    for col, h in enumerate(headers, start=2):
        c = ws.cell(row=5, column=col, value=h)
        c.border = border
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 30
    
    row = 6
    prev_name = ''
    no = 0
    
    for tr, pg, uk in rows:
        current_name = f"{tr.FINGER_ID} - {pg.NAMA}"
        
        if current_name != prev_name:
            # Baris nama pegawai
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            ws.cell(row=row, column=2, value=current_name)
            ws.cell(row=row, column=2).font = Font(bold=True, italic=True)
            for col in range(2, 7):
                ws.cell(row=row, column=col).border = border
            row += 1
            prev_name = current_name
            no = 0
        
        no += 1
        for col in range(2, 7):
            ws.cell(row=row, column=col).border = border
        
        ws.cell(row=row, column=2, value=no)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3, value=tr.WAKTU.strftime('%d %b %Y') if tr.WAKTU else '')
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='left')
        ws.cell(row=row, column=4, value=tr.WAKTU.strftime('%H:%M:%S') if tr.WAKTU else '')
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=tr.STATUS or '')
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=6, value=f"{tr.MESIN or '-'} @ {uk.NAMA_UNIT_KERJA or '-'}")
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='left')
        
        row += 1
    
    # Stream download
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Rekap_Log_Finger_{tgl_awal:%Y%m%d}_{(tgl_akhir - timedelta(days=1)):%Y%m%d}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def laporan_rekap_clock_exception():
    """Render halaman Laporan Rekap Clock Exception."""
    unit_kerja_list = (
        MfUnitKerja.query
        .filter(MfUnitKerja.IS_USE == 'Y')
        .order_by(
            MfUnitKerja.URUT_REPORT.asc(),
            MfUnitKerja.NAMA_UNIT_KERJA.asc()
        )
        .all()
    )
    return render_template(
        'pages/dashboard_1/Laporan Rekap Clock Exception.html',
        unit_kerja_list=unit_kerja_list
    )


def preview_rekap_clock_exception():

    unit_list = request.form.getlist(
        'unit_kerja[]'
    )

    tgl_awal_str = request.form.get(
        'tgl_awal'
    )

    tgl_akhir_str = request.form.get(
        'tgl_akhir'
    )


    if not unit_list or not tgl_awal_str or not tgl_akhir_str:

        return {
            "error": "Unit atau periode kosong"
        },400


    unit_ids = [
        int(x)
        for x in unit_list
    ]


    tgl_awal = datetime.strptime(
        tgl_awal_str,
        '%Y-%m-%d'
    )

    tgl_akhir = datetime.strptime(
        tgl_akhir_str,
        '%Y-%m-%d'
    )


    data = generate_rekap_absensi_matrix(
        unit_ids,
        tgl_awal,
        tgl_akhir
    )


    return {

        "success": True,


        "tanggal": [

            {
                "tgl": x.TGL_KERJA.strftime(
                    "%Y-%m-%d"
                ),

                "hari": x.TGL_KERJA.strftime(
                    "%a"
                ),

                "is_libur": (
                    (x.IS_LIBUR or "N").upper() == "Y"
                ),

                "keterangan": x.KET or ""

            }

            for x in data["kalender"]

        ],


        "pegawai": [

            {
                "nip": p.NIP,
                "nama": p.NAMA
            }

            for p in data["pegawai"]

        ],


        "matrix": data["matrix"]

    }




def export_rekap_clock_exception():
    """Export Rekap Exception Clock (matriks pegawai x tanggal)."""
    unit_list = request.form.getlist('unit_kerja[]')
    tgl_awal_str = request.form.get('tgl_awal')
    tgl_akhir_str = request.form.get('tgl_akhir')
    
    if not unit_list or not tgl_awal_str or not tgl_akhir_str:
        return {'error': 'Unit kosong atau format tanggal salah'}, 400
    
    try:
        unit_ids = [int(u) for u in unit_list]
    except ValueError:
        return {'error': 'Unit Kerja ID harus berupa angka'}, 400
    
    tgl_awal = datetime.strptime(tgl_awal_str, '%Y-%m-%d')
    tgl_akhir = datetime.strptime(tgl_akhir_str, '%Y-%m-%d')
    
    # 1. Ambil kalender (semua, termasuk libur)
    kalender_rows = (
        MfKalender.query
        .filter(MfKalender.TGL_KERJA.between(tgl_awal, tgl_akhir))
        .order_by(MfKalender.TGL_KERJA.asc())
        .all()
    )
    
    if not kalender_rows:
        # Fallback: generate dari rentang tanggal
        kalender_rows = []
        d = tgl_awal
        while d <= tgl_akhir:
            kalender_rows.append(type('obj', (object,), {
                'TGL_KERJA': datetime.combine(d, datetime.min.time()),
                'IS_LIBUR': 'N',
                'KET': None
            })())
            d += timedelta(days=1)
    
    n_tgl = len(kalender_rows)
    
    # 2. Ambil data absensi
    absensi_rows = (
        db.session.query(Absensi, Pegawai, MfUnitKerja)
        .join(Pegawai, Absensi.FINGER_ID == Pegawai.FINGER_ID)
        .join(MfUnitKerja, Pegawai.UNIT_KERJA_ID == MfUnitKerja.UNIT_KERJA_ID)
        .filter(Absensi.TGL_KERJA.between(tgl_awal, tgl_akhir + timedelta(days=1)))
        .filter(Pegawai.UNIT_KERJA_ID.in_(unit_ids))
        .all()
    )
    
    # ============================================================
    # 2B. Ambil DINAS LUAR untuk normalisasi absensi
    #
    # Prioritas normalisasi:
    #
    # DINAS_LUAR
    #       >
    # ABSENSI FINGER
    #
    # Digunakan oleh:
    # merge_absensi_dinas_luar()
    #
    # ============================================================

    dinas_luar_rows = (
        db.session.query(
            DinasLuar,
            Pegawai
        )
        .join(
            Pegawai,
            DinasLuar.FINGER_ID == Pegawai.FINGER_ID
        )
        .filter(
            DinasLuar.TGL_AKHIR_DINAS_LUAR >= tgl_awal
        )
        .filter(
            DinasLuar.TGL_AWAL_DINAS_LUAR <= tgl_akhir
        )
        .filter(
            Pegawai.UNIT_KERJA_ID.in_(unit_ids)
        )
        .all()
    )



    # 3. Ambil data pegawai distinct
    pegawai_list = (
        Pegawai.query
        .join(
            MfUnitKerja,
            Pegawai.UNIT_KERJA_ID == MfUnitKerja.UNIT_KERJA_ID
        )
        .outerjoin(
            MfJabatan,
            Pegawai.JABATAN_ID == MfJabatan.JABATAN_ID
        )
        .outerjoin(
            MfEselon,
            Pegawai.ESELON == MfEselon.ESELON
        )
        .outerjoin(
            MfGolongan,
            Pegawai.GOL == MfGolongan.GOL
        )
        .filter(Pegawai.UNIT_KERJA_ID.in_(unit_ids))
        .filter(
            Pegawai.TGL_MASUK <= tgl_akhir
        )
        .all()
    )

    # ============================================================
    # HRIS REBORN BUSINESS RULE
    #
    # Hanya pegawai aktif pada periode laporan yang ditampilkan.
    #
    # Mendukung:
    #   0 / N  = aktif
    #   1 / Y  = keluar
    #
    # Pegawai yang keluar setelah periode laporan masih dihitung.
    # ============================================================

    pegawai_list = [
        p for p in pegawai_list
        if is_pegawai_aktif_periode(
            p,
            tgl_awal,
            tgl_akhir
        )
    ]
    

    # ============================================================
    # SORTING TERPUSAT HRIS REBORN
    #
    # Single Source of Sorting:
    #
    # 1. Eselon
    # 2. Urut Jabatan
    # 3. Class Jabatan descending
    # 4. NIP ascending
    #
    # Rule:
    # app/utils/pegawaiSortHelper.py
    # ============================================================

    pegawai_list = sort_pegawai_rows(
        pegawai_list
    )


    print("===== DEBUG EXPORT EXCEPTION CLOCK SORT =====")
    print("TOTAL PEGAWAI =", len(pegawai_list))

    for x in pegawai_list[:10]:
        print(
            "NAMA=",
            x.NAMA,
            "| JABATAN_ID=",
            x.JABATAN_ID,
            "| CLASS_ID=",
            x.CLASS_ID
        )

    print("===== END DEBUG EXPORT EXCEPTION CLOCK SORT =====")

    if not pegawai_list:
        return {'error': 'Pegawai tidak ditemukan'}, 400
    
    # ============================================================
    # 3B. Generate ABSENSI NORMALISASI FINAL
    #
    # Sumber:
    #
    # PEGAWAI
    # ABSENSI
    # DINAS_LUAR
    #
    # Prioritas:
    #
    # DINAS_LUAR
    #       >
    # ABSENSI FINGER
    #
    # ============================================================

    normalisasi_rows = merge_absensi_dinas_luar(
        pegawai_list,
        absensi_rows,
        dinas_luar_rows,
        tgl_awal,
        tgl_akhir
    )


    # ============================================================
    # INDEX NORMALISASI
    #
    # key:
    #   (NIP, tanggal)
    #
    # value:
    #   hasil merge:
    #   ABSENSI + DINAS_LUAR
    #
    # ============================================================

    normalisasi_dict = {}

    for item in normalisasi_rows:

        key = (
            item["nip"],
            item["tanggal"].date()
        )

        normalisasi_dict[key] = item



    # 4. Build dict absensi: {nip: {tgl_str: absensi_obj}}
    absensi_dict = {}
    for a, p, uk in absensi_rows:
        tgl_key = a.TGL_KERJA.strftime('%Y-%m-%d') if a.TGL_KERJA else None
        if p.NIP not in absensi_dict:
            absensi_dict[p.NIP] = {}
        absensi_dict[p.NIP][tgl_key] = a
    
    # 5. Nama unit
    unit_names = ', '.join([u.NAMA_UNIT_KERJA for u in MfUnitKerja.query.filter(MfUnitKerja.UNIT_KERJA_ID.in_(unit_ids)).all()])
    
    # 6. Build Excel
    from openpyxl.styles import PatternFill, Font as OpFont
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Exception Clock"
    ws.sheet_properties.tabColor = "FF7B00"
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    col_paraf = 4 + n_tgl  # Kolom terakhir
    
    # Logo
    try:
        img = XLImage('static/img/LogoSAR.png')
        img.width, img.height = 50, 50
        ws.add_image(img, 'A1')
    except:
        pass
    
    # Judul
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=col_paraf)
    ws.cell(row=2, column=4, value='Rekap Exception Clock').font = Font(bold=True, size=12)
    ws.cell(row=2, column=4).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=col_paraf)
    ws.cell(row=3, column=4, value=f"Periode {tgl_awal:%d.%m.%Y} s/d {tgl_akhir:%d.%m.%Y}")
    ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=col_paraf)
    ws.cell(row=4, column=4, value=f"Unit : {unit_names}").font = Font(bold=True)
    ws.cell(row=4, column=4).alignment = Alignment(horizontal='center')
    
    thin = Side(style='thin')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    # Header: No, Nama
    ws.merge_cells(start_row=5, start_column=2, end_row=6, end_column=2)
    ws.cell(row=5, column=2, value='No').border = border
    ws.cell(row=5, column=2).alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells(start_row=5, start_column=3, end_row=6, end_column=3)
    ws.cell(row=5, column=3, value='Nama').border = border
    ws.cell(row=5, column=3).alignment = Alignment(horizontal='center', vertical='center')
    
    # Header: Tanggal (merge row 5)
    ws.merge_cells(start_row=5, start_column=4, end_row=5, end_column=col_paraf)
    ws.cell(row=5, column=4, value='Tanggal').border = border
    ws.cell(row=5, column=4).alignment = Alignment(horizontal='center')
    
    # Isi tanggal per kolom
    red_font = Font(color='FF0000')
    for i, kl in enumerate(kalender_rows):
        col = 4 + i
        tgl_val = kl.TGL_KERJA
        hari = tgl_val.strftime('%a').lower() if hasattr(kl, 'TGL_KERJA') else ''
        is_libur = kl.IS_LIBUR == 'Y' if hasattr(kl, 'IS_LIBUR') else (tgl_val.weekday() >= 5)
        
        cell = ws.cell(row=6, column=col, value=f"{tgl_val.day}\n{hari}")
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if is_libur:
            cell.font = Font(color='FF0000')
    
    # Lebar kolom
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 30
    for i in range(4, col_paraf + 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = 6
    
    # Isi data
    row = 7
    no = 0
    fill_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    fill_green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    fill_blue = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    fill_orange = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    
    print("========== DEBUG EXPORT SORT ==========")
    for idx, x in enumerate(pegawai_list[:10], start=1):
        print(idx, x.NAMA, x.JABATAN_ID, x.CLASS_ID)

    for peg in pegawai_list:
        for c in range(2, col_paraf + 1):
            ws.cell(row=row, column=c).border = border
        
        ws.cell(row=row, column=2, value=no + 1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3, value=peg.NAMA)
        
        for i, kl in enumerate(kalender_rows):
            col = 4 + i
            tgl_str = kl.TGL_KERJA.strftime('%Y-%m-%d') if hasattr(kl, 'TGL_KERJA') else ''
            is_libur = kl.IS_LIBUR == 'Y' if hasattr(kl, 'IS_LIBUR') else (kl.TGL_KERJA.weekday() >= 5)
            
            normalisasi = normalisasi_dict.get(
                (
                    peg.NIP,
                    kl.TGL_KERJA.date()
                )
            )

            cell = ws.cell(row=row, column=col)

            # Format cell harian:
            # - IN dan OUT ditampilkan dalam SATU cell
            # - Jika keduanya ada, tampil dua baris
            # - Jika hanya salah satu yang ada, tampil satu baris
            # - wrap_text diperlukan agar \n benar-benar dirender Excel
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            
            if normalisasi and normalisasi.get("sumber") == "DINAS_LUAR":

                cell.value = normalisasi.get(
                    "label",
                    "DL"
                )


                if normalisasi.get("warna") == "orange":

                    cell.font = Font(
                        color='FFA500'
                    )


                elif normalisasi.get("warna") == "blue":

                    cell.font = Font(
                        color='0000FF'
                    )


            elif normalisasi and normalisasi.get("sumber") == "ABSENSI":

                absensi = absensi_dict.get(
                    peg.NIP,
                    {}
                ).get(
                    tgl_str
                )


                transaksi = (absensi.TRANSAKSI_IN or '').upper()
                # HRIS legacy menggunakan 1900-01-01 00:00:00
                # sebagai sentinel "tidak ada jam".
                # Jangan tampilkan sentinel tersebut sebagai 00:00.
                def format_jam_aktual(value):
                    if not value:
                        return ''
                    if value.year == 1900 and value.month == 1 and value.day == 1:
                        return ''
                    return value.strftime('%H:%M')

                jam_in = format_jam_aktual(absensi.TGL_JAM_IN)
                jam_out = format_jam_aktual(absensi.TGL_JAM_OUT)
                
                if transaksi == 'WFH':
                    cell.value = 'WFH'
                elif transaksi == 'DINASLUAR':
                    cell.value = f"{jam_in}\n{jam_out}"
                    if absensi.STATUS_UM in [1, 2]:
                        cell.font = Font(color='FFA500')  # Orange
                    else:
                        cell.font = Font(color='0000FF')  # Blue
                elif transaksi == 'CUTI':
                    cell.value = '- CT -'
                    cell.font = Font(color='FFA500')
                elif transaksi == 'SAKIT':
                    cell.value = '- S -'
                    cell.font = Font(color='FFA500')
                elif transaksi == 'ALPA':
                    cell.value = 'i'
                    cell.font = Font(color='FFA500')
                else:
                    # Absensi normal:
                    # IN dan OUT ditampilkan dua baris.
                    # Warna exception ditentukan dari jam aktual
                    # dibandingkan dengan jam baku, bukan IS_INVALID.
                    if jam_in and jam_out:
                        cell.value = f"{jam_in}\n{jam_out}"
                    elif jam_in:
                        cell.value = jam_in
                    elif jam_out:
                        cell.value = jam_out

                    # ------------------------------------------------
                    # STATUS ABSENSI NORMAL / EXCEPTION
                    #
                    # Database legacy menunjukkan:
                    #   LogFP / LogFP       = fingerprint normal
                    #   LogFP / OUT NonFP   = OUT tidak fingerprint
                    #   IN NonFP / LogFP   = IN tidak fingerprint
                    #
                    # IsInValid/isOutValid TIDAK dipakai sebagai
                    # penentu warna karena record LogFP/LogFP normal
                    # juga mempunyai flag Y/Y.
                    #
                    # Jam aktual TIDAK dibandingkan secara exact dengan
                    # jam baku di sini. Contoh 07:31 vs 07:30 tidak
                    # otomatis menjadi merah.
                    # ------------------------------------------------

                    transaksi_out = (
                        getattr(absensi, 'TRANSAKSI_OUT', '') or ''
                    ).upper()

                    is_exception = False

                    # Pasangan fingerprint lengkap = NORMAL.
                    if transaksi == 'LOGFP' and transaksi_out == 'LOGFP':
                        if not jam_in or not jam_out:
                            is_exception = True

                    # Salah satu sisi bukan fingerprint = EXCEPTION.
                    elif transaksi == 'LOGFP' and transaksi_out == 'OUT NONFP':
                        is_exception = True

                    elif transaksi == 'IN NONFP' and transaksi_out == 'LOGFP':
                        is_exception = True

                    else:
                        # Kombinasi transaksi lain yang bukan transaksi
                        # khusus di atas dianggap exception.
                        is_exception = True

                    if is_exception:
                        cell.font = Font(color='FF0000')
            else:
                cell.value = ''
            
            if is_libur:
                cell.font = Font(color='FF0000')
        
        # Beri ruang untuk dua baris IN / OUT.
        # Tetap satu baris secara visual jika hanya ada satu nilai.
        ws.row_dimensions[row].height = 30

        no += 1
        row += 1
    
    # Legend
    row += 2
    ws.cell(row=row, column=2).fill = fill_red
    ws.cell(row=row, column=3, value='HARI LIBUR')
    row += 1
    ws.cell(row=row, column=2).fill = fill_green
    ws.cell(row=row, column=3, value='SIAGA')
    row += 1
    ws.cell(row=row, column=2).fill = fill_blue
    ws.cell(row=row, column=3, value='DL TIDAK TERPOTONG UANG MAKAN')
    row += 1
    ws.cell(row=row, column=2).fill = fill_orange
    ws.cell(row=row, column=3, value='TERPOTONG UANG MAKAN')
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        download_name=f"Rekap_Exception_Clock_{tgl_awal:%Y%m%d}_{tgl_akhir:%Y%m%d}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def laporan_rekap_ketidakhadiran_pegawai():
    """Render halaman Laporan Rekap Ketidakhadiran Pegawai."""
    unit_kerja_list = MfUnitKerja.query.order_by(
        MfUnitKerja.URUT_REPORT.asc(),
        MfUnitKerja.NAMA_UNIT_KERJA.asc()
    ).all()
    return render_template(
        'pages/dashboard_1/Laporan Rekap Ketidakhadiran Pegawai.html',
        unit_kerja_list=unit_kerja_list
    )

def export_rekap_ketidakhadiran_pegawai():
    """Export Rekap Sprin (DL, Sakit, Ijin, Cuti)."""
    unit_list = request.form.getlist('unit_kerja[]')
    jenis_list = request.form.getlist('jenis[]')
    tgl_awal_str = request.form.get('tgl_awal')
    tgl_akhir_str = request.form.get('tgl_akhir')
    
    if not unit_list or not tgl_awal_str or not tgl_akhir_str:
        return {'error': 'Unit kosong atau format tanggal salah'}, 400
    
    try:
        unit_ids = [int(u) for u in unit_list]
    except ValueError:
        return {'error': 'Unit Kerja ID harus berupa angka'}, 400
    
    tgl_awal = datetime.strptime(tgl_awal_str, '%Y-%m-%d')
    tgl_akhir = datetime.strptime(tgl_akhir_str, '%Y-%m-%d')
    
    # Query DinasLuar join Pegawai via NIP
    query = (
        db.session.query(DinasLuar, Pegawai, MfUnitKerja)
        .join(Pegawai, DinasLuar.NIP == Pegawai.NIP)
        .join(MfUnitKerja, Pegawai.UNIT_KERJA_ID == MfUnitKerja.UNIT_KERJA_ID)
        .filter(DinasLuar.TGL_AWAL_DINAS_LUAR <= tgl_akhir)
        .filter(DinasLuar.TGL_AKHIR_DINAS_LUAR >= tgl_awal)
        .filter(Pegawai.UNIT_KERJA_ID.in_(unit_ids))
    )
    
    if jenis_list:
        query = query.filter(DinasLuar.TRANSAKSI.in_(jenis_list))
    
    rows = query.order_by(DinasLuar.TGL_AWAL_DINAS_LUAR, DinasLuar.TRANSAKSI, Pegawai.NAMA).all()
    
    if not rows:
        return {'error': 'Data tidak ada'}, 400
    
    unit_names = ', '.join([u.NAMA_UNIT_KERJA for u in MfUnitKerja.query.filter(MfUnitKerja.UNIT_KERJA_ID.in_(unit_ids)).all()])
    jenis_names = ', '.join(jenis_list) if jenis_list else 'Semua'
    
    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Sprin"
    ws.sheet_properties.tabColor = "FF7B00"
    
    try:
        img = XLImage('static/img/LogoSAR.png')
        img.width, img.height = 50, 50
        ws.add_image(img, 'A1')
    except:
        pass
    
    thin = Side(style='thin')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    # Judul
    ws.merge_cells('D2:F2')
    ws.cell(row=2, column=4, value=f"Rekap {jenis_names}").font = Font(bold=True, size=12)
    ws.cell(row=2, column=4).alignment = Alignment(horizontal='center')
    
    ws.merge_cells('D3:F3')
    ws.cell(row=3, column=4, value=f"Periode {tgl_awal:%d.%m.%Y} s/d {tgl_akhir:%d.%m.%Y}")
    ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
    
    ws.merge_cells('D4:F4')
    ws.cell(row=4, column=4, value=f"Unit : {unit_names}").font = Font(bold=True)
    ws.cell(row=4, column=4).alignment = Alignment(horizontal='center')
    
    # Header
    for col, val in {2: 'No', 3: 'Nama', 4: 'Tanggal', 5: 'Penempatan', 6: 'Keterangan'}.items():
        c = ws.cell(row=5, column=col, value=val)
        c.border = border; c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 26
    
    # Isi data
    for i, (dl, peg, uk) in enumerate(rows, 1):
        for c in range(2, 7):
            ws.cell(row=5+i, column=c).border = border
        
        transaksi = dl.TRANSAKSI or ''
        if transaksi.upper() == 'ALPA':
            transaksi = 'Ijin'
        
        tgl_a = dl.TGL_AWAL_DINAS_LUAR.strftime('%d.%b.%Y') if dl.TGL_AWAL_DINAS_LUAR else ''
        tgl_b = dl.TGL_AKHIR_DINAS_LUAR.strftime('%d.%b.%Y') if dl.TGL_AKHIR_DINAS_LUAR else ''
        
        ws.cell(row=5+i, column=2, value=i).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=5+i, column=3, value=f"{peg.NIP}\n{peg.NAMA}").alignment = Alignment(vertical='top', wrap_text=True)
        ws.cell(row=5+i, column=4, value=f"{tgl_a}\n{tgl_b}").alignment = Alignment(vertical='top', wrap_text=True)
        ws.cell(row=5+i, column=5, value=dl.PENEMPATAN_DINAS_LUAR or '').alignment = Alignment(vertical='top', wrap_text=True)
        ws.cell(row=5+i, column=6, value=f"({transaksi}) {dl.KETERANGAN_DINAS_LUAR or ''}").alignment = Alignment(vertical='top', wrap_text=True)
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
        download_name=f"Rekap_Ketidakhadiran_{tgl_awal:%Y%m%d}_{tgl_akhir:%Y%m%d}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def laporan_rekap_pelanggaran_disiplin():
    """Render halaman Laporan Rekap Pelanggaran Disiplin."""
    unit_kerja_list = MfUnitKerja.query.order_by(
        MfUnitKerja.URUT_REPORT.asc(),
        MfUnitKerja.NAMA_UNIT_KERJA.asc()
    ).all()
    return render_template(
        'pages/dashboard_1/Laporan Rekap Pelanggaran Disiplin.html',
        unit_kerja_list=unit_kerja_list
    )

def export_rekap_pelanggaran_disiplin():
    """Export Ranking Pelanggaran Disiplin Absensi."""
    unit_list = request.form.getlist('unit_kerja[]')
    tgl_awal_str = request.form.get('tgl_awal')
    tgl_akhir_str = request.form.get('tgl_akhir')
    
    if not unit_list or not tgl_awal_str or not tgl_akhir_str:
        return {'error': 'Unit kosong atau format tanggal salah'}, 400
    
    try:
        unit_ids = [int(u) for u in unit_list]
    except ValueError:
        return {'error': 'Unit Kerja ID harus berupa angka'}, 400
    
    tgl_awal = datetime.strptime(tgl_awal_str, '%Y-%m-%d')
    tgl_akhir = datetime.strptime(tgl_akhir_str, '%Y-%m-%d')
    
    tgl_server = datetime.now()
    if tgl_server.date() < tgl_awal.date():
        return {'error': 'Tgl server lebih kecil dari tanggal awal periode'}, 400
    if tgl_server.date() < tgl_akhir.date():
        tgl_akhir = tgl_server
    
    # Kalender hari kerja
    kalender_rows = (
        MfKalender.query
        .filter(MfKalender.TGL_KERJA.between(tgl_awal, tgl_akhir))
        .filter(MfKalender.IS_LIBUR == 'N')
        .all()
    )
    default_tgl_kerja = len(kalender_rows)
    
    # Ambil data absensi (join via NIP)
    absensi_rows = (
        db.session.query(Absensi, Pegawai)
        .join(Pegawai, Absensi.FINGER_ID == Pegawai.FINGER_ID)
        .join(MfKalender, Absensi.TGL_KERJA == MfKalender.TGL_KERJA)
        .filter(Absensi.TGL_KERJA.between(tgl_awal, tgl_akhir))
        .filter(MfKalender.IS_LIBUR == 'N')
        .filter(Pegawai.UNIT_KERJA_ID.in_(unit_ids))
        .filter(Pegawai.IS_VIP == 0)  # Exclude VIP
        .all()
    )
    
    if not absensi_rows:
        return {'error': 'Record tidak ada atau kalender belum dibuat'}, 400
    
    # Ambil data pegawai
    pegawai_list = (
        Pegawai.query
        .filter(Pegawai.UNIT_KERJA_ID.in_(unit_ids))
        .filter(Pegawai.IS_VIP == 0)
        .filter(Pegawai.TGL_MASUK <= tgl_akhir)
        .filter(
            db.or_(
                Pegawai.IS_KELUAR == 'N',
                db.and_(Pegawai.IS_KELUAR == 'Y', Pegawai.TGL_KELUAR >= tgl_awal)
            )
        )
        .order_by(Pegawai.NAMA)
        .all()
    )
    
    # Build dict absensi per NIP
    from collections import defaultdict
    absensi_dict = defaultdict(list)
    for a, p in absensi_rows:
        absensi_dict[p.NIP].append(a)
    
    # Hitung pelanggaran per pegawai
    hasil = []
    for peg in pegawai_list:
        abs_list = absensi_dict.get(peg.NIP, [])
        
        # TLM tanpa keterangan
        tot_tlm_a = sum(a.AWAL_TLM or 0 for a in abs_list 
                       if a.TINGKAT_TLM in ('TLM-1','TLM-2','TLM-3','TLM-4') 
                       and a.PENDUKUNG_IN == 'N' 
                       and a.TRANSAKSI_IN in ('LogFP','Manual','-','IN NonFP'))
        
        # PSW tanpa keterangan
        tot_psw_a = sum(a.TOTAL_PSW or 0 for a in abs_list 
                       if a.TINGKAT_PSW in ('PSW-1','PSW-2','PSW-3','PSW-4') 
                       and a.PENDUKUNG_OUT == 'N' 
                       and a.TRANSAKSI_OUT in ('LogFP','Manual','-','OUT NonFP'))
        
        # Sakit tanpa keterangan
        sakit_a = sum(1 for a in abs_list 
                     if a.TRANSAKSI_IN and a.TRANSAKSI_IN.strip().upper() == 'SAKIT' 
                     and a.PENDUKUNG_IN == 'N')
        
        # Alpa tanpa keterangan + tidak masuk tanpa ket
        tgl_masuk = peg.TGL_MASUK
        if tgl_masuk and tgl_masuk.date() > tgl_awal.date():
            xn_tgl_kerja = len([k for k in kalender_rows if k.TGL_KERJA and k.TGL_KERJA > tgl_masuk])
        else:
            xn_tgl_kerja = default_tgl_kerja
        
        alpa_a = sum(1 for a in abs_list 
                    if a.TRANSAKSI_IN and a.TRANSAKSI_IN.strip().upper() == 'ALPA' 
                    and a.PENDUKUNG_IN == 'N')
        
        tdk_masuk_tanpa_ket = alpa_a + (xn_tgl_kerja - len(abs_list))
        if tdk_masuk_tanpa_ket < 0:
            tdk_masuk_tanpa_ket = 0
        
        # Grand total menit
        grand_tot_menit = abs(tot_psw_a) + tot_tlm_a + (sakit_a * 60 * 8) + (tdk_masuk_tanpa_ket * 60 * 8)
        
        if grand_tot_menit > 0:
            tot_hr = grand_tot_menit // (60 * 8)
            sisa_menit = grand_tot_menit % (60 * 8)
            tot_jam = sisa_menit // 60
            tot_menit = sisa_menit % 60
            
            hasil.append({
                'nip': peg.NIP,
                'nama': peg.NAMA,
                'grand_tot_menit': grand_tot_menit,
                'tot_tlm_a': tot_tlm_a,
                'tot_psw_a': abs(tot_psw_a),
                'sakit_a': sakit_a,
                'tdk_masuk': tdk_masuk_tanpa_ket,
                'tot_hr': tot_hr,
                'tot_jam': tot_jam,
                'tot_menit': tot_menit,
            })
    
    if not hasil:
        return {'error': 'Tidak ada pegawai yang melanggar disiplin absensi'}, 400
    
    # Sort by grand_tot_menit descending
    hasil.sort(key=lambda x: x['grand_tot_menit'], reverse=True)
    
    # Nama unit
    unit_names = ', '.join([u.NAMA_UNIT_KERJA for u in MfUnitKerja.query.filter(MfUnitKerja.UNIT_KERJA_ID.in_(unit_ids)).all()])
    
    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Pelanggaran"
    ws.sheet_properties.tabColor = "FF7B00"
    
    try:
        img = XLImage('static/img/LogoSAR.png')
        img.width, img.height = 50, 50
        ws.add_image(img, 'A1')
    except:
        pass
    
    thin = Side(style='thin')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    col_paraf = 10
    
    # Judul
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=col_paraf)
    ws.cell(row=2, column=4, value='Ranking Pelanggaran Disiplin Absensi').font = Font(bold=True, size=12)
    ws.cell(row=2, column=4).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=col_paraf)
    ws.cell(row=3, column=4, value=f"Periode {tgl_awal:%d.%m.%Y} s/d {tgl_akhir:%d.%m.%Y}")
    ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=col_paraf)
    ws.cell(row=4, column=4, value=f"Unit : {unit_names}").font = Font(bold=True)
    ws.cell(row=4, column=4).alignment = Alignment(horizontal='center')
    
    # Header
    ws.merge_cells('B5:B6')
    ws.cell(row=5, column=2, value='No').border = border
    ws.cell(row=5, column=2).font = Font(bold=True)
    ws.cell(row=5, column=2).alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('C5:C6')
    ws.cell(row=5, column=3, value='Nama').border = border
    ws.cell(row=5, column=3).font = Font(bold=True)
    ws.cell(row=5, column=3).alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('D5:G5')
    ws.cell(row=5, column=4, value='Tanpa Keterangan').border = border
    ws.cell(row=5, column=4).font = Font(bold=True)
    ws.cell(row=5, column=4).alignment = Alignment(horizontal='center')
    
    for col, val in {4: 'TLM\n(menit)', 5: 'PSW\n(menit)', 6: 'Sakit\n(Hr Kerja)', 7: 'Ketidakhadiran\n(Hr Kerja)'}.items():
        c = ws.cell(row=6, column=col, value=val)
        c.border = border; c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.merge_cells('H5:J5')
    ws.cell(row=5, column=8, value='Total').border = border
    ws.cell(row=5, column=8).font = Font(bold=True)
    ws.cell(row=5, column=8).alignment = Alignment(horizontal='center')
    
    for col, val in {8: 'Hr\nKerja', 9: 'Jam', 10: 'Menit'}.items():
        c = ws.cell(row=6, column=col, value=val)
        c.border = border; c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Lebar kolom
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 30
    for c in ['D','E','F','G','H','I','J']:
        ws.column_dimensions[c].width = 10
    
    # Isi data
    for i, row_data in enumerate(hasil, 1):
        for c in range(2, 11):
            ws.cell(row=6+i, column=c).border = border
        
        ws.cell(row=6+i, column=2, value=i).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=6+i, column=3, value=f"{row_data['nama']}\n{row_data['nip']}").alignment = Alignment(vertical='top', wrap_text=True)
        ws.cell(row=6+i, column=4, value=row_data['tot_tlm_a'] or None).alignment = Alignment(horizontal='right')
        ws.cell(row=6+i, column=5, value=row_data['tot_psw_a'] or None).alignment = Alignment(horizontal='right')
        ws.cell(row=6+i, column=6, value=row_data['sakit_a'] or None).alignment = Alignment(horizontal='right')
        ws.cell(row=6+i, column=7, value=row_data['tdk_masuk'] or None).alignment = Alignment(horizontal='right')
        ws.cell(row=6+i, column=8, value=row_data['tot_hr'] or None).alignment = Alignment(horizontal='right')
        ws.cell(row=6+i, column=9, value=row_data['tot_jam'] or None).alignment = Alignment(horizontal='right')
        ws.cell(row=6+i, column=10, value=row_data['tot_menit'] or None).alignment = Alignment(horizontal='right')
    
    last_row = 6 + len(hasil)
    ws.cell(row=last_row+2, column=3, value='Keterangan')
    ws.cell(row=last_row+3, column=3, value='TLM : Terlambat Masuk')
    ws.cell(row=last_row+4, column=3, value='PSW : Pulang Sebelum Waktu')
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
        download_name=f"Rekap_Pelanggaran_Disiplin_{tgl_awal:%Y%m%d}_{tgl_akhir:%Y%m%d}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def laporan_rekap_uang_makan():
    """Render halaman Laporan Rekap Uang Makan."""
    unit_kerja_list = MfUnitKerja.query.filter(
        MfUnitKerja.IS_USE == 'Y'
    ).order_by(
        MfUnitKerja.URUT_REPORT.asc(),
        MfUnitKerja.NAMA_UNIT_KERJA.asc()
    ).all()

    current_year = datetime.now().year

    return render_template(
        'pages/dashboard_1/Laporan Rekap Uang Makan.html',
        unit_kerja_list=unit_kerja_list,
        current_year=current_year
    )

def export_rekap_uang_makan(preview=False):
    """Export / Preview Rekap Uang Makan (seperti RekapUM di VB.NET)."""
    unit_list = request.form.getlist('unit_kerja[]')
    bulan_str = request.form.get('bulan', '')  # Format: YYYY-MM
    staf_kepegawaian = request.form.get('staf_kepegawaian', '')
    kasubag_umum = request.form.get('kasubag_umum', '')
    
    if not unit_list or not bulan_str:
        return {'error': 'Unit atau bulan kosong'}, 400
    
    try:
        unit_ids = [int(u) for u in unit_list]
    except ValueError:
        return {'error': 'Unit Kerja ID harus berupa angka'}, 400
    
    # Parse bulan -> tgl_awal & tgl_akhir
    tahun, bulan = map(int, bulan_str.split('-'))
    tgl_awal = datetime(tahun, bulan, 1)
    # Akhir bulan
    if bulan == 12:
        tgl_akhir = datetime(tahun + 1, 1, 1) - timedelta(days=1)
    else:
        tgl_akhir = datetime(tahun, bulan + 1, 1) - timedelta(days=1)
    
    # Cek tgl server
    tgl_server = datetime.now()
    if tgl_server.date() < tgl_awal.date():
        return {'error': 'Tgl server lebih kecil dari tanggal awal periode'}, 400
    if tgl_server.date() < tgl_akhir.date():
        tgl_akhir = tgl_server
    
    # Ambil kalender hari kerja
    kalender_rows = (
        MfKalender.query
        .filter(MfKalender.TGL_KERJA.between(tgl_awal, tgl_akhir))
        .filter(MfKalender.IS_LIBUR == 'N')
        .all()
    )
    default_tgl_kerja = len(kalender_rows)
    
    # Ambil nominal Uang Makan reguler yang berlaku.
    # Master HRIS:
    #   JenisTunjangan = U.Makan
    #   Activity       = Intern
    #   HariKerja      = 0
    # Nominal tidak di-hardcode; tetap mengikuti MF_TUNJANGAN.
    um_row = (
        MfTunjangan.query
        .filter(MfTunjangan.JENIS_TUNJANGAN == 'U.Makan')
        .filter(MfTunjangan.ACTIVITY == 'Intern')
        .filter(MfTunjangan.HARI_KERJA == 0)
        .filter(MfTunjangan.TGL_MULAI <= tgl_akhir.date())
        .order_by(
            MfTunjangan.TGL_MULAI.desc(),
            MfTunjangan.IDTUNJANGAN.desc()
        )
        .first()
    )
    nominal_um = um_row.NOMINAL if um_row else 0
    
    # Ambil data absensi (join via NIP)
    absensi_rows = (
        db.session.query(Absensi, Pegawai)
        .join(Pegawai, Absensi.FINGER_ID == Pegawai.FINGER_ID)
        .join(MfKalender, Absensi.TGL_KERJA == MfKalender.TGL_KERJA)
        .filter(Absensi.TGL_KERJA.between(tgl_awal, tgl_akhir))
        .filter(MfKalender.IS_LIBUR == 'N')
        .filter(Pegawai.UNIT_KERJA_ID.in_(unit_ids))
        .all()
    )
    
    # Ambil data pegawai
    pegawai_list = (
        Pegawai.query
        .filter(Pegawai.UNIT_KERJA_ID.in_(unit_ids))
        .filter(
            db.or_(
                db.and_(Pegawai.TGL_MASUK <= tgl_akhir, Pegawai.IS_KELUAR == 'N'),
                db.and_(Pegawai.IS_KELUAR == 'Y', Pegawai.TGL_KELUAR >= tgl_awal)
            )
        )
        .all()
    )

    # ================================================================
    # SORTING STANDARD HRIS REBORN
    #
    # Single Source of Sorting:
    #   1. Eselon
    #   2. Urut Jabatan
    #   3. Class Jabatan DESC
    #   4. NIP ASC
    # ================================================================
    pegawai_list = sort_pegawai_rows(pegawai_list)

    # ================================================================
    # MASTER PANGKAT
    #
    # Sumber resmi nama pangkat adalah MF_GOLONGAN berdasarkan GOL_ID.
    # Jangan menggunakan Pegawai.PANGKAT karena merupakan data legacy
    # yang dapat berbeda antarpegawai dengan golongan yang sama.
    # ================================================================
    pangkat_map = {
        row.GOL_ID: row.PANGKAT
        for row in MfGolongan.query.all()
        if row.GOL_ID
    }
    
    if not pegawai_list:
        return {'error': 'Pegawai tidak ditemukan'}, 400
    
    # ================================================================
    # DATA ABSENSI PER PEGAWAI
    # ================================================================
    from collections import defaultdict

    absensi_dict = defaultdict(list)
    for a, p in absensi_rows:
        absensi_dict[p.NIP].append(a)

    # ================================================================
    # DATA DINAS LUAR UNTUK POTONG UANG MAKAN
    #
    # STATUS_UM:
    #   0 = tidak memotong UM
    #   1 = MEMOTONG UM
    #   2 = tidak memotong UM / penempatan
    #
    # Sumber resmi transaksi DL adalah DINAS_LUAR.
    # Tanggal disimpan dalam SET agar overlapping SPRIN pada tanggal
    # yang sama tidak dihitung dua kali.
    # ================================================================
    kalender_dates = {
        k.TGL_KERJA.date()
        if hasattr(k.TGL_KERJA, 'date')
        else k.TGL_KERJA
        for k in kalender_rows
        if k.TGL_KERJA
    }

    dinas_luar_rows = (
        DinasLuar.query
        .filter(DinasLuar.TRANSAKSI == 'DinasLuar')
        .filter(DinasLuar.STATUS_UM == 1)
        .filter(DinasLuar.TGL_AWAL_DINAS_LUAR <= tgl_akhir)
        .filter(DinasLuar.TGL_AKHIR_DINAS_LUAR >= tgl_awal)
        .all()
    )

    dl_dates_dict = defaultdict(set)

    for dl in dinas_luar_rows:
        if (
            not dl.FINGER_ID
            or not dl.TGL_AWAL_DINAS_LUAR
            or not dl.TGL_AKHIR_DINAS_LUAR
        ):
            continue

        dl_awal = max(
            dl.TGL_AWAL_DINAS_LUAR.date(),
            tgl_awal.date()
        )
        dl_akhir = min(
            dl.TGL_AKHIR_DINAS_LUAR.date(),
            tgl_akhir.date()
        )

        if dl_awal > dl_akhir:
            continue

        current_date = dl_awal
        while current_date <= dl_akhir:
            if current_date in kalender_dates:
                dl_dates_dict[str(dl.FINGER_ID)].add(current_date)
            current_date += timedelta(days=1)

    # Nama unit
    unit_names = ', '.join([
        u.NAMA_UNIT_KERJA
        for u in MfUnitKerja.query
        .filter(MfUnitKerja.UNIT_KERJA_ID.in_(unit_ids))
        .all()
    ])

    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Uang Makan"
    ws.sheet_properties.tabColor = "FF7B00"
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    col_paraf = 16

    # Logo
    try:
        img = XLImage('static/img/LogoSAR.png')
        img.width, img.height = 50, 50
        ws.add_image(img, 'A1')
    except:
        pass

    thin = Side(style='thin')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    green_fill = PatternFill(start_color='ADFF2F', end_color='ADFF2F', fill_type='solid')
    gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # === JUDUL ===
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=col_paraf-1)
    ws.cell(row=2, column=4, value='REKAP UANG MAKAN').font = Font(bold=True, size=14)
    ws.cell(row=2, column=4).alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=col_paraf-1)
    ws.cell(row=3, column=4, value=f"PEGAWAI KANTOR PENCARIAN DAN PERTOLONGAN {unit_names}")
    ws.cell(row=3, column=4).font = Font(bold=True)
    ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')

    # Info bulan & hari kerja
    ws.cell(row=5, column=14, value='Bulan').font = Font(bold=True)
    ws.cell(row=5, column=15, value=f": {tgl_awal:%B %Y}")
    ws.cell(row=6, column=14, value='Hari Kerja').font = Font(bold=True)
    ws.cell(row=6, column=15, value=f": {default_tgl_kerja}")

    # === HEADER (row 7-9) ===
    ws.cell(row=7, column=2, value='No')
    ws.cell(row=7, column=3, value='NIP')
    ws.cell(row=7, column=4, value='Nama')
    ws.cell(row=7, column=5, value='Pangkat')
    ws.cell(row=7, column=6, value='Jenis Absensi')
    ws.merge_cells(start_row=7, start_column=6, end_row=7, end_column=10)
    ws.cell(row=7, column=11, value='JUMLAH HARI')
    ws.merge_cells(start_row=7, start_column=11, end_row=7, end_column=12)
    ws.cell(row=7, column=13, value='JUMLAH UANG (Rp)')
    ws.cell(row=7, column=14, value='TANDA TANGAN')
    ws.merge_cells(start_row=7, start_column=14, end_row=7, end_column=15)

    ws.cell(row=8, column=6, value='DL')
    ws.cell(row=8, column=7, value='CT')
    ws.cell(row=8, column=8, value='i')
    ws.cell(row=8, column=9, value='S')
    ws.cell(row=8, column=10, value='TA')

    ws.merge_cells('B7:B9')
    ws.merge_cells('C7:C9')
    ws.merge_cells('D7:D9')
    ws.merge_cells('E7:E9')
    ws.merge_cells('M7:M9')
    ws.merge_cells('K9:L9')
    ws.merge_cells('N9:O9')

    # Styling header
    for r in range(7, 10):
        for c in range(2, 16):
            try:
                cell = ws.cell(row=r, column=c)
                cell.border = border
                cell.fill = green_fill
                cell.font = Font(bold=True, size=9)
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
            except AttributeError:
                pass

    # Lebar kolom
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20

    for c in ['F','G','H','I','J','K','L']:
        ws.column_dimensions[c].width = 5

    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 17
    ws.column_dimensions['O'].width = 17

    # === ISI DATA ===
    row = 10
    no = 1
    total_um = 0
    preview_rows = []

    for peg in pegawai_list:
        abs_list = absensi_dict.get(peg.NIP, [])

        # ============================================================
        # KUMPULKAN TANGGAL CUTI / SAKIT / IJIN
        # ============================================================
        cuti_dates = set()
        sakit_dates = set()
        ijin_dates = set()
        alpa_tanpa_dates = set()

        for a in abs_list:
            if not a.TGL_KERJA:
                continue

            abs_date = (
                a.TGL_KERJA.date()
                if hasattr(a.TGL_KERJA, 'date')
                else a.TGL_KERJA
            )

            if abs_date not in kalender_dates:
                continue

            transaksi = (
                a.TRANSAKSI_IN.strip().upper()
                if a.TRANSAKSI_IN
                else ''
            )

            if transaksi == 'CUTI':
                cuti_dates.add(abs_date)

            elif transaksi == 'SAKIT':
                sakit_dates.add(abs_date)

            elif transaksi == 'ALPA':
                if a.PENDUKUNG_IN == 'Y':
                    ijin_dates.add(abs_date)
                else:
                    alpa_tanpa_dates.add(abs_date)

        # ============================================================
        # HARI KERJA PEGAWAI
        # ============================================================
        tgl_masuk = peg.TGL_MASUK

        if tgl_masuk:
            tgl_masuk_date = (
                tgl_masuk.date()
                if hasattr(tgl_masuk, 'date')
                else tgl_masuk
            )
        else:
            tgl_masuk_date = None

        if tgl_masuk_date and tgl_masuk_date > tgl_awal.date():
            # Tanggal mulai bekerja tetap dihitung apabila merupakan
            # hari kerja kalender.
            xn_tgl_kerja = len([
                k for k in kalender_dates
                if k >= tgl_masuk_date
            ])
        else:
            xn_tgl_kerja = default_tgl_kerja

        # TA tetap ditampilkan sebagai informasi.
        # TA TIDAK mengurangi Uang Makan.
        ta = len(alpa_tanpa_dates) + max(
            0,
            xn_tgl_kerja - len({
                (
                    a.TGL_KERJA.date()
                    if hasattr(a.TGL_KERJA, 'date')
                    else a.TGL_KERJA
                )
                for a in abs_list
                if a.TGL_KERJA
            })
        )

        # ============================================================
        # DINAS LUAR POTONG UM
        # ============================================================
        dl_dates = dl_dates_dict.get(str(peg.FINGER_ID), set())

        # Satu tanggal hanya boleh mengurangi UM satu kali.
        # Prioritaskan DL sebagai alasan potongan apabila tanggal yang
        # sama juga mempunyai record absensi cuti/sakit/ijin.
        cuti_dates -= dl_dates
        sakit_dates -= dl_dates
        ijin_dates -= dl_dates

        dl_count = len(dl_dates)
        total_cuti = len(cuti_dates)
        sakit_all = len(sakit_dates)
        alpa_ijin = len(ijin_dates)

        # ============================================================
        # FORMULA UANG MAKAN
        #
        # Hari Kerja Efektif =
        #   Hari Kerja Master
        #   - Cuti
        #   - Sakit
        #   - Ijin
        #   - DL StatusUM=1
        #   - TA (Tidak Hadir)
        # ============================================================
        jumlah_hari = xn_tgl_kerja - (
            total_cuti
            + sakit_all
            + alpa_ijin
            + dl_count
            + ta
        )

        if jumlah_hari < 0:
            jumlah_hari = 0

        um = jumlah_hari * nominal_um
        total_um += um

        preview_rows.append({
            "no": no,
            "nama": peg.NAMA or "",
            "pangkat": f"{peg.GOL_ID} - {pangkat_map.get(peg.GOL_ID, '-')}",
            "nip": peg.NIP or "",
            "dl": dl_count,
            "cuti": total_cuti,
            "ijin": alpa_ijin,
            "sakit": sakit_all,
            "ta": ta,
            "jumlah_hari": jumlah_hari,
            "jumlah_uang": um,
        })

        # Tulis data
        for c in range(2, 16):
            ws.cell(row=row, column=c).border = border

        ws.cell(
            row=row,
            column=2,
            value=no
        ).alignment = Alignment(horizontal='center')

        ws.cell(
            row=row,
            column=3,
            value=peg.NIP
        ).alignment = Alignment(horizontal='left')

        ws.cell(
            row=row,
            column=4,
            value=peg.NAMA
        ).alignment = Alignment(horizontal='left')

        ws.cell(
            row=row,
            column=5,
            value=f"{peg.GOL_ID} - {pangkat_map.get(peg.GOL_ID, '-')}"
        ).alignment = Alignment(horizontal='left')

        if dl_count > 0:
            ws.cell(row=row, column=6, value=dl_count).alignment = Alignment(horizontal='center')

        if total_cuti > 0:
            ws.cell(row=row, column=7, value=total_cuti).alignment = Alignment(horizontal='center')

        if alpa_ijin > 0:
            ws.cell(row=row, column=8, value=alpa_ijin).alignment = Alignment(horizontal='center')

        if sakit_all > 0:
            ws.cell(row=row, column=9, value=sakit_all).alignment = Alignment(horizontal='center')

        if ta > 0:
            ws.cell(row=row, column=10, value=ta).alignment = Alignment(horizontal='center')

        ws.cell(row=row, column=11, value=jumlah_hari).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=12, value='hari').alignment = Alignment(horizontal='center')

        if um > 0:
            ws.cell(row=row, column=13, value=um)
            ws.cell(row=row, column=13).number_format = '#,##0'

        # Tanda tangan
        if no % 2 == 1:
            ws.merge_cells(start_row=row, start_column=14, end_row=row+1, end_column=14)
            ws.merge_cells(start_row=row, start_column=15, end_row=row+1, end_column=15)
            ws.cell(row=row, column=14, value=no).alignment = Alignment(vertical='top')

        no += 1
        row += 1

    # ================================================================
    # PREVIEW
    # ================================================================
    if preview:
        return {
            "success": True,
            "bulan": bulan_str,
            "hari_kerja": default_tgl_kerja,
            "nominal_um": nominal_um,
            "unit_names": unit_names,
            "total_um": total_um,
            "rows": preview_rows,
        }

    # === BARIS TOTAL ===
    if no % 2 == 0:
        row += 1
    
    for c in range(2, 16):
        ws.cell(row=row, column=c).border = border
        ws.cell(row=row, column=c).fill = gray_fill
    
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row=row, column=2, value='TOTAL').font = Font(bold=True)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=12, value='Rp').font = Font(bold=True)
    ws.cell(row=row, column=13, value=total_um)
    ws.cell(row=row, column=13).number_format = '#,##0'
    
    # === TANDA TANGAN ===
    row += 2
    ws.cell(row=row, column=13, value=f"Surabaya, {tgl_akhir:%B %Y}")
    ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=15)
    
    row += 1
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=9)
    ws.cell(row=row, column=6, value='Mengetahui,')
    ws.cell(row=row, column=13, value='Staf Kepegawaian')
    ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=15)
    
    row += 1
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=9)
    ws.cell(row=row, column=6, value='Kasubag Umum')
    
    row += 4
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=11)
    ws.cell(row=row, column=6, value=kasubag_umum).font = Font(underline='single')
    ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=15)
    ws.cell(row=row, column=13, value=staf_kepegawaian).font = Font(underline='single')
    
    row += 1
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=11)
    ws.cell(row=row, column=6, value='NIP. ................................')
    ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=15)
    ws.cell(row=row, column=13, value='NIP. ................................')
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        download_name=f"Rekap_Uang_Makan_{tgl_awal:%Y%m}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def laporan_rekap_tunjangan_kinerja():
    """Render halaman Laporan Rincian Pembayaran Tunjangan."""
    unit_kerja_list = MfUnitKerja.query.order_by(
        MfUnitKerja.URUT_REPORT.asc(),
        MfUnitKerja.NAMA_UNIT_KERJA.asc()
    ).all()
    return render_template(
        'pages/dashboard_1/Laporan Rincian Pembayaran Tunjangan.html',
        unit_kerja_list=unit_kerja_list
    )

def export_rekap_tunjangan_kinerja():
    """Export Rincian Pembayaran Tunjangan Kinerja."""
    unit_list = request.form.getlist('unit_kerja[]')
    tgl_awal_str = request.form.get('tgl_awal')
    tgl_akhir_str = request.form.get('tgl_akhir')
    
    if not unit_list or not tgl_awal_str or not tgl_akhir_str:
        return {'error': 'Unit kosong atau format tanggal salah'}, 400
    
    try:
        unit_ids = [int(u) for u in unit_list]
    except ValueError:
        return {'error': 'Unit Kerja ID harus berupa angka'}, 400
    
    tgl_awal = datetime.strptime(tgl_awal_str, '%Y-%m-%d')
    tgl_akhir = datetime.strptime(tgl_akhir_str, '%Y-%m-%d')
    
    # Cek tgl server
    tgl_server = datetime.now()
    if tgl_server.date() < tgl_awal.date():
        return {'error': 'Tgl server lebih kecil dari tanggal awal periode'}, 400
    if tgl_server.date() < tgl_akhir.date():
        tgl_akhir = tgl_server
    
    # Ambil kalender
    kalender_rows = (
        MfKalender.query
        .filter(MfKalender.TGL_KERJA.between(tgl_awal, tgl_akhir))
        .all()
    )
    
    # Ambil data absensi (join via NIP)
    absensi_rows = (
        db.session.query(Absensi, Pegawai)
        .join(Pegawai, Absensi.FINGER_ID == Pegawai.FINGER_ID)
        .join(MfKalender, Absensi.TGL_KERJA == MfKalender.TGL_KERJA)
        .filter(Absensi.TGL_KERJA.between(tgl_awal, tgl_akhir))
        .filter(MfKalender.IS_LIBUR == 'N')
        .filter(Pegawai.UNIT_KERJA_ID.in_(unit_ids))
        .all()
    )
    
    # Ambil data pegawai
    pegawai_list = (
        Pegawai.query
        .filter(Pegawai.UNIT_KERJA_ID.in_(unit_ids))
        .filter(Pegawai.TGL_MASUK <= tgl_akhir)
        .filter(
            db.or_(
                Pegawai.IS_KELUAR == 'N',
                db.and_(Pegawai.IS_KELUAR == 'Y', Pegawai.TGL_KELUAR >= tgl_awal)
            )
        )
        .order_by(Pegawai.NAMA)
        .all()
    )
    
    if not pegawai_list:
        return {'error': 'Pegawai tidak ditemukan'}, 400
    
    # Ambil MFPot untuk persentase potongan
    potongan_list = (
        MfPot.query
        .filter(MfPot.TGL_MULAI <= tgl_akhir)
        .all()
    )
    
    # Ambil tunjangan per class
    class_tunjangan = {}
    for c in MfClass.query.filter(MfClass.TGL_MULAI <= tgl_akhir.date()).order_by(MfClass.TGL_MULAI.desc()).all():
        if c.CLASS_ID not in class_tunjangan:
            class_tunjangan[c.CLASS_ID] = c.TUNJANGAN
    
    # DinasLuar > 4 bulan
    dl_rows = (
        DinasLuar.query
        .join(Pegawai, DinasLuar.NIP == Pegawai.NIP)
        .filter(DinasLuar.TRANSAKSI == 'DinasLuar')
        .filter(DinasLuar.STATUS_UM == 1)
        .filter(DinasLuar.TGL_AKHIR_DINAS_LUAR >= DinasLuar.TGL_AWAL_DINAS_LUAR)  # > 4 bulan
        .filter(DinasLuar.TGL_AWAL_DINAS_LUAR <= tgl_akhir)
        .filter(DinasLuar.TGL_AKHIR_DINAS_LUAR >= tgl_awal)
        .filter(Pegawai.UNIT_KERJA_ID.in_(unit_ids))
        .all()
    )
    
    # Build dict absensi
    absensi_dict = defaultdict(list)
    for a, p in absensi_rows:
        tgl_key = a.TGL_KERJA.strftime('%Y-%m-%d') if a.TGL_KERJA else None
        absensi_dict[p.NIP].append(a)
    
    # Build dict DL
    dl_dict = defaultdict(list)
    for dl in dl_rows:
        dl_dict[dl.NIP].append(dl)
    
    pot_dict = {}
    for p in potongan_list:
        pot_dict[p.KATEGORI] = p.PERSEN_POT or 0
    
    pot_ta = pot_dict.get('TA', 0)  # 5%
    pot_dl = pot_dict.get('DINASLUAR', 0)  # 1%
    
    # Hitung per pegawai
    hasil = []
    for peg in pegawai_list:
        abs_list = absensi_dict.get(peg.NIP, [])
        tunjangan = class_tunjangan.get(peg.CLASS_ID, 0) or 0
        
        # ✅ Buat dict lookup per tanggal untuk akses cepat
        abs_lookup = {}
        for a in abs_list:
            tgl_key = a.TGL_KERJA.strftime('%Y-%m-%d') if a.TGL_KERJA else None
            abs_lookup[tgl_key] = a
        
        persen_pot = 0
        tgl_masuk = peg.TGL_MASUK
        tgl_hitung = tgl_masuk if tgl_masuk and tgl_masuk > tgl_awal else tgl_awal
        
        d = tgl_hitung
        while d.date() <= tgl_akhir.date():
            # Cek libur dari kalender_rows
            is_libur = False
            tgl_str = d.strftime('%Y-%m-%d')
            
            kl = [k for k in kalender_rows if k.TGL_KERJA and k.TGL_KERJA.strftime('%Y-%m-%d') == tgl_str]
            if kl:
                is_libur = kl[0].IS_LIBUR == 'Y'
            elif d.weekday() >= 5:
                is_libur = True
            
            if not is_libur:
                # ✅ Cari record absensi via lookup dict
                a = abs_lookup.get(tgl_str)
                
                if a:
                    transaksi = (a.TRANSAKSI_IN or '').strip().lower()
                    
                    if transaksi in ('alpa', 'sakit', 'ijin'):
                        persen_pot += a.PERSEN_POT_TLM or 0
                    elif transaksi == 'dinasluar':
                        pass
                    else:
                        persen_pot += (a.PERSEN_POT_TLM or 0) + (a.PERSEN_POT_PSW or 0)
                else:
                    persen_pot += pot_ta
                
                # Cek DL > 4 bulan
                dl_peg = dl_dict.get(peg.NIP, [])
                for dl in dl_peg:
                    if dl.TGL_AWAL_DINAS_LUAR:
                        limit_dl = dl.TGL_AWAL_DINAS_LUAR + timedelta(days=120)
                        tgl_akhir_dl = dl.TGL_AKHIR_DINAS_LUAR.date() if dl.TGL_AKHIR_DINAS_LUAR else d.date()
                        if limit_dl.date() <= d.date() <= tgl_akhir_dl:
                            persen_pot += pot_dl
                            break
            
            d += timedelta(days=1)
        
        # ✅ Hitung nilai potongan (seperti VB.NET: tunjangan * persen_pot / 100)
        nilai_pot = tunjangan * (persen_pot / 100) if persen_pot > 0 else 0
        jumlah_dibayarkan = tunjangan - nilai_pot
        
        hasil.append({
            'nama': peg.NAMA,
            'nip': peg.NIP,
            'jabatan': f"TMT: {peg.TMT_JABATAN.strftime('%d/%m/%Y') if peg.TMT_JABATAN else '-'}",
            'class_id': peg.CLASS_ID,
            'tunjangan': tunjangan,
            'persen_pot': persen_pot,
            'nilai_pot': nilai_pot,
            'jumlah_pot': nilai_pot,
            'dibayarkan': jumlah_dibayarkan,
        })
    
    if not hasil:
        return {'error': 'Record tidak ada'}, 400
    
    # Nama unit
    unit_names = ', '.join([u.NAMA_UNIT_KERJA for u in MfUnitKerja.query.filter(MfUnitKerja.UNIT_KERJA_ID.in_(unit_ids)).all()])
    
    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Tunjangan"
    ws.sheet_properties.tabColor = "FF7B00"
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    try:
        img = XLImage('static/img/LogoSAR.png')
        img.width, img.height = 50, 50
        ws.add_image(img, 'A1')
    except:
        pass
    
    thin = Side(style='thin')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    col_paraf = 12
    
    # Judul
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=col_paraf)
    ws.cell(row=2, column=4, value='Rincian Pembayaran Tunjangan').font = Font(bold=True, size=12)
    ws.cell(row=2, column=4).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=col_paraf)
    ws.cell(row=3, column=4, value=f"Periode {tgl_awal:%d.%m.%Y} s.d {tgl_akhir:%d.%m.%Y}")
    ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=col_paraf)
    ws.cell(row=4, column=4, value=f"Unit : {unit_names}").font = Font(bold=True)
    ws.cell(row=4, column=4).alignment = Alignment(horizontal='center')
    
    # Header
    headers = {2: 'No', 3: 'Nama', 4: 'NIP', 5: 'Status\nKepeg', 6: 'Jabatan\nTMT', 
               7: 'Kelas\nJabatan', 8: 'Tunjangan\nKerja', 9: 'Persen\nPot.', 
               10: 'Nilai\nPot.', 11: 'Jumlah\nPot.', 12: 'Jumlah\ndibayarkan'}
    
    for col, val in headers.items():
        c = ws.cell(row=5, column=col, value=val)
        c.border = border; c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Lebar kolom
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 15
    
    # Isi data
    grand_total = 0
    for i, row_data in enumerate(hasil, 1):
        for c in range(2, 13):
            ws.cell(row=5+i, column=c).border = border
        
        ws.cell(row=5+i, column=2, value=i).alignment = Alignment(horizontal='center')
        ws.cell(row=5+i, column=3, value=row_data['nama'])
        ws.cell(row=5+i, column=4, value=row_data['nip'])
        ws.cell(row=5+i, column=5, value='PNS')
        ws.cell(row=5+i, column=6, value=row_data['jabatan'])
        ws.cell(row=5+i, column=7, value=row_data['class_id'] if row_data['class_id'] else '').alignment = Alignment(horizontal='center')
        
        if row_data['tunjangan'] > 0:
            ws.cell(row=5+i, column=8, value=row_data['tunjangan']).number_format = '#,##0.00'
        if row_data['persen_pot'] > 0:
            ws.cell(row=5+i, column=9, value=row_data['persen_pot']).number_format = '0.00'
        if row_data['nilai_pot'] > 0:
            ws.cell(row=5+i, column=10, value=row_data['nilai_pot']).number_format = '#,##0.00'
        if row_data['jumlah_pot'] > 0:
            ws.cell(row=5+i, column=11, value=row_data['jumlah_pot']).number_format = '#,##0.00'
        if row_data['dibayarkan'] > 0:
            ws.cell(row=5+i, column=12, value=row_data['dibayarkan']).number_format = '#,##0.00'
        
        grand_total += row_data['dibayarkan']
    
    # Baris total
    last_row = 5 + len(hasil) + 1
    for c in range(2, 13):
        ws.cell(row=last_row, column=c).border = border
    
    ws.merge_cells(start_row=last_row, start_column=2, end_row=last_row, end_column=11)
    ws.cell(row=last_row, column=2, value='Total').font = Font(bold=True)
    ws.cell(row=last_row, column=12, value=grand_total).number_format = '#,##0.00'
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
        download_name=f"Rincian_Pembayaran_Tunjangan_{tgl_awal:%Y%m%d}_{tgl_akhir:%Y%m%d}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
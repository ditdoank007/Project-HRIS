# controllers/dashboard_1MasterFileController.py
from sqlite3 import IntegrityError

import requests
from datetime import date, datetime, timedelta
from io import BytesIO
from urllib.parse import quote
from flask import render_template, request, jsonify, session, current_app, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import func, text as sa_text, text
from app import db
from app.models.classModel import MfClass
from app.models.pegawaiModel import Pegawai
from app.models.unitKerjaModel import MfUnitKerja
from app.models.kalenderModel import MfKalender
from app.models.potModel import MfPot
from app.models.jamKerjaModel import MfJamKerja
from app.models.jabatanModel import MfJabatan
from app.models.groupJabatanModel import MfGroupJabatan
from app.models.subGroupJabatanModel import MfSubGroupJabatan
from app.models.loadFingerModel import MfLoadFinger
from app.models.logTransaksiModel import LogTransaksi
from app.models.logTransaksiBackupModel import LogTransaksiBackup
from app.models.joblistModel import MfJoblist
from app.models.jabatanKegiatanModel import MfJabatanKegiatan
from app.models.tunjanganModel import MfTunjangan
from app.models.userAccountModel import UserAccount
from app.models.hrisAuthConfigModel import HrisAuthConfig
from app.models.formModel import MfForm
from app.models.hakAksesFormModel import HakAksesForm
from app.utils.pegawaiSortHelper import sort_pegawai_rows
from app.helpers.masterJamKerjaHelper import (
    get_shift_kerja_definition,
)

GOOGLE_ID_HOLIDAY_CALENDAR_ID = 'id.indonesian#holiday@group.v.calendar.google.com'

LEVEL_LABEL = {0: 'Admin', 1: 'Operator'}
LEVEL_VALUE = {'admin': 0, 'operator': 1}

def get_auth_config():
    config = HrisAuthConfig.query.first()

    if not config:
        return jsonify({
            'success': False,
            'message': 'Konfigurasi login HRIS belum tersedia.'
        }), 404

    return jsonify({
        'success': True,
        'data': config.to_dict()
    })


def save_auth_config():
    data = request.get_json() or {}

    auth_mode = str(data.get('auth_mode', '')).strip().upper()
    sso_server = str(data.get('sso_server', '')).strip() or None
    sso_callback = str(data.get('sso_callback', '')).strip() or None

    if auth_mode not in ('LOCAL', 'SSO'):
        return jsonify({
            'success': False,
            'message': 'Mode Login harus LOCAL atau SSO.'
        }), 400

    if auth_mode == 'SSO' and (not sso_server or not sso_callback):
        return jsonify({
            'success': False,
            'message': 'SSO Server dan SSO Callback wajib diisi untuk mode SSO.'
        }), 400

    config = HrisAuthConfig.query.first()

    if not config:
        return jsonify({
            'success': False,
            'message': 'Konfigurasi login HRIS tidak ditemukan.'
        }), 404

    config.AUTH_MODE = auth_mode

    if auth_mode == 'SSO':
        config.SSO_SERVER = sso_server
        config.SSO_CALLBACK = sso_callback
    else:
        config.SSO_SERVER = None
        config.SSO_CALLBACK = None

    config.UPDATED_AT = datetime.now()

    db.session.commit()

    return jsonify({
        'success': True,
        'message': 'Konfigurasi Master Login berhasil disimpan.',
        'data': config.to_dict()
    })


def master_butir_kegiatan():
    """
    Render halaman Master File Butir Kegiatan.
    Group Jabatan dropdown diisi dari MF_GROUP_JABATAN (server-side render),
    supaya user WAJIB pilih salah satu sebelum tabel Item Kegiatan bisa
    di-refresh (composite key MF_JOBLIST butuh GROUP_JABATAN_ID).
    """
    group_jabatan_list = MfGroupJabatan.query.order_by(
        MfGroupJabatan.GROUP_JABATAN_ID.asc()
    ).all()

    return render_template(
        'pages/dashboard_1/Master File Butir Kegiatan.html',
        group_jabatan_list=group_jabatan_list,
    )

def get_joblist_list():
    """
    Ambil data MF_JOBLIST (Item ID + Diskripsi) untuk tabel di halaman
    Butir Kegiatan, difilter berdasarkan Group Jabatan.

    Query param:
      - group_jabatan_id : WAJIB. Tanpa ini, tabel tidak boleh menampilkan
        apa-apa (mencegah query seluruh MF_JOBLIST tanpa filter).
    """
    group_jabatan_id = request.args.get('group_jabatan_id', type=int)

    if group_jabatan_id is None:
        return jsonify({'status': 'error', 'message': 'Group Jabatan wajib dipilih terlebih dahulu'}), 400

    group_jabatan = MfGroupJabatan.query.get(group_jabatan_id)
    if group_jabatan is None:
        return jsonify({'status': 'error', 'message': 'Group Jabatan tidak ditemukan'}), 404

    rows = (
        MfJoblist.query
        .filter(MfJoblist.GROUP_JABATAN_ID == group_jabatan_id)
        .order_by(MfJoblist.ITEM_ID.asc())
        .all()
    )

    data = [
        {
            'no': idx + 1,
            'item_id': row.ITEM_ID,
            'deskripsi': row.DESKRIPSI or '-',
        }
        for idx, row in enumerate(rows)
    ]

    return jsonify({'status': 'success', 'data': data})

def save_joblist():
    """
    Simpan baris baru ke MF_JOBLIST (Item Kegiatan) untuk sebuah Group Jabatan.

    Body JSON yang diharapkan:
    {
        "group_jabatan_id": 10,
        "item_id": "ITM-001",
        "deskripsi": "Menyusun laporan bulanan"
    }

    Catatan penting:
    ITEM_ID punya FK ke MF_JABATAN_KEGIATAN.ITEM_ID -- jadi ID Kegiatan
    yang diinput HARUS sudah terdaftar di master MF_JABATAN_KEGIATAN
    lebih dulu. Kalau belum ada, insert akan ditolak database (FK
    constraint) dan kita kembalikan pesan yang jelas ke user, bukan
    error 500 mentah.
    """
    payload = request.get_json(silent=True) or {}

    group_jabatan_id_raw = payload.get('group_jabatan_id')
    item_id = (payload.get('item_id') or '').strip()
    deskripsi = (payload.get('deskripsi') or '').strip()

    # --- Validasi field wajib ---
    if group_jabatan_id_raw in (None, ''):
        return jsonify({'status': 'error', 'message': 'Group Jabatan wajib dipilih'}), 400
    if not item_id:
        return jsonify({'status': 'error', 'message': 'ID Kegiatan wajib diisi'}), 400
    if not deskripsi:
        return jsonify({'status': 'error', 'message': 'Deskripsi Kegiatan wajib diisi'}), 400

    # --- Validasi & konversi Group Jabatan ID ---
    try:
        group_jabatan_id = int(group_jabatan_id_raw)
    except (TypeError, ValueError):
        return jsonify({'status': 'error', 'message': 'Group Jabatan ID harus berupa angka'}), 400

    group_jabatan = MfGroupJabatan.query.get(group_jabatan_id)
    if group_jabatan is None:
        return jsonify({'status': 'error', 'message': 'Group Jabatan tidak ditemukan'}), 400

    # --- Pastikan ITEM_ID sudah terdaftar di master MF_JABATAN_KEGIATAN (FK) ---
    jabatan_kegiatan = MfJabatanKegiatan.query.get(item_id)
    if jabatan_kegiatan is None:
        return jsonify({
            'status': 'error',
            'message': f'ID Kegiatan "{item_id}" belum terdaftar di Master Jabatan Kegiatan. '
                       f'Tambahkan dulu di master tersebut sebelum dikaitkan ke Group Jabatan ini.'
        }), 400

    # --- Cegah duplikat composite key (GROUP_JABATAN_ID + ITEM_ID) ---
    existing = MfJoblist.query.get((group_jabatan_id, item_id))
    if existing is not None:
        return jsonify({
            'status': 'error',
            'message': f'ID Kegiatan "{item_id}" sudah terdaftar untuk Group Jabatan ini'
        }), 409

    joblist = MfJoblist(
        GROUP_JABATAN_ID=group_jabatan_id,
        ITEM_ID=item_id,
        DESKRIPSI=deskripsi,
    )

    try:
        db.session.add(joblist)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': 'Gagal menyimpan: referensi data tidak valid (FK constraint)'
        }), 400

    return jsonify({
        'status': 'success',
        'message': 'Item kegiatan berhasil disimpan',
        'data': joblist.to_dict(),
    })

def master_jabatan():
    """
    Render halaman Master File Master Jabatan.
    Group Jabatan & SubGroup Jabatan diisi dari tabel MF_GROUP_JABATAN dan
    MF_SUB_GROUP_JABATAN (server-side render) — bukan hardcode di HTML,
    supaya kalau ada group/subgroup baru, cukup tambah row di DB.
    """
    group_jabatan_list = MfGroupJabatan.query.order_by(
        MfGroupJabatan.GROUP_JABATAN_ID.asc()
    ).all()
    sub_group_jabatan_list = MfSubGroupJabatan.query.order_by(
        MfSubGroupJabatan.SUB_GROUP_JABATAN_ID.asc()
    ).all()

    return render_template(
        'pages/dashboard_1/Master File Master Jabatan.html',
        group_jabatan_list=group_jabatan_list,
        sub_group_jabatan_list=sub_group_jabatan_list,
    )

def save_jabatan():
    """
    Simpan data Master Jabatan baru dari form.
    Body JSON yang diharapkan:
    {
        "group_jabatan_id": 10,
        "sub_group_jabatan_id": 10101010,
        "jabatan_id": 5001,
        "nama_jabatan": "Analis SAR",
        "parent_jabatan_id": 0,      // -> JABATAN_MANAGE (opsional)
        "level_jabatan": 1,          // -> URUT_JABATAN (wajib)
        "type_jabatan": "FT",
        "is_aktif": true             // true=Aktif -> IS_USE=1, false=Non Aktif -> IS_USE=0
    }

    Catatan: JABATAN_ID_BARU (FK ke PERUBAHAN_JABATAN, NOT NULL) belum
    ada field-nya di form ini. Untuk sementara diisi otomatis sama
    dengan jabatan_id (asumsi: belum ada perubahan jabatan). Sesuaikan
    kalau logika sebenarnya berbeda atau tabel PERUBAHAN_JABATAN sudah
    ada modelnya.
    """
    payload = request.get_json(silent=True) or {}

    group_jabatan_id_raw = payload.get('group_jabatan_id')
    sub_group_jabatan_id_raw = payload.get('sub_group_jabatan_id')
    jabatan_id_raw = payload.get('jabatan_id')
    nama_jabatan = (payload.get('nama_jabatan') or '').strip()
    parent_jabatan_id_raw = payload.get('parent_jabatan_id')
    level_jabatan_raw = payload.get('level_jabatan')
    type_jabatan = (payload.get('type_jabatan') or '').strip()
    is_aktif_raw = payload.get('is_aktif')

    # --- Validasi field wajib ---
    if group_jabatan_id_raw in (None, ''):
        return jsonify({'status': 'error', 'message': 'Group Jabatan wajib dipilih'}), 400
    if sub_group_jabatan_id_raw in (None, ''):
        return jsonify({'status': 'error', 'message': 'SubGroup Jabatan wajib dipilih'}), 400
    if jabatan_id_raw in (None, ''):
        return jsonify({'status': 'error', 'message': 'Jabatan ID wajib diisi'}), 400
    if not nama_jabatan:
        return jsonify({'status': 'error', 'message': 'Nama Jabatan wajib diisi'}), 400
    if level_jabatan_raw in (None, ''):
        return jsonify({'status': 'error', 'message': 'Level Jabatan wajib diisi'}), 400
    if not type_jabatan:
        return jsonify({'status': 'error', 'message': 'Type wajib dipilih'}), 400
    if is_aktif_raw is None:
        return jsonify({'status': 'error', 'message': 'Isi Aktif wajib dipilih'}), 400

    # --- Validasi & konversi Jabatan ID ---
    try:
        jabatan_id = int(jabatan_id_raw)
    except (TypeError, ValueError):
        return jsonify({'status': 'error', 'message': 'Jabatan ID harus berupa angka'}), 400

    # Cegah duplikat primary key
    existing = MfJabatan.query.get(jabatan_id)
    if existing is not None:
        return jsonify({
            'status': 'error',
            'message': f'Jabatan ID {jabatan_id} sudah terdaftar ({existing.NAMA_JABATAN})'
        }), 409

    # --- Validasi & konversi Group Jabatan ID (harus ada di MF_GROUP_JABATAN) ---
    try:
        group_jabatan_id = int(group_jabatan_id_raw)
    except (TypeError, ValueError):
        return jsonify({'status': 'error', 'message': 'Group Jabatan ID harus berupa angka'}), 400

    group_jabatan = MfGroupJabatan.query.get(group_jabatan_id)
    if group_jabatan is None:
        return jsonify({'status': 'error', 'message': 'Group Jabatan tidak ditemukan'}), 400

    # --- Validasi & konversi SubGroup Jabatan ID (harus ada di MF_SUB_GROUP_JABATAN) ---
    try:
        sub_group_jabatan_id = int(sub_group_jabatan_id_raw)
    except (TypeError, ValueError):
        return jsonify({'status': 'error', 'message': 'SubGroup Jabatan ID harus berupa angka'}), 400

    sub_group_jabatan = MfSubGroupJabatan.query.get(sub_group_jabatan_id)
    if sub_group_jabatan is None:
        return jsonify({'status': 'error', 'message': 'SubGroup Jabatan tidak ditemukan'}), 400

    # --- Validasi & konversi Parent Jabatan ID (opsional) ---
    parent_jabatan_id = None
    if parent_jabatan_id_raw not in (None, ''):
        try:
            parent_jabatan_id = int(parent_jabatan_id_raw)
        except (TypeError, ValueError):
            return jsonify({'status': 'error', 'message': 'Parent Jabatan ID harus berupa angka'}), 400

    # --- Validasi & konversi Level Jabatan ---
    try:
        level_jabatan = int(level_jabatan_raw)
    except (TypeError, ValueError):
        return jsonify({'status': 'error', 'message': 'Level Jabatan harus berupa angka bulat'}), 400

    # --- Validasi Type ---
    if type_jabatan not in ('FT', 'FU'):
        return jsonify({'status': 'error', 'message': 'Type harus "FT" atau "FU"'}), 400

    # --- Konversi Isi Aktif: Aktif -> 1, Non Aktif -> 0 ---
    is_use = 1 if is_aktif_raw else 0

    jabatan = MfJabatan(
        JABATAN_ID=jabatan_id,
        JABATAN_ID_BARU=jabatan_id,  # asumsi sementara: belum ada perubahan jabatan
        GROUP_JABATAN_ID=group_jabatan_id,
        SUB_GROUP_JABATAN_ID=sub_group_jabatan_id,
        JABATAN_MANAGE=parent_jabatan_id,
        NAMA_JABATAN=nama_jabatan,
        URUT_JABATAN=level_jabatan,
        TYPE_JABATAN=type_jabatan,
        IS_USE=is_use,
        UPDATE_BY=session.get('nip', 'system'),
        UPDATE_DATE=datetime.utcnow(),
    )

    db.session.add(jabatan)
    db.session.commit()

    return jsonify({
        'status': 'success',
        'message': 'Data jabatan berhasil disimpan',
        'data': jabatan.to_dict(),
    })

def master_jam_finger():
    """Render halaman Master File Master Jam Finger."""
    return render_template('pages/dashboard_1/Master File Master Jam Finger.html')

def save_jam_finger():
    """
    Simpan data Master Jam Finger baru dari form.

    LOG_TRANSAKSI dan LOG_TRANSAKSI_BACKUP punya circular foreign key
    (masing-masing mereferensikan satu sama lain), jadi urutan insert
    HARUS:
      1. Insert LOG_TRANSAKSI_BACKUP dulu TANPA mengisi TRAKSAKSI_ID
         (kosongkan/None dulu -- kolom ini nullable).
      2. Insert LOG_TRANSAKSI dengan TRAKSAKSI_BACKUP_ID dari langkah 1.
      3. UPDATE balik LOG_TRANSAKSI_BACKUP.TRAKSAKSI_ID dengan ID dari
         langkah 2 (baru sekarang aman, karena baris LOG_TRANSAKSI
         sudah ada).
      4. Insert MF_LOAD_FINGER dengan TRAKSAKSI_ID dari langkah 2.

    Kedua tabel (LOG_TRANSAKSI, LOG_TRANSAKSI_BACKUP) BUKAN auto_increment
    di database (dikonfirmasi via DESCRIBE), jadi ID dihitung manual
    lewat MAX()+1.
    """
    payload = request.get_json(silent=True) or {}

    shift_raw = (payload.get('shift') or '').strip()
    tgl_mulai_raw = (payload.get('tgl_mulai') or '').strip()
    jam_in_start_raw = (payload.get('jam_in_start') or '').strip()
    jam_in_end_raw = (payload.get('jam_in_end') or '').strip()
    jam_out_start_raw = (payload.get('jam_out_start') or '').strip()
    jam_out_end_raw = (payload.get('jam_out_end') or '').strip()

    if not shift_raw:
        return jsonify({'status': 'error', 'message': 'Shift wajib dipilih'}), 400
    if not tgl_mulai_raw:
        return jsonify({'status': 'error', 'message': 'Tanggal Mulai wajib diisi'}), 400
    if not jam_in_start_raw or not jam_in_end_raw:
        return jsonify({'status': 'error', 'message': 'Rentang Start Jam In wajib diisi'}), 400
    if not jam_out_start_raw or not jam_out_end_raw:
        return jsonify({'status': 'error', 'message': 'Rentang Start Jam Out wajib diisi'}), 400

    try:
        tgl_mulai = datetime.strptime(tgl_mulai_raw, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'status': 'error', 'message': 'Format Tanggal Mulai harus YYYY-MM-DD'}), 400

    def _parse_jam(value):
        try:
            jam = datetime.strptime(value, '%H:%M').time()
            return datetime.combine(tgl_mulai, jam)
        except ValueError:
            return None

    start_finger = _parse_jam(jam_in_start_raw)
    end_finger = _parse_jam(jam_in_end_raw)
    start_finger_out = _parse_jam(jam_out_start_raw)
    end_finger_out = _parse_jam(jam_out_end_raw)

    if start_finger is None or end_finger is None:
        return jsonify({'status': 'error', 'message': 'Format Start Jam In harus HH:MM'}), 400
    if start_finger_out is None or end_finger_out is None:
        return jsonify({'status': 'error', 'message': 'Format Start Jam Out harus HH:MM'}), 400
    if start_finger >= end_finger:
        return jsonify({'status': 'error', 'message': 'Rentang Start Jam In tidak valid'}), 400
    if start_finger_out >= end_finger_out:
        return jsonify({'status': 'error', 'message': 'Rentang Start Jam Out tidak valid'}), 400

    current_nip = session.get('nip')
    if not current_nip:
        return jsonify({'status': 'error', 'message': 'Sesi login tidak valid (NIP tidak ditemukan)'}), 401

    now = datetime.utcnow()

    try:
        # --- Langkah 1: insert LOG_TRANSAKSI_BACKUP TANPA isi TRAKSAKSI_ID dulu ---
        next_backup_id = db.session.query(
            func.coalesce(func.max(LogTransaksiBackup.TRAKSAKSI_BACKUP_ID), 0)
        ).scalar() + 1

        log_transaksi_backup = LogTransaksiBackup(
            TRAKSAKSI_BACKUP_ID=next_backup_id,
            TRAKSAKSI_ID=None,  # dikosongkan dulu -- diisi belakangan setelah LOG_TRANSAKSI ada
            TRANSAKSI='MASTER_JAM_FINGER',
            ACTIVITY='INSERT',
            UPDATE_DATE=now,
        )
        db.session.add(log_transaksi_backup)
        db.session.flush()  # commit sementara insert ini, supaya baris backup benar-benar ada di DB

        # --- Langkah 2: insert LOG_TRANSAKSI dengan TRAKSAKSI_BACKUP_ID dari langkah 1 ---
        next_transaksi_id = db.session.query(
            func.coalesce(func.max(LogTransaksi.TRAKSAKSI_ID), 0)
        ).scalar() + 1

        log_transaksi = LogTransaksi(
            TRAKSAKSI_ID=next_transaksi_id,
            NIP=current_nip,
            TRAKSAKSI_BACKUP_ID=next_backup_id,
            TRANSAKSI='MASTER_JAM_FINGER',
            ACTIVITY='INSERT',
            UPDATE_DATE=now,
        )
        db.session.add(log_transaksi)
        db.session.flush()  # sekarang baris LOG_TRANSAKSI juga benar-benar ada di DB

        # --- Langkah 3: baru sekarang aman untuk update balik TRAKSAKSI_ID ---
        log_transaksi_backup.TRAKSAKSI_ID = next_transaksi_id
        db.session.flush()

        # --- Langkah 4: insert MF_LOAD_FINGER ---
        jam_finger = MfLoadFinger(
            TRAKSAKSI_ID=next_transaksi_id,
            START_FINGER=start_finger,
            END_FINGER=end_finger,
            TGL_MULAI_BERLAKU=tgl_mulai,
            UPDATE_BY=current_nip,
            UPDATE_DATE=now,
            SHIFT_KERJA=shift_raw,
            START_FINGER_OUT=start_finger_out,
            END_FINGER_OUT=end_finger_out,
        )
        db.session.add(jam_finger)
        db.session.commit()

    except Exception:
        db.session.rollback()
        raise

    return jsonify({
        'status': 'success',
        'message': 'Data jam finger berhasil disimpan',
        'data': jam_finger.to_dict(),
    })


def get_jam_finger_by_id():
    """Ambil satu record Master Jam Finger berdasarkan TransaksiID."""
    transaksi_id_raw = request.args.get('transaksi_id', '').strip()

    if not transaksi_id_raw.isdigit():
        return jsonify({
            'status': 'error',
            'message': 'TransaksiID tidak valid'
        }), 400

    transaksi_id = int(transaksi_id_raw)

    row = db.session.get(MfLoadFinger, transaksi_id)

    if row is None:
        return jsonify({
            'status': 'error',
            'message': 'Data Master Jam Finger tidak ditemukan'
        }), 404

    def _time_value(value):
        if value is None:
            return ''

        if isinstance(value, str):
            value = value.strip()

            for fmt in (
                '%Y-%m-%d %H:%M:%S',
                '%H:%M:%S',
                '%H:%M',
            ):
                try:
                    return datetime.strptime(value, fmt).strftime('%H:%M')
                except ValueError:
                    continue

            return ''

        if hasattr(value, 'strftime'):
            return value.strftime('%H:%M')

        return ''

    return jsonify({
        'status': 'success',
        'data': {
            'transaksi_id': row.TRAKSAKSI_ID,
            'shift': row.SHIFT_KERJA or '',
            'tgl_mulai': (
                row.TGL_MULAI_BERLAKU.strftime('%Y-%m-%d')
                if row.TGL_MULAI_BERLAKU and hasattr(row.TGL_MULAI_BERLAKU, 'strftime')
                else ''
            ),
            'jam_in_start': _time_value(row.START_FINGER),
            'jam_in_end': _time_value(row.END_FINGER),
            'jam_out_start': _time_value(row.START_FINGER_OUT),
            'jam_out_end': _time_value(row.END_FINGER_OUT),
        }
    })


def update_jam_finger():
    """Update satu record Master Jam Finger berdasarkan TransaksiID."""
    payload = request.get_json(silent=True) or {}

    transaksi_id_raw = str(payload.get('transaksi_id') or '').strip()
    shift_raw = str(payload.get('shift') or '').strip()
    tgl_mulai_raw = str(payload.get('tgl_mulai') or '').strip()
    jam_in_start_raw = str(payload.get('jam_in_start') or '').strip()
    jam_in_end_raw = str(payload.get('jam_in_end') or '').strip()
    jam_out_start_raw = str(payload.get('jam_out_start') or '').strip()
    jam_out_end_raw = str(payload.get('jam_out_end') or '').strip()

    if not transaksi_id_raw.isdigit():
        return jsonify({'status': 'error', 'message': 'TransaksiID tidak valid'}), 400

    if shift_raw not in ('1', '2'):
        return jsonify({'status': 'error', 'message': 'Shift tidak valid'}), 400

    if not tgl_mulai_raw:
        return jsonify({'status': 'error', 'message': 'Tanggal Mulai wajib diisi'}), 400

    if not jam_in_start_raw or not jam_in_end_raw:
        return jsonify({
            'status': 'error',
            'message': 'Rentang Start Jam In wajib diisi'
        }), 400

    if not jam_out_start_raw or not jam_out_end_raw:
        return jsonify({
            'status': 'error',
            'message': 'Rentang Start Jam Out wajib diisi'
        }), 400

    try:
        transaksi_id = int(transaksi_id_raw)
        tgl_mulai = datetime.strptime(
            tgl_mulai_raw, '%Y-%m-%d'
        ).date()
    except ValueError:
        return jsonify({
            'status': 'error',
            'message': 'Format tanggal atau TransaksiID tidak valid'
        }), 400

    def _parse_jam(value):
        try:
            jam = datetime.strptime(value, '%H:%M').time()
            return datetime.combine(tgl_mulai, jam)
        except ValueError:
            return None

    start_finger = _parse_jam(jam_in_start_raw)
    end_finger = _parse_jam(jam_in_end_raw)
    start_finger_out = _parse_jam(jam_out_start_raw)
    end_finger_out = _parse_jam(jam_out_end_raw)

    if start_finger is None or end_finger is None:
        return jsonify({
            'status': 'error',
            'message': 'Format Start Jam In harus HH:MM'
        }), 400

    if start_finger_out is None or end_finger_out is None:
        return jsonify({
            'status': 'error',
            'message': 'Format Start Jam Out harus HH:MM'
        }), 400

    if start_finger >= end_finger:
        return jsonify({
            'status': 'error',
            'message': 'Rentang Start Jam In tidak valid'
        }), 400

    if start_finger_out >= end_finger_out:
        return jsonify({
            'status': 'error',
            'message': 'Rentang Start Jam Out tidak valid'
        }), 400

    current_nip = session.get('nip')

    if not current_nip:
        return jsonify({
            'status': 'error',
            'message': 'Sesi login tidak valid (NIP tidak ditemukan)'
        }), 401

    row = db.session.get(MfLoadFinger, transaksi_id)

    if row is None:
        return jsonify({
            'status': 'error',
            'message': 'Data Master Jam Finger tidak ditemukan'
        }), 404

    try:
        row.START_FINGER = start_finger
        row.END_FINGER = end_finger
        row.TGL_MULAI_BERLAKU = tgl_mulai
        row.UPDATE_BY = current_nip
        row.UPDATE_DATE = datetime.utcnow()
        row.SHIFT_KERJA = shift_raw
        row.START_FINGER_OUT = start_finger_out
        row.END_FINGER_OUT = end_finger_out

        db.session.commit()

    except Exception:
        db.session.rollback()
        raise

    return jsonify({
        'status': 'success',
        'message': 'Data jam finger berhasil diperbarui',
    })


def delete_jam_finger():
    """Hapus satu record Master Jam Finger berdasarkan TransaksiID."""
    payload = request.get_json(silent=True) or {}

    transaksi_id_raw = str(payload.get('transaksi_id') or '').strip()

    if not transaksi_id_raw.isdigit():
        return jsonify({
            'status': 'error',
            'message': 'TransaksiID tidak valid'
        }), 400

    transaksi_id = int(transaksi_id_raw)

    row = db.session.get(MfLoadFinger, transaksi_id)

    if row is None:
        return jsonify({
            'status': 'error',
            'message': 'Data Master Jam Finger tidak ditemukan'
        }), 404

    try:
        db.session.delete(row)
        db.session.commit()

    except Exception:
        db.session.rollback()
        raise

    return jsonify({
        'status': 'success',
        'message': 'Data jam finger berhasil dihapus',
    })

def master_jam_kerja():
    """Render halaman Master File Master Jam Kerja."""
    return render_template('pages/dashboard_1/Master File Master Jam Kerja.html')

def save_jam_kerja():
    """
    Simpan data Master Jam Kerja baru dari form.
    Body JSON yang diharapkan:
    {
        "shift": "1",                    // SK-1 sampai SK-4
        "penggantian_tlm1": true,        // true=Ada Penggantian -> 'Y', false=Tidak Ada -> 'N'
        "tgl_mulai": "2026-01-01",
        "hari_kerja": "1",                // HK-1=Senin-Kamis, HK-2=Jumat
        "jam_masuk": "HH:MM",
        "jam_pulang": "HH:MM"
    }

    Catatan:
    Jam masuk dan jam pulang berasal dari input Master Jam Kerja
    dan tidak ditentukan secara hardcode oleh SK.
    """
    payload = request.get_json(silent=True) or {}

    shift_raw = payload.get('shift', '').strip() if payload.get('shift') else ''
    penggantian_tlm1_raw = payload.get('penggantian_tlm1')
    tgl_mulai_raw = payload.get('tgl_mulai', '').strip() if payload.get('tgl_mulai') else ''
    hari_kerja_raw = payload.get('hari_kerja', '').strip() if payload.get('hari_kerja') else ''
    jam_masuk_raw = payload.get('jam_masuk', '').strip() if payload.get('jam_masuk') else ''
    jam_pulang_raw = payload.get('jam_pulang', '').strip() if payload.get('jam_pulang') else ''

    # --- Validasi field wajib ---
    if not shift_raw:
        return jsonify({'status': 'error', 'message': 'Shift wajib dipilih'}), 400
    if penggantian_tlm1_raw is None:
        return jsonify({'status': 'error', 'message': 'Penggantian TLM 1 wajib dipilih'}), 400
    if not tgl_mulai_raw:
        return jsonify({'status': 'error', 'message': 'Tanggal Mulai wajib diisi'}), 400
    if not hari_kerja_raw:
        return jsonify({'status': 'error', 'message': 'Hari Kerja wajib dipilih'}), 400
    if not jam_masuk_raw:
        return jsonify({'status': 'error', 'message': 'Jam Masuk wajib diisi'}), 400
    if not jam_pulang_raw:
        return jsonify({'status': 'error', 'message': 'Jam Pulang wajib diisi'}), 400

    # --- Validasi & konversi Tanggal Mulai ---
    try:
        tgl_mulai = datetime.strptime(tgl_mulai_raw, '%Y-%m-%d')
    except ValueError:
        return jsonify({'status': 'error', 'message': 'Format Tanggal Mulai harus YYYY-MM-DD'}), 400

    # --- Validasi & konversi Jam Masuk / Jam Pulang, digabung dengan Tanggal Mulai ---
    try:
        jam_masuk_time = datetime.strptime(jam_masuk_raw, '%H:%M').time()
        std_jam_in = datetime.combine(tgl_mulai.date(), jam_masuk_time)
    except ValueError:
        return jsonify({'status': 'error', 'message': 'Format Jam Masuk harus HH:MM'}), 400

    try:
        jam_pulang_time = datetime.strptime(jam_pulang_raw, '%H:%M').time()
        std_jam_out = datetime.combine(tgl_mulai.date(), jam_pulang_time)
    except ValueError:
        return jsonify({'status': 'error', 'message': 'Format Jam Pulang harus HH:MM'}), 400

    # --- Konversi Penggantian TLM 1: Ada -> 'Y', Tidak Ada -> 'N' ---
    penggantian_tlm1 = 'Y' if penggantian_tlm1_raw else 'N'

    # --- Validasi Hari Kerja ---
    if hari_kerja_raw not in ('1', '2'):
        return jsonify({
            'status': 'error',
            'message': 'Hari Kerja tidak valid'
        }), 400

    # --- Validasi dan mapping Shift Kerja bisnis HK/SK ---
    shift_definition = get_shift_kerja_definition(shift_raw)

    if shift_definition is None:
        return jsonify({
            'status': 'error',
            'message': 'Shift Kerja tidak valid. Pilih SK-1 sampai SK-4.'
        }), 400

    if shift_definition['hari_kerja'] != hari_kerja_raw:
        return jsonify({
            'status': 'error',
            'message': (
                f"Shift Kerja SK-{shift_raw} tidak sesuai dengan "
                f"Hari Kerja HK-{hari_kerja_raw}."
            )
        }), 400

    # --- Mapping kode bisnis SK ke struktur legacy database ---
    shift_db = shift_definition['hari_kerja']

    if shift_raw in ('1', '3'):
        shift_kerja_db = '1'
    else:
        shift_kerja_db = '2'

    # --- Generate IDJKerja karena kolom database bukan AUTO_INCREMENT ---
    max_id = db.session.query(
        db.func.max(MfJamKerja.IDJKERJA)
    ).scalar()
    next_id = (max_id or 0) + 1

    # Agenda legacy MF_JAM_KERJA saat ini selalu kosong.
    agenda = None

    jam_kerja = MfJamKerja(
        IDJKERJA=next_id,
        STD_JAM_IN=std_jam_in,
        STD_JAM_OUT=std_jam_out,
        TGL_MULAI_BERLAKU=tgl_mulai,
        SHIFT=shift_db,
        AGENDA=agenda,
        PENGGANTIAN_TLM1=penggantian_tlm1,
        UPDATE_BY=session.get('nip', 'system'),
        UPDATE_DATE=datetime.now(),
        SHIFT_KERJA=shift_kerja_db,
    )

    db.session.add(jam_kerja)
    db.session.commit()

    return jsonify({
        'status': 'success',
        'message': 'Data jam kerja berhasil disimpan',
        'data': jam_kerja.to_dict(),
    })

def master_kalender():
    """Render halaman Master File Master Kalender."""
    return render_template('pages/dashboard_1/Master File Master Kalender.html')

def _get_indonesian_holidays(tahun):
    """
    Ambil daftar hari libur nasional Indonesia untuk 1 tahun tertentu
    dari Google Calendar (public holiday calendar).
    Return: dict { date(YYYY, M, D): "Nama Hari Libur" }
    """
    api_key = current_app.config.get('GOOGLE_CALENDAR_API_KEY')
    if not api_key:
        # Tanpa API key: kalender tetap dibuat, hanya Sabtu/Minggu yang
        # otomatis ditandai libur — tanggal merah nasional di-skip.
        return {}

    calendar_id = quote(
        GOOGLE_ID_HOLIDAY_CALENDAR_ID,
        safe=''
    )

    url = (
        'https://www.googleapis.com/calendar/v3/calendars/'
        f'{calendar_id}/events'
    )
    params = {
        'key': api_key,
        'timeMin': f'{tahun}-01-01T00:00:00Z',
        'timeMax': f'{tahun}-12-31T23:59:59Z',
        'singleEvents': 'true',
        'orderBy': 'startTime',
    }

    holidays = {}
    try:
        resp = requests.get(url, params=params, timeout=10)
        resp.raise_for_status()
        for event in resp.json().get('items', []):
            tgl_str = event.get('start', {}).get('date')  # all-day event -> 'YYYY-MM-DD'
            if not tgl_str:
                continue
            holidays[date.fromisoformat(tgl_str)] = event.get('summary', 'Hari Libur Nasional')
    except requests.RequestException as e:
        current_app.logger.warning(f'Gagal mengambil data libur nasional dari Google Calendar: {e}')

    return holidays


def create_kalender_tahun():
    """
    Generate seluruh baris KALENDER untuk 1 tahun penuh.
    Body JSON: { "tahun": 2026 }
    Kalau baris tanggal tertentu sudah ada, akan di-update (bukan duplikat).
    """
    payload = request.get_json(silent=True) or {}
    tahun_raw = payload.get('tahun')

    if not tahun_raw or not str(tahun_raw).isdigit():
        return jsonify({'status': 'error', 'message': 'Tahun wajib diisi dan berupa angka'}), 400

    tahun = int(tahun_raw)
    if tahun < 1900 or tahun > 2200:
        return jsonify({'status': 'error', 'message': 'Tahun tidak valid'}), 400

    holidays = _get_indonesian_holidays(tahun)
    hari_nama = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']

    current_nip = session.get('nip', 'system')
    now = datetime.utcnow()

    inserted, updated = 0, 0
    d = date(tahun, 1, 1)
    akhir_tahun = date(tahun, 12, 31)

    while d <= akhir_tahun:
        holiday_name = holidays.get(d)
        is_weekend = d.weekday() >= 5  # 5=Sabtu, 6=Minggu

        if holiday_name:
            is_libur, ket = 'Y', holiday_name
        elif is_weekend:
            is_libur, ket = 'Y', hari_nama[d.weekday()]
        else:
            is_libur, ket = 'N', None

        tgl_kerja = datetime.combine(d, datetime.min.time())
        row = MfKalender.query.get(tgl_kerja)

        if row is None:
            db.session.add(MfKalender(
                TGL_KERJA=tgl_kerja, IS_LIBUR=is_libur, KET=ket,
                UPDATE_BY=current_nip, UPDATE_DATE=now,
            ))
            inserted += 1
        else:
            row.IS_LIBUR = is_libur
            row.KET = ket
            row.UPDATE_BY = current_nip
            row.UPDATE_DATE = now
            updated += 1

        d += timedelta(days=1)

    db.session.commit()

    return jsonify({
        'status': 'success',
        'tahun': tahun,
        'inserted': inserted,
        'updated': updated,
        'holiday_source_available': bool(current_app.config.get('GOOGLE_CALENDAR_API_KEY')),
    })


def save_kalender_changes():
    """
    Simpan perubahan status kalender dari halaman Master Kalender.

    Payload:
    {
        "tahun": 2026,
        "changes": [
            {
                "tanggal": "2026-08-24",
                "status": "KERJA"
            },
            {
                "tanggal": "2026-08-26",
                "status": "LIBUR",
                "keterangan": "Libur Khusus"
            },
            {
                "tanggal": "2026-08-27",
                "status": "WFH"
            }
        ]
    }

    Status yang diperbolehkan:
    - KERJA
    - LIBUR
    - WFH
    """

    payload = request.get_json(silent=True) or {}

    tahun_raw = payload.get('tahun')
    changes = payload.get('changes')

    if not tahun_raw or not str(tahun_raw).isdigit():
        return jsonify({
            'status': 'error',
            'message': 'Tahun wajib diisi dan berupa angka'
        }), 400

    if not isinstance(changes, list):
        return jsonify({
            'status': 'error',
            'message': 'Data perubahan kalender tidak valid'
        }), 400

    tahun = int(tahun_raw)

    if tahun < 1900 or tahun > 2200:
        return jsonify({
            'status': 'error',
            'message': 'Tahun tidak valid'
        }), 400

    current_nip = session.get('nip', 'system')
    now = datetime.utcnow()

    updated = 0

    try:

        for item in changes:

            if not isinstance(item, dict):
                continue

            tanggal_raw = item.get('tanggal')
            status = str(
                item.get('status', '')
            ).upper().strip()

            if not tanggal_raw:
                continue

            try:
                tanggal = date.fromisoformat(
                    tanggal_raw
                )
            except ValueError:
                return jsonify({
                    'status': 'error',
                    'message': f'Tanggal tidak valid: {tanggal_raw}'
                }), 400

            if tanggal.year != tahun:
                return jsonify({
                    'status': 'error',
                    'message': (
                        f'Tanggal {tanggal_raw} '
                        f'tidak sesuai tahun {tahun}'
                    )
                }), 400

            if status not in (
                'KERJA',
                'LIBUR',
                'WFH'
            ):
                return jsonify({
                    'status': 'error',
                    'message': (
                        f'Status kalender tidak valid: {status}'
                    )
                }), 400

            tgl_kerja = datetime.combine(
                tanggal,
                datetime.min.time()
            )

            row = MfKalender.query.get(tgl_kerja)

            if row is None:
                return jsonify({
                    'status': 'error',
                    'message': (
                        f'Data kalender {tanggal_raw} '
                        'belum tersedia. Silakan buat kalender '
                        'untuk tahun tersebut terlebih dahulu.'
                    )
                }), 404

            if status == 'KERJA':

                row.IS_LIBUR = 'N'
                row.KET = None

            elif status == 'WFH':

                row.IS_LIBUR = 'N'
                row.KET = 'WFH'

            elif status == 'LIBUR':

                row.IS_LIBUR = 'Y'

                keterangan = (
                    item.get('keterangan')
                    or 'Libur'
                )

                row.KET = str(
                    keterangan
                )[:50]

            row.UPDATE_BY = current_nip
            row.UPDATE_DATE = now

            updated += 1

        db.session.commit()

    except Exception:

        db.session.rollback()
        raise

    return jsonify({
        'status': 'success',
        'tahun': tahun,
        'updated': updated,
        'message': (
            f'{updated} perubahan kalender berhasil disimpan'
        )
    })


def get_kalender_list():
    """
    Ambil data KALENDER untuk 1 tahun tertentu.
    Query param: tahun (wajib) — sengaja wajib supaya tidak menarik
    seluruh histori kalender sekaligus (bisa ribuan baris).
    """
    tahun = request.args.get('tahun', type=int)
    if not tahun:
        return jsonify({'status': 'error', 'message': 'Parameter tahun wajib diisi'}), 400

    start = datetime(tahun, 1, 1)
    end = datetime(tahun, 12, 31, 23, 59, 59)

    rows = (
        MfKalender.query
        .filter(MfKalender.TGL_KERJA >= start, MfKalender.TGL_KERJA <= end)
        .order_by(MfKalender.TGL_KERJA.asc())
        .all()
    )

    data = [
        {
            'no': idx + 1,
            'tanggal': row.TGL_KERJA.strftime('%d-%m-%Y'),
            'tanggal_iso': row.TGL_KERJA.strftime('%Y-%m-%d'),
            'is_libur': row.IS_LIBUR,
            'ket': row.KET or '-',
            'updated': _format_jam_finger_updated(row.UPDATE_DATE),
        }
        for idx, row in enumerate(rows)
    ]

    return jsonify({'status': 'success', 'data': data})

def master_pegawai_vip():
    """
    Render halaman Master File Master Pegawai VIP.
    Unit Kerja dropdown diisi dari tabel MF_UNIT_KERJA (server-side render),
    bukan hardcode di HTML — supaya kalau ada unit kerja baru, cukup tambah
    row di DB tanpa perlu edit template.
    """
    # Dropdown Unit Kerja menampilkan seluruh Unit Kerja yang aktif.
    unit_kerja_list = (
        MfUnitKerja.query
        .filter(
            MfUnitKerja.IS_USE == 'Y'
        )
        .order_by(
            MfUnitKerja.URUT_REPORT.asc(),
            MfUnitKerja.NAMA_UNIT_KERJA.asc()
        )
        .all()
    )

    return render_template(
        'pages/dashboard_1/Master File Master Pegawai VIP.html',
        unit_kerja_list=unit_kerja_list
    )

def get_pegawai_vip_list():
    """
    Ambil seluruh data pegawai untuk tabel VIP List.

    Sorting mengikuti Single Source of Sorting HRIS Reborn:
        1. Eselon
        2. Urut Jabatan
        3. Class Jabatan descending
        4. NIP ascending

    Filter:
        - unit_kerja_id
        - field1 + keyword1
        - field2 + keyword2
    """

    unit_kerja_id = request.args.get('unit_kerja_id', type=int)

    field1 = request.args.get('field1')
    keyword1 = request.args.get('keyword1', '').strip()

    field2 = request.args.get('field2')
    keyword2 = request.args.get('keyword2', '').strip()

    # ============================================================
    # FIELD FILTER
    # ============================================================

    field_map = {
        'Gol': Pegawai.GOL,
        'Jabatan': MfJabatan.NAMA_JABATAN,
        'Jenis Kelamin': Pegawai.JENIS_KEL,
        'Nama Peg': Pegawai.NAMA,
        'NIP': Pegawai.NIP,
        'No Finger': Pegawai.FINGER_ID,
        'Unit Kerja': MfUnitKerja.NAMA_UNIT_KERJA,
    }

    # ============================================================
    # HRIS REBORN BUSINESS RULE
    #
    # Pegawai menggunakan JABATAN_ID sebagai referensi master.
    #
    # Sumber nama jabatan:
    #   Pegawai.JABATAN_ID
    #          ↓
    #   MF_JABATAN.JABATAN_ID
    #          ↓
    #   MF_JABATAN.NAMA_JABATAN
    #
    # Pegawai.JABATAN adalah field legacy dan tidak digunakan
    # untuk operasional HRIS Reborn.
    # ============================================================

    # Pegawai selalu dihubungkan ke Master Unit Kerja.
    # VIP List hanya menampilkan pegawai dari Unit Kerja yang aktif.
    query = (
        Pegawai.query
        .outerjoin(
            MfJabatan,
            Pegawai.JABATAN_ID == MfJabatan.JABATAN_ID
        )
        .join(
            MfUnitKerja,
            Pegawai.UNIT_KERJA_ID == MfUnitKerja.UNIT_KERJA_ID
        )
        .filter(
            MfUnitKerja.IS_USE == 'Y'
        )
    )

    if unit_kerja_id:
        query = query.filter(
            Pegawai.UNIT_KERJA_ID == unit_kerja_id
        )

    for field, keyword in (
        (field1, keyword1),
        (field2, keyword2),
    ):
        if not field or not keyword:
            continue

        column = field_map.get(field)

        if column is not None:
            query = query.filter(
                column.ilike(f'%{keyword}%')
            )

    # ============================================================
    # HANYA PEGAWAI AKTIF
    #
    # Standar HRIS Reborn:
    #   N = masih aktif
    #   Y = sudah keluar / tidak aktif
    #
    # Master Pegawai VIP hanya menampilkan
    # pegawai dengan IS_KELUAR = N.
    # ============================================================

    query = query.filter(
        Pegawai.IS_KELUAR == 'N'
    )

    # ============================================================
    # AMBIL DATA
    # ============================================================

    pegawai_list = query.all()

    # ============================================================
    # SORTING STANDARD HRIS REBORN
    #
    # Eselon
    # -> Urut Jabatan
    # -> Class Jabatan DESC
    # -> NIP ASC
    # ============================================================

    pegawai_list = sort_pegawai_rows(
        pegawai_list
    )

    # ============================================================
    # SERIALIZE
    #
    # IS_VIP adalah VARCHAR/Y-N di database legacy.
    # Jangan menggunakan bool("0") karena Python akan menghasilkan True.
    # ============================================================

    def is_vip_value(value):
        return str(value or '').strip().upper() in (
            'Y',
            'YES',
            'TRUE',
            '1',
        )

    # ============================================================
    # MASTER JABATAN
    #
    # Pegawai.JABATAN adalah field legacy.
    # Nama jabatan resmi selalu diambil dari MF_JABATAN
    # berdasarkan Pegawai.JABATAN_ID.
    #
    # Jika JABATAN_ID = 0 atau tidak memiliki master,
    # jabatan dikosongkan.
    # ============================================================

    jabatan_ids = {
        p.JABATAN_ID
        for p in pegawai_list
        if p.JABATAN_ID not in (None, 0)
    }

    jabatan_map = {}

    if jabatan_ids:
        jabatan_rows = (
            MfJabatan.query
            .filter(
                MfJabatan.JABATAN_ID.in_(jabatan_ids)
            )
            .all()
        )

        jabatan_map = {
            j.JABATAN_ID: j.NAMA_JABATAN
            for j in jabatan_rows
        }

    data = [
        {
            'no': idx + 1,
            'nip': p.NIP,
            'nama': p.NAMA,
            'jabatan': jabatan_map.get(p.JABATAN_ID),
            'is_vip': is_vip_value(p.IS_VIP),
        }
        for idx, p in enumerate(pegawai_list)
    ]

    return jsonify({
        'status': 'success',
        'data': data
    })



def toggle_pegawai_vip():
    """
    Ubah status VIP satu pegawai.

    Body JSON:
        {
            "nip": "...",
            "is_vip": true|false
        }

    Database legacy menggunakan:
        Y = VIP
        N = bukan VIP
    """

    payload = request.get_json(silent=True) or {}

    nip = str(
        payload.get('nip') or ''
    ).strip()

    is_vip = payload.get('is_vip')

    # ============================================================
    # VALIDASI
    # ============================================================

    if not nip or is_vip is None:
        return jsonify({
            'status': 'error',
            'message': 'nip dan is_vip wajib diisi'
        }), 400

    pegawai = (
        Pegawai.query
        .filter(Pegawai.NIP == nip)
        .first()
    )

    if pegawai is None:
        return jsonify({
            'status': 'error',
            'message': 'Pegawai tidak ditemukan'
        }), 404

    # ============================================================
    # SIMPAN Y / N
    # ============================================================

    pegawai.IS_VIP = 'Y' if bool(is_vip) else 'N'

    pegawai.UPDATE_BY = (
        session.get('nip')
        or 'system'
    )

    pegawai.UPDATE_DATE = datetime.now()

    db.session.commit()

    return jsonify({
        'status': 'success',
        'nip': pegawai.NIP,
        'is_vip': pegawai.IS_VIP == 'Y'
    })



def master_potongan():
    """Render halaman Master File Master Potongan."""
    return render_template('pages/dashboard_1/Master File Master Potongan.html')

def save_potongan():
    """
    Simpan data Master Potongan baru dari form.
    Body JSON yang diharapkan:
    {
        "kategori": "Cuti",
        "tingkat": "Ringan",
        "diskripsi": "Terlambat 1-15 menit",
        "persen_pot": 1.5,
        "is_pendukung": true,      // true=Ada -> 'Y', false=Tidak Ada -> 'N'
        "tgl_mulai": "2026-01-01",
        "range_awal": 1,
        "range_akhir": 15
    }
    """
    payload = request.get_json(silent=True) or {}

    kategori = payload.get('kategori', '').strip()
    tingkat = payload.get('tingkat', '').strip()
    diskripsi = payload.get('diskripsi', '').strip()
    persen_pot_raw = payload.get('persen_pot')
    is_pendukung_raw = payload.get('is_pendukung')
    tgl_mulai_raw = payload.get('tgl_mulai', '').strip()
    range_awal_raw = payload.get('range_awal')
    range_akhir_raw = payload.get('range_akhir')

    # --- Validasi field wajib ---
    if not kategori:
        return jsonify({'status': 'error', 'message': 'Kategori wajib diisi'}), 400
    if not diskripsi:
        return jsonify({'status': 'error', 'message': 'Diskripsi wajib diisi'}), 400
    if is_pendukung_raw is None:
        return jsonify({'status': 'error', 'message': 'Bukti Pendukung wajib dipilih'}), 400

    # --- Validasi & konversi tipe data numerik ---
    persen_pot = None
    if persen_pot_raw not in (None, ''):
        try:
            persen_pot = float(persen_pot_raw)
            if not (0 <= persen_pot <= 100):
                return jsonify({'status': 'error', 'message': 'Potongan (%) harus antara 0-100'}), 400
        except (TypeError, ValueError):
            return jsonify({'status': 'error', 'message': 'Potongan (%) harus berupa angka'}), 400

    range_awal = None
    if range_awal_raw not in (None, ''):
        try:
            range_awal = float(range_awal_raw)
        except (TypeError, ValueError):
            return jsonify({'status': 'error', 'message': 'Range awal harus berupa angka'}), 400

    range_akhir = None
    if range_akhir_raw not in (None, ''):
        try:
            range_akhir = float(range_akhir_raw)
        except (TypeError, ValueError):
            return jsonify({'status': 'error', 'message': 'Range akhir harus berupa angka'}), 400

    if range_awal is not None and range_akhir is not None and range_awal > range_akhir:
        return jsonify({'status': 'error', 'message': 'Range awal tidak boleh lebih besar dari range akhir'}), 400

    # --- Validasi & konversi tanggal ---
    tgl_mulai = None
    if tgl_mulai_raw:
        try:
            tgl_mulai = datetime.strptime(tgl_mulai_raw, '%Y-%m-%d')
        except ValueError:
            return jsonify({'status': 'error', 'message': 'Format tanggal harus YYYY-MM-DD'}), 400

    # --- Konversi Bukti Pendukung: Ada -> 'Y', Tidak Ada -> 'N' ---
    is_pendukung = 'Y' if is_pendukung_raw else 'N'

    potongan = MfPot(
        KATEGORI=kategori,
        TINGKAT=tingkat or None,
        NAMA_POT=diskripsi,
        PERSEN_POT=persen_pot,
        IS_PENDUKUNG=is_pendukung,
        TGL_MULAI=tgl_mulai,
        RANGE_AWAL=range_awal,
        RANGE_AKHIR=range_akhir,
        UPDATE_BY=session.get('nip', 'system'),
        UPDATE_DATE=datetime.utcnow(),
    )

    db.session.add(potongan)
    db.session.commit()

    return jsonify({
        'status': 'success',
        'message': 'Data potongan berhasil disimpan',
        'data': potongan.to_dict(),
    })

def get_potongan_list():
    """
    Ambil data Master Potongan untuk tabel Cari Master Potongan.

    Filter opsional (semua bisa kosong -> berlaku seperti klik Refresh biasa):
      - periode        : filter TGL_MULAI pada tanggal tertentu (format YYYY-MM-DD)
      - field1/keyword1 dan field2/keyword2 : dua dropdown "Filter" (Kategori,
        Potongan(%), Tingkat), digabung dengan AND
    """
    periode_raw = request.args.get('periode', '').strip()
    field1 = request.args.get('field1')
    keyword1 = request.args.get('keyword1', '').strip()
    field2 = request.args.get('field2')
    keyword2 = request.args.get('keyword2', '').strip()

    query = MfPot.query

    # --- Filter Periode: cocokkan TGL_MULAI pada tanggal yang dipilih ---
    if periode_raw:
        try:
            periode_date = datetime.strptime(periode_raw, '%Y-%m-%d')
        except ValueError:
            return jsonify({'status': 'error', 'message': 'Format periode harus YYYY-MM-DD'}), 400

        awal_hari = periode_date
        akhir_hari = periode_date + timedelta(days=1)
        query = query.filter(MfPot.TGL_MULAI >= awal_hari, MfPot.TGL_MULAI < akhir_hari)

    # --- Filter field1/field2 (Kategori, Potongan(%), Tingkat) ---
    # Kategori & Tingkat -> kolom teks, pakai partial match (ilike)
    # Potongan(%) -> kolom Float, tidak bisa ilike -> exact match angka
    text_field_map = {
        'Kategori': MfPot.KATEGORI,
        'Tingkat': MfPot.TINGKAT,
    }

    for field, keyword in [(field1, keyword1), (field2, keyword2)]:
        if not field or not keyword:
            continue  # filter tidak dipakai -> skip, tidak wajib diisi

        if field == 'Potongan(%)':
            try:
                nilai = float(keyword)
            except ValueError:
                return jsonify({'status': 'error', 'message': 'Potongan(%) harus berupa angka'}), 400
            query = query.filter(MfPot.PERSEN_POT == nilai)
        else:
            column = text_field_map.get(field)
            if column is not None:
                query = query.filter(column.ilike(f'%{keyword}%'))

    pot_list = query.order_by(MfPot.UPDATE_DATE.desc()).all()

    def format_range(row):
        if row.RANGE_AWAL is None and row.RANGE_AKHIR is None:
            return '-'
        awal = row.RANGE_AWAL if row.RANGE_AWAL is not None else '-'
        akhir = row.RANGE_AKHIR if row.RANGE_AKHIR is not None else '-'
        return f'{awal} s/d {akhir}'

    data = [
        {
            'no': idx + 1,
            'kategori': row.KATEGORI or '-',
            'tingkat': row.TINGKAT or '-',
            'deskripsi': row.NAMA_POT or '-',
            'persen_pot': row.PERSEN_POT if row.PERSEN_POT is not None else '-',
            'range': format_range(row),
            'tgl_mulai': row.TGL_MULAI.strftime('%d-%m-%Y') if row.TGL_MULAI else '-',
            'updated': _format_jam_finger_updated(row.UPDATE_DATE),
        }
        for idx, row in enumerate(pot_list)
    ]

    return jsonify({'status': 'success', 'data': data})

def master_trt():
    """Render halaman Master File Master TRT."""
    return render_template('pages/dashboard_1/Master File Master TRT.html')

def master_tunkin_class():
    """
    Render halaman Master File Master Tunkin Class.
    Dropdown Class diisi dari data MF_CLASS yang sudah ada di DB
    (fix 14 baris, CLASS_ID 1-14) — bukan hardcode di HTML, karena
    fitur ini hanya untuk EDIT data yang sudah ada, bukan insert baru.
    """
    class_list = MfClass.query.order_by(MfClass.CLASS_ID.asc()).all()
    return render_template(
        'pages/dashboard_1/Master File Master Tunkin Class.html',
        class_list=class_list,
    )

def get_tunkin_class_detail(class_id):
    """
    Ambil detail 1 baris MF_CLASS berdasarkan CLASS_ID, dipakai untuk
    autofill form saat user memilih Class di dropdown.
    """
    row = MfClass.query.get(class_id)
    if row is None:
        return jsonify({'status': 'error', 'message': 'Class tidak ditemukan'}), 404

    return jsonify({
        'status': 'success',
        'data': {
            'class_id': row.CLASS_ID,
            'tunjangan': row.TUNJANGAN,
            'tgl_mulai': row.TGL_MULAI.strftime('%Y-%m-%d') if row.TGL_MULAI else '',
            'dokreff': row.DOKREFF or '',
        }
    })

def save_tunkin_class():
    """
    Update data Master Tunkin/Class yang SUDAH ADA (edit-only).
    Data MF_CLASS bersifat fix (14 baris, CLASS_ID 1-14) sesuai PERBAN
    NO 4 TAHUN 2024 — form ini tidak membuat baris baru, hanya mengubah
    isi (TUNJANGAN, TGL_MULAI, DOKREFF) dari Class yang dipilih.

    Body JSON yang diharapkan:
    {
        "class_id": 3,
        "tunjangan": 1500000,
        "tgl_mulai": "2026-01-01",
        "dokreff": "SE-001/2026"
    }
    """
    payload = request.get_json(silent=True) or {}

    class_id_raw = payload.get('class_id')
    tunjangan_raw = payload.get('tunjangan')
    tgl_mulai_raw = payload.get('tgl_mulai', '').strip() if payload.get('tgl_mulai') else ''
    dokreff = (payload.get('dokreff') or '').strip()

    # --- Validasi field wajib ---
    if class_id_raw in (None, ''):
        return jsonify({'status': 'error', 'message': 'Class wajib dipilih'}), 400
    if tunjangan_raw in (None, ''):
        return jsonify({'status': 'error', 'message': 'Tunjangan wajib diisi'}), 400
    if not dokreff:
        return jsonify({'status': 'error', 'message': 'No Surat wajib diisi'}), 400

    # --- Validasi & konversi Class ID ---
    try:
        class_id = int(class_id_raw)
    except (TypeError, ValueError):
        return jsonify({'status': 'error', 'message': 'Class harus berupa angka'}), 400

    # --- Wajib sudah ada — form ini edit-only, tidak boleh buat baris baru ---
    existing = MfClass.query.get(class_id)
    if existing is None:
        return jsonify({
            'status': 'error',
            'message': f'Class ID {class_id} tidak ditemukan. Data Master Tunkin/Class bersifat tetap (1-14), tidak bisa membuat Class baru.'
        }), 404

    # --- Validasi & konversi Tunjangan ---
    try:
        tunjangan = float(tunjangan_raw)
        if tunjangan < 0:
            return jsonify({'status': 'error', 'message': 'Tunjangan tidak boleh negatif'}), 400
    except (TypeError, ValueError):
        return jsonify({'status': 'error', 'message': 'Tunjangan harus berupa angka'}), 400

    # --- Validasi & konversi Tanggal (opsional) ---
    tgl_mulai = existing.TGL_MULAI  # default: pertahankan nilai lama kalau tidak diisi
    if tgl_mulai_raw:
        try:
            tgl_mulai = datetime.strptime(tgl_mulai_raw, '%Y-%m-%d')
        except ValueError:
            return jsonify({'status': 'error', 'message': 'Format tanggal harus YYYY-MM-DD'}), 400

    # --- Update baris yang sudah ada ---
    existing.TUNJANGAN = tunjangan
    existing.TGL_MULAI = tgl_mulai
    existing.DOKREFF = dokreff
    existing.UPDATE_IN_BY = session.get('nip', 'system')
    existing.UPDATE_DATE = datetime.utcnow()

    db.session.commit()

    return jsonify({
        'status': 'success',
        'message': f'Data Tunkin/Class {class_id} berhasil diperbarui',
        'data': existing.to_dict(),
    })

def master_unit_kerja():
    """Render halaman Master File Master Unit Kerja."""
    return render_template('pages/dashboard_1/Master File Master Unit Kerja.html')

def save_unit_kerja():
    """
    Simpan data Master Unit Kerja baru dari form.
    Body JSON yang diharapkan:
    {
        "unit_kerja_id": 20,
        "nama_unit_kerja": "Surabaya",
        "urut_report": 1,
        "tipe": "Pusat"   // "Pusat" -> IS_PUSAT=1, "Pos" -> IS_PUSAT=0
    }

    Catatan: field "Isi Aktif" di form saat ini TIDAK punya kolom
    yang sesuai di model MfUnitKerja (hanya ada IS_PUSAT untuk Tipe),
    jadi nilai itu diterima tapi tidak disimpan. Kalau memang perlu
    disimpan, tambahkan kolom IS_AKTIF ke model dulu.
    """
    payload = request.get_json(silent=True) or {}

    unit_kerja_id_raw = payload.get('unit_kerja_id')
    nama_unit_kerja = payload.get('nama_unit_kerja', '').strip()
    urut_report_raw = payload.get('urut_report')
    tipe_raw = payload.get('tipe', '').strip()

    # --- Validasi field wajib ---
    if unit_kerja_id_raw in (None, ''):
        return jsonify({'status': 'error', 'message': 'Unit Kerja ID wajib diisi'}), 400
    if not nama_unit_kerja:
        return jsonify({'status': 'error', 'message': 'Nama Unit Kerja wajib diisi'}), 400
    if not tipe_raw:
        return jsonify({'status': 'error', 'message': 'Tipe wajib dipilih'}), 400

    # --- Validasi & konversi Unit Kerja ID ---
    try:
        unit_kerja_id = int(unit_kerja_id_raw)
    except (TypeError, ValueError):
        return jsonify({'status': 'error', 'message': 'Unit Kerja ID harus berupa angka'}), 400

    # Cegah duplikat primary key — kalau sudah ada, tolak (bukan overwrite diam-diam)
    existing = MfUnitKerja.query.get(unit_kerja_id)
    if existing is not None:
        return jsonify({
            'status': 'error',
            'message': f'Unit Kerja ID {unit_kerja_id} sudah terdaftar ({existing.NAMA_UNIT_KERJA})'
        }), 409

    # --- Validasi & konversi Urut Report ---
    urut_report = 0
    if urut_report_raw not in (None, ''):
        try:
            urut_report = int(urut_report_raw)
        except (TypeError, ValueError):
            return jsonify({'status': 'error', 'message': 'Urut Report harus berupa angka bulat'}), 400
        if urut_report < 0:
            return jsonify({'status': 'error', 'message': 'Urut Report tidak boleh negatif'}), 400

    # --- Konversi Tipe: Pusat -> 1, Pos -> 0 ---
    tipe_map = {'Pusat': 1, 'Pos': 2}
    if tipe_raw not in tipe_map:
        return jsonify({'status': 'error', 'message': 'Tipe harus "Pusat" atau "Pos"'}), 400
    is_pusat = tipe_map[tipe_raw]

    unit_kerja = MfUnitKerja(
        UNIT_KERJA_ID=unit_kerja_id,
        NAMA_UNIT_KERJA=nama_unit_kerja,
        URUT_REPORT=urut_report,
        IS_PUSAT=is_pusat,
        UPDATE_IN_BY=session.get('nip', 'system'),
        UPDATE_DATE=datetime.utcnow(),
    )

    db.session.add(unit_kerja)
    db.session.commit()

    return jsonify({
        'status': 'success',
        'message': 'Data unit kerja berhasil disimpan',
        'data': unit_kerja.to_dict(),
    })


def toggle_unit_kerja():
    """
    Mengubah status penggunaan Unit Kerja.

    MF_UNIT_KERJA.isUse:
        Y = Aktif / digunakan HRIS
        N = Nonaktif / tidak digunakan HRIS

    Tidak mengubah struktur database.
    """

    payload = request.get_json(silent=True) or {}

    unit_kerja_id = str(
        payload.get('unit_kerja_id') or ''
    ).strip()

    is_aktif = str(
        payload.get('is_aktif')
        if payload.get('is_aktif') is not None
        else payload.get('is_use')
        or ''
    ).strip().upper()

    if not unit_kerja_id:
        return jsonify({
            'status': 'error',
            'message': 'Unit Kerja ID wajib diisi'
        }), 400

    if is_aktif not in ('Y', 'N'):
        return jsonify({
            'status': 'error',
            'message': 'Status Unit Kerja hanya boleh Y atau N'
        }), 400

    unit = (
        MfUnitKerja.query
        .filter(MfUnitKerja.UNIT_KERJA_ID == unit_kerja_id)
        .first()
    )

    if unit is None:
        return jsonify({
            'status': 'error',
            'message': f'Unit Kerja {unit_kerja_id} tidak ditemukan'
        }), 404

    unit.IS_AKTIF = is_aktif
    unit.UPDATE_BY = session.get('nip', 'system')
    unit.UPDATE_DATE = datetime.utcnow()

    db.session.commit()

    return jsonify({
        'status': 'success',
        'message': (
            f'Unit {unit.NAMA_UNIT_KERJA} '
            f'berhasil di{"aktifkan" if is_aktif == "Y" else "nonaktifkan"}'
        ),
        'data': {
            'unit_kerja_id': unit.UNIT_KERJA_ID,
            'nama_unit_kerja': unit.NAMA_UNIT_KERJA,
            'is_aktif': unit.IS_AKTIF,
            'is_use': unit.IS_USE,
        }
    })


def master_user():
    """Render halaman Master File Master User."""
    return render_template('pages/dashboard_1/Master File Master User.html')

def _get_operator_forms(nip):
    """
    Bangun daftar form untuk tabel "Khusus Operator", meniru query UNION
    di VB.NET Filldata():
      1. Form yang SUDAH punya baris HAK_AKSES_FORM untuk NIP ini
         (Modul='HRIS', join ke MF_FORM dengan MODEL=2) -> tampilkan
         is_akses & type_akses apa adanya.
      2. Form yang BELUM punya baris HAK_AKSES_FORM untuk NIP ini,
         DIBATASI hanya FORM_TYPE 'Transaksi' atau 'Report'
         -> default is_akses='N', type_akses='M'.

    Diurutkan: is_akses desc dulu (yang aktif duluan), lalu nama form.
    """
    user_account = UserAccount.query.filter(
        UserAccount.NIP == nip,
        UserAccount.MODUL == 'HRIS'
    ).first()

    is_admin = (
        user_account is not None
        and user_account.INIT_LEVEL == 0
    )

    existing_access = {
        row.FORM_ID: row
        for row in HakAksesForm.query.filter(
            HakAksesForm.NIP == nip,
            db.or_(
                HakAksesForm.MODUL == 'HRIS',
                db.and_(
                    HakAksesForm.MODUL == 'eDoc',
                    HakAksesForm.FORM_ID.in_({
                        'KehadiranPiket.aspx',
                        'InJadwalSiaga.aspx',
                        'ReJadwalSiaga.aspx',
                        'Rekapsiaga.aspx',
                        'TTUPiket.aspx',
                        'MFTimSiaga.aspx',
                        'MFTunjPiket.aspx',
                        'MFKGR.aspx',
                        'MFEmail.aspx',
                        'DaftarLemburSiaga.aspx',
                    })
                )
            )
        ).all()
    }

    # ============================================================
    # HRIS REBORN
    #
    # Account User di Bagian Umum menjadi pusat otorisasi seluruh
    # menu HRIS Reborn.
    #
    # Menu Siaga legacy masih tercatat pada MODUL='eDoc', sehingga
    # ikut ditampilkan di Account User tanpa mengubah database.
    # ============================================================
    SIAGA_FORM_IDS = {
        'KehadiranPiket.aspx',
        'InJadwalSiaga.aspx',
        'ReJadwalSiaga.aspx',
        'Rekapsiaga.aspx',
        'TTUPiket.aspx',
        'MFTimSiaga.aspx',
        'MFTunjPiket.aspx',
        'MFKGR.aspx',
        'MFEmail.aspx',
        'DaftarLemburSiaga.aspx',
    }

    all_forms = MfForm.query.filter(
        db.or_(
            db.and_(
                MfForm.MODUL == 'HRIS',
                MfForm.MODEL == 2
            ),
            MfForm.FORM_ID.in_(SIAGA_FORM_IDS)
        )
    ).all()

    result = []
    for form in all_forms:
        access = existing_access.get(form.FORM_ID)

        if is_admin:
            # Administrator selalu memiliki akses penuh.
            # Tidak bergantung HAK_AKSES_FORM.
            result.append({
                'form_id': form.FORM_ID,
                'form_name': form.FORM_NAME or '-',
                'form_type': form.FORM_TYPE or '-',
                'is_akses': True,
                'type_akses': 'M',
            })

        elif access is not None:
            # Bagian 1: sudah punya baris akses -> tampil apa adanya
            result.append({
                'form_id': form.FORM_ID,
                'form_name': form.FORM_NAME or '-',
                'form_type': form.FORM_TYPE or '-',
                'is_akses': access.IS_AKSES == 'Y',
                'type_akses': access.TYPE_AKSES or 'M',
            })
        elif form.FORM_TYPE in ('Transaksi', 'Report'):
            # Administrator HRIS (INIT_LEVEL=0)
            # selalu mempunyai seluruh akses.
            #
            # Checkbox hanya representasi visual,
            # bukan sumber authorization administrator.

            result.append({
                'form_id': form.FORM_ID,
                'form_name': form.FORM_NAME or '-',
                'form_type': form.FORM_TYPE or '-',
                'is_akses': True if is_admin else False,
                'type_akses': 'M',
            })
        # Form lain (bukan Transaksi/Report, belum punya akses) -> di-skip,
        # sesuai logic VB.NET yang tidak menampilkannya sama sekali.

    # ============================================================
    # Urutan Account User:
    #
    # 1. UMUM / HRIS        -> paling atas
    # 2. OPERASI / SIAGA    -> di bawah
    #
    # Di dalam kelompok:
    #   - akses aktif lebih dahulu
    #   - kemudian berdasarkan nama form
    #
    # Menu Siaga legacy berada pada MODUL='eDoc', tetapi tetap
    # dikelola dari Account User HRIS.
    # ============================================================
    operasi_siaga_ids = {
        'KehadiranPiket.aspx',
        'InJadwalSiaga.aspx',
        'ReJadwalSiaga.aspx',
        'Rekapsiaga.aspx',
        'TTUPiket.aspx',
        'MFTimSiaga.aspx',
        'MFTunjPiket.aspx',
        'MFKGR.aspx',
        'MFEmail.aspx',
        'DaftarLemburSiaga.aspx',
    }

    result.sort(
        key=lambda x: (
            1 if x['form_id'] in operasi_siaga_ids else 0,
            not x['is_akses'],
            x['form_name'].lower(),
        )
    )

    return result

def get_user_account_detail():
    """
    Cari data user account berdasarkan NIP -- dipakai oleh tombol
    "Cari Peg." dan "Refresh" di halaman Master User.

    Sekarang termasuk daftar form "Khusus Operator" (dari MF_FORM +
    HAK_AKSES_FORM), meniru logic Filldata() di VB.NET.

    Query param:
      - nip : wajib
    """
    nip = request.args.get('nip', '').strip()

    if not nip:
        return jsonify({'status': 'error', 'message': 'NIP wajib diisi'}), 400

    # Account User HRIS hanya berlaku untuk pegawai yang berada
    # pada Unit Kerja yang masih aktif.
    pegawai = (
        db.session.query(Pegawai)
        .join(
            MfUnitKerja,
            Pegawai.UNIT_KERJA_ID == MfUnitKerja.UNIT_KERJA_ID
        )
        .filter(
            Pegawai.NIP == nip,
            MfUnitKerja.IS_AKTIF == 'Y'
        )
        .first()
    )

    if pegawai is None:
        return jsonify({
            'status': 'error',
            'message': f'Pegawai dengan NIP {nip} tidak ditemukan atau Unit Kerjanya tidak aktif'
        }), 404

    user_account = UserAccount.query.filter(
        UserAccount.NIP == nip,
        UserAccount.MODUL == 'HRIS'
    ).first()

    operator_forms = _get_operator_forms(nip)

    return jsonify({
        'status': 'success',
        'data': {
            'nip': pegawai.NIP,
            'nama': pegawai.NAMA,
            'init_level': user_account.INIT_LEVEL if user_account else 1,
            'level_label': LEVEL_LABEL.get(
                user_account.INIT_LEVEL if user_account else 1, 'Operator'
            ),
            'has_account': user_account is not None,
            'operator_forms': operator_forms,
        }
    })

def save_user_account():
    """
    Simpan/update level akses (Operator/Admin) DAN hak akses per-form
    (Khusus Operator) seorang pegawai.

    Body JSON yang diharapkan:
    {
        "nip": "1985...",
        "level": "operator",
        "akses_list": [
            { "form_id": "F001", "type_akses": "M" },
            { "form_id": "F002", "type_akses": "R" }
        ]
    }

    "akses_list" hanya berisi form yang checkbox-nya DICENTANG di UI --
    form yang tidak dicentang tidak perlu dikirim sama sekali.

    Meniru pola VB.NET BtnSave_Click(): hapus SEMUA baris HAK_AKSES_FORM
    milik NIP ini dulu, baru insert ulang baris baru untuk form yang
    dicentang. Untuk USER_ACCOUNT (Level), tetap pakai pola upsert
    seperti sebelumnya.
    """
    payload = request.get_json(silent=True) or {}

    nip = (payload.get('nip') or '').strip()
    level_raw = (payload.get('level') or '').strip().lower()
    akses_list = payload.get('akses_list') or []

    if not nip:
        return jsonify({'status': 'error', 'message': 'NIP wajib diisi'}), 400
    if level_raw not in LEVEL_VALUE:
        return jsonify({'status': 'error', 'message': 'Level harus "operator" atau "admin"'}), 400
    if not isinstance(akses_list, list):
        return jsonify({'status': 'error', 'message': 'akses_list harus berupa list'}), 400

    pegawai = Pegawai.query.filter(Pegawai.NIP == nip).first()
    if pegawai is None:
        return jsonify({'status': 'error', 'message': f'Pegawai dengan NIP {nip} tidak ditemukan'}), 404

    # --- Validasi setiap entri akses_list sebelum ada perubahan apapun ke DB ---
    #
    # Existing HAK_AKSES_FORM dapat berisi legacy FormID yang tidak lagi
    # memiliki pasangan MF_FORM Model 2. Permission legacy tersebut tetap
    # boleh dipertahankan.
    #
    # Form baru wajib memiliki definisi MF_FORM HRIS Model 2.

    valid_type_akses = ('M', 'R')
    cleaned_akses = []

    existing_access_ids = {
        row.FORM_ID
        for row in HakAksesForm.query.filter(
            HakAksesForm.NIP == nip,
            HakAksesForm.MODUL == 'HRIS'
        ).all()
    }

    for item in akses_list:
        if not isinstance(item, dict):
            return jsonify({
                'status': 'error',
                'message': 'Setiap item akses_list harus berupa object'
            }), 400

        form_id = (item.get('form_id') or '').strip()
        type_akses = (item.get('type_akses') or 'M').strip().upper()

        if not form_id:
            return jsonify({
                'status': 'error',
                'message': 'form_id wajib diisi pada setiap item akses_list'
            }), 400

        if type_akses not in valid_type_akses:
            return jsonify({
                'status': 'error',
                'message': (
                    f'type_akses harus "M" (Modify) atau "R" (Read Only), '
                    f'diterima: "{type_akses}"'
                )
            }), 400

        # Permission lama/legacy milik user boleh dipertahankan walaupun
        # FormID tersebut sudah tidak mempunyai definisi MF_FORM Model 2.
        if form_id in existing_access_ids:
            cleaned_akses.append({
                'form_id': form_id,
                'type_akses': type_akses
            })
            continue

        # Permission baru wajib mempunyai definisi menu HRIS Model 2.
        # Form HRIS Model 2 adalah menu utama HRIS.
        # Form Siaga legacy tertentu tetap boleh dikelola dari
        # Account User Umum walaupun MODUL legacy-nya adalah eDoc.
        SIAGA_FORM_IDS = {
            'KehadiranPiket.aspx',
            'InJadwalSiaga.aspx',
            'ReJadwalSiaga.aspx',
            'Rekapsiaga.aspx',
            'TTUPiket.aspx',
            'MFTimSiaga.aspx',
            'MFTunjPiket.aspx',
            'MFKGR.aspx',
            'MFEmail.aspx',
            'DaftarLemburSiaga.aspx',
        }

        form_query = MfForm.query.filter(
            MfForm.FORM_ID == form_id
        )

        if form_id in SIAGA_FORM_IDS:
            form = form_query.filter(
                MfForm.MODUL == 'eDoc'
            ).first()
        else:
            form = form_query.filter(
                MfForm.MODUL == 'HRIS',
                MfForm.MODEL == 2
            ).first()

        if form is None:
            return jsonify({
                'status': 'error',
                'message': (
                    f'Form HRIS Model 2 dengan ID "{form_id}" '
                    f'tidak ditemukan'
                )
            }), 400

        cleaned_akses.append({
            'form_id': form_id,
            'type_akses': type_akses
        })

    init_level = LEVEL_VALUE[level_raw]
    now = datetime.utcnow()
    current_nip_login = session.get('nip', 'system')

    try:
        # --- Bagian 1: upsert USER_ACCOUNT (Level) ---
        user_account = UserAccount.query.filter(
            UserAccount.NIP == nip,
            UserAccount.MODUL == 'HRIS'
        ).first()

        if user_account is None:
            user_account = UserAccount(
                NIP=nip,
                INIT_LEVEL=init_level,
                MODUL='HRIS',
                UPDATE_BY=current_nip_login,
                UPDATE_DATE=now,
            )
            db.session.add(user_account)
        else:
            user_account.INIT_LEVEL = init_level
            user_account.MODUL = 'HRIS'
            user_account.UPDATE_BY = current_nip_login
            user_account.UPDATE_DATE = now

        # --- Bagian 2: HAK_AKSES_FORM hanya untuk Operator ---
        #
        # Administrator (INIT_LEVEL=0):
        #   tidak membutuhkan row HAK_AKSES_FORM.
        #
        # Operator (INIT_LEVEL>0):
        #   mengikuti checkbox menu.
        #
        if init_level > 0:

            HakAksesForm.query.filter(
                HakAksesForm.NIP == nip,
                db.or_(
                    HakAksesForm.MODUL == 'HRIS',
                    db.and_(
                        HakAksesForm.MODUL == 'eDoc',
                        HakAksesForm.FORM_ID.in_(SIAGA_FORM_IDS)
                    )
                )
            ).delete(
                synchronize_session=False
            )

            for item in cleaned_akses:

                modul_form = (
                    'eDoc'
                    if item['form_id'] in SIAGA_FORM_IDS
                    else 'HRIS'
                )

                db.session.add(HakAksesForm(
                    NIP=nip,
                    FORM_ID=item['form_id'],
                    IS_AKSES='Y',
                    TYPE_AKSES=item['type_akses'],
                    MODUL=modul_form,
                    UPDATE_BY=current_nip_login,
                    UPDATE_DATE=now,
                ))

        db.session.commit()

    except Exception:
        db.session.rollback()
        raise

    return jsonify({
        'status': 'success',
        'message': f'Data akun {pegawai.NAMA} berhasil disimpan ({len(cleaned_akses)} hak akses form)',
        'data': user_account.to_dict(),
    })


def delete_user_account():
    """
    Hapus account HRIS seorang pegawai.

    Yang dihapus:
      1. HAK_AKSES_FORM untuk NIP + Modul HRIS
      2. USER_ACCOUNT untuk UserID + Modul HRIS

    Yang TIDAK dihapus:
      - PEGAWAI
      - data kepegawaian lainnya
      - account pada modul selain HRIS
    """
    payload = request.get_json(silent=True) or {}

    nip = (payload.get('nip') or '').strip()

    if not nip:
        return jsonify({
            'status': 'error',
            'message': 'NIP wajib diisi'
        }), 400

    user_account = UserAccount.query.filter(
        UserAccount.NIP == nip,
        UserAccount.MODUL == 'HRIS'
    ).first()

    if user_account is None:
        return jsonify({
            'status': 'error',
            'message': f'User Account HRIS untuk NIP {nip} tidak ditemukan'
        }), 404

    try:
        # Hapus hak akses form HRIS milik account ini.
        HakAksesForm.query.filter(
            HakAksesForm.NIP == nip,
            HakAksesForm.MODUL == 'HRIS'
        ).delete()

        # Hapus account HRIS.
        db.session.delete(user_account)

        db.session.commit()

    except Exception:
        db.session.rollback()
        raise

    return jsonify({
        'status': 'success',
        'message': f'User Account HRIS {nip} berhasil dihapus'
    })


def master_uang_makan():
    """
    Render halaman Master Uang Makan.

    Halaman ini secara khusus hanya menangani:
      JenisTunjangan = U.Makan
      Activity       = Intern
    """
    edit_id = request.args.get("edit", type=int)

    return render_template(
        "pages/dashboard_1/Master File Uang Makan.html",
        edit_id=edit_id,
    )


def save_uang_makan():
    """
    Insert Master Uang Makan.

    Field bisnis mengikuti HRIS 2013:
      JenisTunjangan = U.Makan
      Activity       = Intern
      HariKerja      = 1
      Fungsional     = All

    IDTunjangan dibuat oleh aplikasi karena tabel legacy tidak
    memiliki AUTO_INCREMENT.
    """
    payload = request.get_json(silent=True) or {}

    tgl_mulai_raw = (payload.get("tgl_mulai") or "").strip()
    nominal_raw = payload.get("nominal")
    no_surat = (payload.get("no_surat") or "").strip()

    if not tgl_mulai_raw:
        return jsonify({
            "status": "error",
            "message": "Tanggal Mulai wajib diisi"
        }), 400

    if nominal_raw in (None, ""):
        return jsonify({
            "status": "error",
            "message": "Nominal wajib diisi"
        }), 400

    try:
        tgl_mulai = datetime.strptime(
            tgl_mulai_raw, "%Y-%m-%d"
        ).date()
    except ValueError:
        return jsonify({
            "status": "error",
            "message": "Format Tanggal Mulai tidak valid"
        }), 400

    try:
        nominal = float(nominal_raw)
    except (TypeError, ValueError):
        return jsonify({
            "status": "error",
            "message": "Nominal harus berupa angka"
        }), 400

    if nominal < 0:
        return jsonify({
            "status": "error",
            "message": "Nominal tidak boleh negatif"
        }), 400

    nip = session.get("nip", "system")

    try:
        # Karena tabel tidak mempunyai AUTO_INCREMENT, ambil ID
        # terbesar kemudian +1.
        last_id = db.session.execute(
            sa_text("""
                SELECT IDTunjangan
                FROM MF_TUNJANGAN
                ORDER BY IDTunjangan DESC
                LIMIT 1
                FOR UPDATE
            """)
        ).scalar()

        next_id = (int(last_id) + 1) if last_id is not None else 1

        row = MfTunjangan(
            IDTUNJANGAN=next_id,
            JENIS_TUNJANGAN="U.Makan",
            ACTIVITY="Intern",
            NOMINAL=nominal,
            TGL_MULAI=tgl_mulai,
            HARI_KERJA=1,
            FUNGSIONAL="All",
            UPDATE_BY=nip,
            UPDATE_DATE=datetime.now(),
            DOKREFF=no_surat,
        )

        db.session.add(row)
        db.session.commit()

        return jsonify({
            "status": "success",
            "message": "Master Uang Makan berhasil disimpan",
            "data": row.to_dict(),
        })

    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception(
            "Gagal menyimpan Master Uang Makan"
        )
        return jsonify({
            "status": "error",
            "message": f"Gagal menyimpan Master Uang Makan: {exc}"
        }), 500


def get_uang_makan_detail():
    """
    Ambil satu record Uang Makan untuk kebutuhan Edit.
    """
    tunjangan_id = request.args.get("id", type=int)

    if tunjangan_id is None:
        return jsonify({
            "status": "error",
            "message": "ID Tunjangan wajib diisi"
        }), 400

    row = (
        MfTunjangan.query
        .filter(
            MfTunjangan.IDTUNJANGAN == tunjangan_id,
            MfTunjangan.JENIS_TUNJANGAN == "U.Makan",
            MfTunjangan.ACTIVITY == "Intern",
        )
        .first()
    )

    if row is None:
        return jsonify({
            "status": "error",
            "message": "Data Uang Makan tidak ditemukan"
        }), 404

    return jsonify({
        "status": "success",
        "data": row.to_dict(),
    })


def update_uang_makan():
    """
    Update Master Uang Makan berdasarkan IDTunjangan.

    IDTunjangan TIDAK BOLEH berubah.
    JenisTunjangan dan Activity dipaksa tetap U.Makan/Intern.
    """
    payload = request.get_json(silent=True) or {}

    tunjangan_id_raw = payload.get("tunjangan_id")
    tgl_mulai_raw = (payload.get("tgl_mulai") or "").strip()
    nominal_raw = payload.get("nominal")
    no_surat = (payload.get("no_surat") or "").strip()

    if tunjangan_id_raw in (None, ""):
        return jsonify({
            "status": "error",
            "message": "ID Tunjangan wajib diisi"
        }), 400

    try:
        tunjangan_id = int(tunjangan_id_raw)
    except (TypeError, ValueError):
        return jsonify({
            "status": "error",
            "message": "ID Tunjangan tidak valid"
        }), 400

    if not tgl_mulai_raw:
        return jsonify({
            "status": "error",
            "message": "Tanggal Mulai wajib diisi"
        }), 400

    try:
        tgl_mulai = datetime.strptime(
            tgl_mulai_raw, "%Y-%m-%d"
        ).date()
    except ValueError:
        return jsonify({
            "status": "error",
            "message": "Format Tanggal Mulai tidak valid"
        }), 400

    try:
        nominal = float(nominal_raw)
    except (TypeError, ValueError):
        return jsonify({
            "status": "error",
            "message": "Nominal harus berupa angka"
        }), 400

    if nominal < 0:
        return jsonify({
            "status": "error",
            "message": "Nominal tidak boleh negatif"
        }), 400

    row = (
        MfTunjangan.query
        .filter(
            MfTunjangan.IDTUNJANGAN == tunjangan_id,
            MfTunjangan.JENIS_TUNJANGAN == "U.Makan",
            MfTunjangan.ACTIVITY == "Intern",
        )
        .first()
    )

    if row is None:
        return jsonify({
            "status": "error",
            "message": "Data Uang Makan tidak ditemukan"
        }), 404

    row.TGL_MULAI = tgl_mulai
    row.NOMINAL = nominal
    row.DOKREFF = no_surat
    row.UPDATE_BY = session.get("nip", "system")
    row.UPDATE_DATE = datetime.now()

    # Field domain Uang Makan tidak diubah saat EDIT.
    # Khusus HariKerja: pertahankan nilai existing agar data historis
    # tidak berubah hanya karena Tanggal/Nominal/No Surat diedit.

    try:
        db.session.commit()

        return jsonify({
            "status": "success",
            "message": "Master Uang Makan berhasil diupdate",
            "data": row.to_dict(),
        })

    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception(
            "Gagal update Master Uang Makan"
        )
        return jsonify({
            "status": "error",
            "message": f"Gagal update Master Uang Makan: {exc}"
        }), 500


def _get_uang_makan_reference_counts(tunjangan_id):
    """
    Cari tabel lain yang mempunyai kolom IDTunjangan.

    Kita tidak mengasumsikan FK legacy karena database memang tidak
    mendeklarasikan seluruh relasi sebagai foreign key.
    """
    tables = db.session.execute(
        sa_text("""
            SELECT TABLE_NAME
            FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
              AND LOWER(COLUMN_NAME) = LOWER('IDTunjangan')
              AND TABLE_NAME <> 'MF_TUNJANGAN'
        """)
    ).scalars().all()

    references = []

    for table_name in tables:
        safe_table = str(table_name).replace("`", "``")

        count = db.session.execute(
            sa_text(
                f"SELECT COUNT(*) FROM `{safe_table}` "
                "WHERE IDTunjangan = :id"
            ),
            {"id": tunjangan_id},
        ).scalar()

        if count and int(count) > 0:
            references.append({
                "table": table_name,
                "count": int(count),
            })

    return references


def delete_uang_makan():
    """
    Hapus Master Uang Makan dengan pemeriksaan referensi.

    Tidak langsung DELETE apabila record masih direferensikan tabel lain.
    """
    payload = request.get_json(silent=True) or {}

    tunjangan_id_raw = payload.get("tunjangan_id")

    if tunjangan_id_raw in (None, ""):
        return jsonify({
            "status": "error",
            "message": "ID Tunjangan wajib diisi"
        }), 400

    try:
        tunjangan_id = int(tunjangan_id_raw)
    except (TypeError, ValueError):
        return jsonify({
            "status": "error",
            "message": "ID Tunjangan tidak valid"
        }), 400

    row = (
        MfTunjangan.query
        .filter(
            MfTunjangan.IDTUNJANGAN == tunjangan_id,
            MfTunjangan.JENIS_TUNJANGAN == "U.Makan",
            MfTunjangan.ACTIVITY == "Intern",
        )
        .first()
    )

    if row is None:
        return jsonify({
            "status": "error",
            "message": "Data Uang Makan tidak ditemukan"
        }), 404

    try:
        references = _get_uang_makan_reference_counts(tunjangan_id)

        if references:
            detail = "; ".join(
                f"{item['table']} ({item['count']} referensi)"
                for item in references
            )

            return jsonify({
                "status": "error",
                "message": (
                    "Data tidak dapat dihapus karena masih digunakan "
                    f"oleh: {detail}"
                ),
                "references": references,
            }), 409

        db.session.delete(row)
        db.session.commit()

        return jsonify({
            "status": "success",
            "message": "Master Uang Makan berhasil dihapus",
        })

    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception(
            "Gagal menghapus Master Uang Makan"
        )
        return jsonify({
            "status": "error",
            "message": f"Gagal menghapus Master Uang Makan: {exc}"
        }), 500


def get_tunjangan_list():
    """
    API daftar Master Uang Makan.

    Selalu hanya mengembalikan:
      JenisTunjangan = U.Makan
      Activity       = Intern
    """
    rows = _query_tunjangan().all()

    data = []

    for idx, row in enumerate(rows, start=1):
        updated = "-"

        if row.UPDATE_DATE:
            updated = row.UPDATE_DATE.strftime(
                "%d/%m/%Y %H:%M:%S"
            )

        data.append({
            "no": idx,
            "tunjangan_id": row.IDTUNJANGAN,
            "jenis_tunjangan": row.JENIS_TUNJANGAN or "-",
            "activity": row.ACTIVITY or "-",
            "tgl_mulai": (
                row.TGL_MULAI.strftime("%d/%m/%Y")
                if row.TGL_MULAI else "-"
            ),
            "nominal": (
                f"{row.NOMINAL:,.0f}"
                if row.NOMINAL is not None else "0"
            ),
            "hari_kerja": row.HARI_KERJA,
            "shift": row.SHIFT or "-",
            "fungsional": row.FUNGSIONAL or "-",
            "no_surat": row.DOKREFF or "-",
            "updated": updated,
        })

    return jsonify({
        "status": "success",
        "data": data,
    })

def cari_master_jabatan():
    """Render halaman Cari Master Jabatan."""
    return render_template('pages/dashboard_1/Cari Master Jabatan.html')

def get_jabatan_list():
    """
    Ambil data Master Jabatan untuk tabel Cari Master Jabatan.

    Filter opsional (semua bisa kosong -> berlaku seperti klik Refresh biasa,
    menampilkan seluruh data):
      - field1/keyword1 dan field2/keyword2 : dua dropdown "Filter"
        (Jabatan ID, Nama Jabatan, Isi Aktif, Tanggal Mulai), digabung dengan AND

    Catatan: "Tanggal Mulai" dipetakan ke kolom UPDATE_DATE karena
    MfJabatan tidak punya kolom tanggal mulai berlaku tersendiri.
    Sesuaikan kalau ternyata dimaksudkan untuk kolom lain.
    """
    field1 = request.args.get('field1')
    keyword1 = request.args.get('keyword1', '').strip()
    field2 = request.args.get('field2')
    keyword2 = request.args.get('keyword2', '').strip()

    query = MfJabatan.query

    for field, keyword in [(field1, keyword1), (field2, keyword2)]:
        if not field or not keyword:
            continue  # filter ini tidak dipakai -> skip, tidak wajib diisi

        if field == 'Jabatan ID':
            try:
                nilai = int(keyword)
            except ValueError:
                return jsonify({'status': 'error', 'message': 'Jabatan ID harus berupa angka'}), 400
            query = query.filter(MfJabatan.JABATAN_ID == nilai)

        elif field == 'Nama Jabatan':
            query = query.filter(MfJabatan.NAMA_JABATAN.ilike(f'%{keyword}%'))

        elif field == 'Isi Aktif':
            keyword_lower = keyword.strip().lower()
            if keyword_lower in ('aktif', '1'):
                query = query.filter(MfJabatan.IS_USE == 1)
            elif keyword_lower in ('non aktif', 'nonaktif', '0'):
                query = query.filter(MfJabatan.IS_USE == 0)
            else:
                return jsonify({
                    'status': 'error',
                    'message': 'Isi Aktif harus diisi "Aktif" atau "Non Aktif"'
                }), 400

        elif field == 'Tanggal Mulai':
            try:
                tgl = datetime.strptime(keyword, '%Y-%m-%d')
            except ValueError:
                return jsonify({'status': 'error', 'message': 'Format Tanggal Mulai harus YYYY-MM-DD'}), 400
            awal_hari = tgl
            akhir_hari = tgl + timedelta(days=1)
            query = query.filter(
                MfJabatan.UPDATE_DATE >= awal_hari,
                MfJabatan.UPDATE_DATE < akhir_hari
            )

    jabatan_list = query.order_by(MfJabatan.JABATAN_ID.asc()).all()

    def _format_jabatan_updated(value):
        # Database legacy bisa mengembalikan UPDATE_DATE sebagai
        # datetime maupun string, termasuk kemungkinan zero-date.
        if value is None:
            return '-'

        if isinstance(value, str):
            value = value.strip()

            if not value or value.startswith('0000-00-00'):
                return '-'

            for fmt in (
                '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%d %H:%M',
                '%d-%m-%Y %H:%M:%S',
                '%d-%m-%Y %H:%M',
            ):
                try:
                    return datetime.strptime(
                        value,
                        fmt
                    ).strftime('%d-%m-%Y %H:%M')
                except ValueError:
                    continue

            return value

        try:
            return value.strftime('%d-%m-%Y %H:%M')
        except AttributeError:
            return str(value)

    def _format_jabatan_is_aktif(value):
        if value is None:
            return '-'

        normalized = str(value).strip().upper()

        if normalized in ('1', 'Y', 'YA', 'TRUE'):
            return 'Aktif'

        if normalized in ('0', 'N', 'TIDAK', 'FALSE'):
            return 'Non Aktif'

        return str(value)

    data = [
        {
            'no': idx + 1,
            'jabatan_id': row.JABATAN_ID,
            'nama_jabatan': row.NAMA_JABATAN or '-',
            'butir_kegiatan': '-',  # belum ada model/relasi Butir Kegiatan
            'urut_jabatan': (
                row.URUT_JABATAN
                if row.URUT_JABATAN is not None
                else '-'
            ),
            'is_aktif': _format_jabatan_is_aktif(row.IS_USE),
            'updated': _format_jabatan_updated(row.UPDATE_DATE),
        }
        for idx, row in enumerate(jabatan_list)
    ]

    return jsonify({'status': 'success', 'data': data})


def get_jabatan_by_id():
    """Ambil satu Master Jabatan berdasarkan JabatanID untuk mode Edit."""
    id_raw = request.args.get('id', '').strip()

    try:
        jabatan_id = int(id_raw)
    except (TypeError, ValueError):
        return jsonify({
            'status': 'error',
            'message': 'Jabatan ID tidak valid'
        }), 400

    row = MfJabatan.query.filter(
        MfJabatan.JABATAN_ID == jabatan_id
    ).first()

    if row is None:
        return jsonify({
            'status': 'error',
            'message': 'Data Master Jabatan tidak ditemukan'
        }), 404

    is_use = str(row.IS_USE).strip().upper() if row.IS_USE is not None else ''

    if is_use in ('Y', 'YA', 'TRUE'):
        is_aktif = '1'
    elif is_use in ('N', 'TIDAK', 'FALSE'):
        is_aktif = '0'
    elif is_use in ('1', '0'):
        is_aktif = is_use
    else:
        is_aktif = ''

    return jsonify({
        'status': 'success',
        'data': {
            'jabatan_id': row.JABATAN_ID,
            'group_jabatan_id': row.GROUP_JABATAN_ID,
            'sub_group_jabatan_id': row.SUB_GROUP_JABATAN_ID,
            'nama_jabatan': row.NAMA_JABATAN or '',
            'parent_jabatan_id': (
                row.PARENT_ID
                if row.PARENT_ID is not None
                else ''
            ),
            'level_jabatan': (
                row.URUT_JABATAN
                if row.URUT_JABATAN is not None
                else ''
            ),
            'type_jabatan': row.TYPE_JABATAN or '',
            'is_aktif': is_aktif,
        }
    })


def update_jabatan():
    """Update Master Jabatan dengan mempertahankan JabatanID."""
    payload = request.get_json(silent=True) or {}

    jabatan_id_raw = payload.get('jabatan_id')
    group_jabatan_id_raw = payload.get('group_jabatan_id')
    sub_group_jabatan_id_raw = payload.get('sub_group_jabatan_id')
    nama_jabatan = (payload.get('nama_jabatan') or '').strip()
    parent_jabatan_id_raw = payload.get('parent_jabatan_id')
    level_jabatan_raw = payload.get('level_jabatan')
    type_jabatan = (payload.get('type_jabatan') or '').strip()
    is_aktif_raw = payload.get('is_aktif')

    if jabatan_id_raw in (None, ''):
        return jsonify({
            'status': 'error',
            'message': 'Jabatan ID wajib diisi'
        }), 400

    try:
        jabatan_id = int(jabatan_id_raw)
    except (TypeError, ValueError):
        return jsonify({
            'status': 'error',
            'message': 'Jabatan ID harus berupa angka'
        }), 400

    jabatan = MfJabatan.query.filter(
        MfJabatan.JABATAN_ID == jabatan_id
    ).first()

    if jabatan is None:
        return jsonify({
            'status': 'error',
            'message': 'Data Master Jabatan tidak ditemukan'
        }), 404

    if group_jabatan_id_raw in (None, ''):
        return jsonify({
            'status': 'error',
            'message': 'Group Jabatan wajib dipilih'
        }), 400

    if sub_group_jabatan_id_raw in (None, ''):
        return jsonify({
            'status': 'error',
            'message': 'SubGroup Jabatan wajib dipilih'
        }), 400

    if not nama_jabatan:
        return jsonify({
            'status': 'error',
            'message': 'Nama Jabatan wajib diisi'
        }), 400

    if level_jabatan_raw in (None, ''):
        return jsonify({
            'status': 'error',
            'message': 'Level Jabatan wajib diisi'
        }), 400

    if not type_jabatan:
        return jsonify({
            'status': 'error',
            'message': 'Type wajib dipilih'
        }), 400

    if is_aktif_raw is None:
        return jsonify({
            'status': 'error',
            'message': 'Isi Aktif wajib dipilih'
        }), 400

    try:
        group_jabatan_id = int(group_jabatan_id_raw)
    except (TypeError, ValueError):
        return jsonify({
            'status': 'error',
            'message': 'Group Jabatan ID harus berupa angka'
        }), 400

    group_jabatan = MfGroupJabatan.query.get(group_jabatan_id)

    if group_jabatan is None:
        return jsonify({
            'status': 'error',
            'message': 'Group Jabatan tidak ditemukan'
        }), 400

    try:
        sub_group_jabatan_id = int(sub_group_jabatan_id_raw)
    except (TypeError, ValueError):
        return jsonify({
            'status': 'error',
            'message': 'SubGroup Jabatan ID harus berupa angka'
        }), 400

    sub_group_jabatan = MfSubGroupJabatan.query.get(
        sub_group_jabatan_id
    )

    if sub_group_jabatan is None:
        return jsonify({
            'status': 'error',
            'message': 'SubGroup Jabatan tidak ditemukan'
        }), 400

    parent_jabatan_id = None

    if parent_jabatan_id_raw not in (None, ''):
        try:
            parent_jabatan_id = int(parent_jabatan_id_raw)
        except (TypeError, ValueError):
            return jsonify({
                'status': 'error',
                'message': 'Parent Jabatan ID harus berupa angka'
            }), 400

    try:
        level_jabatan = int(level_jabatan_raw)
    except (TypeError, ValueError):
        return jsonify({
            'status': 'error',
            'message': 'Level Jabatan harus berupa angka bulat'
        }), 400

    if type_jabatan not in ('FT', 'FU'):
        return jsonify({
            'status': 'error',
            'message': 'Type harus "FT" atau "FU"'
        }), 400

    # Pertahankan format legacy IsUse yang sudah digunakan row lama.
    current_is_use = (
        str(jabatan.IS_USE).strip().upper()
        if jabatan.IS_USE is not None
        else ''
    )

    if current_is_use in ('Y', 'N'):
        is_use = 'Y' if bool(is_aktif_raw) else 'N'
    else:
        is_use = 1 if bool(is_aktif_raw) else 0

    jabatan.GROUP_JABATAN_ID = group_jabatan_id
    jabatan.SUB_GROUP_JABATAN_ID = sub_group_jabatan_id
    jabatan.PARENT_ID = parent_jabatan_id
    jabatan.NAMA_JABATAN = nama_jabatan
    jabatan.URUT_JABATAN = level_jabatan
    jabatan.TYPE_JABATAN = type_jabatan
    jabatan.IS_USE = is_use
    jabatan.UPDATE_BY = session.get('nip', 'system')
    jabatan.UPDATE_DATE = datetime.utcnow()

    db.session.commit()

    return jsonify({
        'status': 'success',
        'message': (
            f'Data Master Jabatan ID {jabatan_id} '
            'berhasil diperbarui'
        ),
        'data': jabatan.to_dict(),
    })


def delete_jabatan():
    """Hapus Master Jabatan hanya jika tidak memiliki referensi data lain."""
    payload = request.get_json(silent=True) or {}
    jabatan_id_raw = payload.get('jabatan_id')

    try:
        jabatan_id = int(jabatan_id_raw)
    except (TypeError, ValueError):
        return jsonify({
            'status': 'error',
            'message': 'Jabatan ID tidak valid'
        }), 400

    jabatan = MfJabatan.query.filter(
        MfJabatan.JABATAN_ID == jabatan_id
    ).first()

    if jabatan is None:
        return jsonify({
            'status': 'error',
            'message': 'Data Master Jabatan tidak ditemukan'
        }), 404

    references = []

    child_count = MfJabatan.query.filter(
        MfJabatan.PARENT_ID == jabatan_id
    ).count()

    if child_count > 0:
        references.append(
            f'{child_count} data Master Jabatan sebagai child/struktur bawahan'
        )

    kegiatan_result = db.session.execute(
        text(
            'SELECT COUNT(*) '
            'FROM MF_JABATAN_KEGIATAN '
            'WHERE JabatanID = :id'
        ),
        {'id': jabatan_id}
    )

    kegiatan_count = kegiatan_result.scalar() or 0

    if kegiatan_count > 0:
        references.append(
            f'{kegiatan_count} data Master Jabatan Kegiatan'
        )

    pegawai_result = db.session.execute(
        text(
            'SELECT COUNT(*) '
            'FROM PEGAWAI '
            'WHERE JabatanID = :id'
        ),
        {'id': jabatan_id}
    )

    pegawai_count = pegawai_result.scalar() or 0

    if pegawai_count > 0:
        references.append(
            f'{pegawai_count} data Pegawai'
        )

    checks = [
        ('BUKU_HARIAN_HEAD', 'SELECT COUNT(*) FROM BUKU_HARIAN_HEAD WHERE JabatanID = :id OR JabatanIDParent = :id'),
        ('DRH', 'SELECT COUNT(*) FROM DRH WHERE JabatanID = :id'),
        ('PERUBAHAN_JABATAN', 'SELECT COUNT(*) FROM PERUBAHAN_JABATAN WHERE JabatanIDBaru = :id OR JabatanIDLama = :id'),
        ('SARAN', 'SELECT COUNT(*) FROM SARAN WHERE JabatanID = :id'),
        ('SKP_PEGAWAI', 'SELECT COUNT(*) FROM SKP_PEGAWAI WHERE JabatanID = :id'),
        ('SKP_PEGAWAI_HEAD', 'SELECT COUNT(*) FROM SKP_PEGAWAI_HEAD WHERE JabatanID = :id OR JabatanIDParent = :id'),
    ]

    for table_name, sql in checks:
        result = db.session.execute(
            text(sql),
            {'id': jabatan_id}
        )
        count = result.scalar() or 0

        if count > 0:
            references.append(
                f'{count} data {table_name}'
            )

    if references:
        return jsonify({
            'status': 'error',
            'message': (
                f'Data Master Jabatan ID {jabatan_id} tidak dapat dihapus '
                'karena masih digunakan oleh data lain.'
            ),
            'references': references
        }), 409

    try:
        db.session.delete(jabatan)
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': (
                f'Data Master Jabatan ID {jabatan_id} gagal dihapus '
                'karena terjadi kesalahan database.'
            )
        }), 500

    return jsonify({
        'status': 'success',
        'message': (
            f'Data Master Jabatan ID {jabatan_id} berhasil dihapus'
        ),
    })


def cari_master_jam_finger():
    """Render halaman Cari Master Jam Finger."""
    return render_template('pages/dashboard_1/Cari Master Jam Finger.html')

def _query_jam_finger(periode_raw, field1, keyword1, field2, keyword2):
    """
    Helper bersama untuk get_jam_finger_list() dan export_jam_finger_excel(),
    supaya logic filter tidak perlu ditulis dua kali dan selalu konsisten
    antara tampilan tabel dan hasil download Excel.

    Filter opsional (semua bisa kosong -> tampilkan semua data):
      - periode : filter TGL_MULAI_BERLAKU pada tanggal tertentu (YYYY-MM-DD)
      - field1/keyword1, field2/keyword2 : dropdown "Filter" (cuma ada
        "Tanggal Mulai" di UI saat ini), digabung dengan AND
    """
    query = MfLoadFinger.query

    if periode_raw:
        try:
            periode_date = datetime.strptime(periode_raw, '%Y-%m-%d')
        except ValueError:
            raise ValueError('Format periode harus YYYY-MM-DD')
        awal_hari = periode_date
        akhir_hari = periode_date + timedelta(days=1)
        query = query.filter(
            MfLoadFinger.TGL_MULAI_BERLAKU >= awal_hari,
            MfLoadFinger.TGL_MULAI_BERLAKU < akhir_hari
        )

    for field, keyword in [(field1, keyword1), (field2, keyword2)]:
        if not field or not keyword:
            continue  # filter tidak dipakai -> skip, tidak wajib diisi

        if field == 'Tanggal Mulai':
            try:
                tgl = datetime.strptime(keyword, '%Y-%m-%d')
            except ValueError:
                raise ValueError('Format Tanggal Mulai harus YYYY-MM-DD')
            awal_hari = tgl
            akhir_hari = tgl + timedelta(days=1)
            query = query.filter(
                MfLoadFinger.TGL_MULAI_BERLAKU >= awal_hari,
                MfLoadFinger.TGL_MULAI_BERLAKU < akhir_hari
            )

    return query.order_by(MfLoadFinger.TGL_MULAI_BERLAKU.desc()).all()


def _format_jam_finger_date(value):
    """
    Format tanggal Master Jam Finger menjadi DD-MM-YYYY.
    Aman untuk datetime/date maupun string legacy dari database.
    """
    if value is None:
        return '-'

    if isinstance(value, str):
        value = value.strip()
        if not value or value.startswith('0000-00-00'):
            return '-'

        for fmt in (
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d',
            '%d-%m-%Y %H:%M:%S',
            '%d-%m-%Y',
        ):
            try:
                return datetime.strptime(value, fmt).strftime('%d-%m-%Y')
            except ValueError:
                continue

        return value

    if hasattr(value, 'strftime'):
        return value.strftime('%d-%m-%Y')

    return str(value)


def _format_jam_finger_time(value):
    """
    Format jam Master Jam Finger menjadi HH:MM.
    Aman untuk datetime/time maupun string legacy dari database.
    """
    if value is None:
        return '-'

    if isinstance(value, str):
        value = value.strip()
        if not value or value.startswith('0000-00-00'):
            return '-'

        for fmt in (
            '%Y-%m-%d %H:%M:%S',
            '%H:%M:%S',
            '%H:%M',
        ):
            try:
                return datetime.strptime(value, fmt).strftime('%H:%M')
            except ValueError:
                continue

        return value

    if hasattr(value, 'strftime'):
        return value.strftime('%H:%M')

    return str(value)


def _format_jam_finger_updated(value):
    """
    Format UpdateDate menjadi DD-MM-YYYY HH:MM.
    Menangani datetime normal dan zero-date legacy
    '0000-00-00 00:00:00' tanpa mengubah data database.
    """
    if value is None:
        return '-'

    if isinstance(value, str):
        value = value.strip()
        if not value or value.startswith('0000-00-00'):
            return '-'

        for fmt in (
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%d-%m-%Y %H:%M:%S',
            '%d-%m-%Y %H:%M',
        ):
            try:
                return datetime.strptime(value, fmt).strftime('%d-%m-%Y %H:%M')
            except ValueError:
                continue

        return value

    if hasattr(value, 'strftime'):
        return value.strftime('%d-%m-%Y %H:%M')

    return str(value)


def get_jam_finger_list():
    """
    Ambil data Master Jam Finger untuk tabel Cari Master Jam Finger.
    Filter opsional -> kalau semua kosong, berlaku seperti klik Refresh biasa.
    """
    periode_raw = request.args.get('periode', '').strip()
    field1 = request.args.get('field1')
    keyword1 = request.args.get('keyword1', '').strip()
    field2 = request.args.get('field2')
    keyword2 = request.args.get('keyword2', '').strip()

    try:
        rows = _query_jam_finger(periode_raw, field1, keyword1, field2, keyword2)
    except ValueError as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400

    data = [
        {
            'no': idx + 1,
            'transaksi_id': row.TRAKSAKSI_ID,
            'tgl_mulai': _format_jam_finger_date(row.TGL_MULAI_BERLAKU),
            'shift': row.SHIFT_KERJA or '-',
            'start_finger': _format_jam_finger_time(row.START_FINGER),
            'end_finger': _format_jam_finger_time(row.END_FINGER),
            'start_finger_out': _format_jam_finger_time(row.START_FINGER_OUT),
            'end_finger_out': _format_jam_finger_time(row.END_FINGER_OUT),
            'updated': _format_jam_finger_updated(row.UPDATE_DATE),
        }
        for idx, row in enumerate(rows)
    ]

    return jsonify({'status': 'success', 'data': data})


def export_jam_finger_excel():
    """
    Export data Master Jam Finger ke file Excel (.xlsx), dengan filter
    yang SAMA PERSIS seperti tabel di layar -- supaya file yang di-download
    selalu cocok dengan apa yang sedang ditampilkan user.
    """
    periode_raw = request.args.get('periode', '').strip()
    field1 = request.args.get('field1')
    keyword1 = request.args.get('keyword1', '').strip()
    field2 = request.args.get('field2')
    keyword2 = request.args.get('keyword2', '').strip()

    try:
        rows = _query_jam_finger(periode_raw, field1, keyword1, field2, keyword2)
    except ValueError as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = 'Jam Finger'

    headers = [
        'No', 'Tgl Mulai', 'Shift',
        'Start Finger (In)', 'End Finger (In)',
        'Start Finger (Out)', 'End Finger (Out)',
        'Updated',
    ]
    ws.append(headers)

    # Styling header supaya konsisten dengan warna orange di tabel web (#EB6831)
    header_fill = PatternFill(start_color='EB6831', end_color='EB6831', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for idx, row in enumerate(rows):
        ws.append([
            idx + 1,
            _format_jam_finger_date(row.TGL_MULAI_BERLAKU),
            row.SHIFT_KERJA or '-',
            _format_jam_finger_time(row.START_FINGER),
            _format_jam_finger_time(row.END_FINGER),
            _format_jam_finger_time(row.START_FINGER_OUT),
            _format_jam_finger_time(row.END_FINGER_OUT),
            _format_jam_finger_updated(row.UPDATE_DATE),
        ])

    # Auto-lebar kolom supaya isinya tidak terpotong
    for col_cells in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col_cells if cell.value is not None)
        col_letter = col_cells[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 4

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'jam_finger_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

def cari_master_jam_kerja():
    """Render halaman Cari Master Jam Kerja."""
    return render_template('pages/dashboard_1/Cari Master Jam Kerja.html')

def _format_jam(value):
    """
    Format nilai jam (STD_JAM_IN/STD_JAM_OUT) menjadi string 'HH:MM',
    aman dipakai baik value berupa objek datetime maupun string mentah
    dari database (mengatasi ketidakcocokan tipe kolom DB vs model).
    """
    if value is None:
        return '-'
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return '-'
        # Coba beberapa format string yang mungkin tersimpan di DB
        for fmt in ('%Y-%m-%d %H:%M:%S', '%H:%M:%S', '%H:%M'):
            try:
                return datetime.strptime(value, fmt).strftime('%H:%M')
            except ValueError:
                continue
        # Kalau tidak ada format yang cocok, tampilkan apa adanya
        return value
    # Kalau sudah objek datetime/time
    try:
        return value.strftime('%H:%M')
    except AttributeError:
        return str(value)


def get_jam_kerja_by_id():
    """Ambil satu Master Jam Kerja berdasarkan IDJKerja untuk mode Edit."""
    id_raw = request.args.get('id', '').strip()

    try:
        id_jkerja = int(id_raw)
    except (TypeError, ValueError):
        return jsonify({
            'status': 'error',
            'message': 'IDJKerja tidak valid'
        }), 400

    row = MfJamKerja.query.filter(
        MfJamKerja.IDJKERJA == id_jkerja
    ).first()

    if row is None:
        return jsonify({
            'status': 'error',
            'message': 'Data Master Jam Kerja tidak ditemukan'
        }), 404

    # Mapping struktur database legacy menjadi kode bisnis SK-1 s.d. SK-4.
    shift_db = str(row.SHIFT).strip() if row.SHIFT is not None else ''
    shift_kerja_db = str(row.SHIFT_KERJA).strip() if row.SHIFT_KERJA is not None else ''

    shift_map = {
        ('1', '1'): '1',
        ('1', '2'): '2',
        ('2', '1'): '3',
        ('2', '2'): '4',
    }

    shift_bisnis = shift_map.get((shift_db, shift_kerja_db))

    if shift_bisnis is None:
        return jsonify({
            'status': 'error',
            'message': 'Data Shift/Hari Kerja pada database tidak valid'
        }), 400

    return jsonify({
        'status': 'success',
        'data': {
            'id_jkerja': row.IDJKERJA,
            'shift': shift_bisnis,
            'hari_kerja': shift_db,
            'tgl_mulai': (
                row.TGL_MULAI_BERLAKU.strftime('%Y-%m-%d')
                if row.TGL_MULAI_BERLAKU else ''
            ),
            'jam_masuk': _format_jam(row.STD_JAM_IN),
            'jam_pulang': _format_jam(row.STD_JAM_OUT),
            'penggantian_tlm1': (
                '1' if str(row.PENGGANTIAN_TLM1).strip() == 'Y' else '0'
            ),
        }
    })


def update_jam_kerja():
    """Update Master Jam Kerja berdasarkan IDJKerja yang sama."""
    payload = request.get_json(silent=True) or {}

    id_raw = payload.get('id_jkerja')
    try:
        id_jkerja = int(id_raw)
    except (TypeError, ValueError):
        return jsonify({
            'status': 'error',
            'message': 'IDJKerja tidak valid'
        }), 400

    jam_kerja = MfJamKerja.query.filter(
        MfJamKerja.IDJKERJA == id_jkerja
    ).first()

    if jam_kerja is None:
        return jsonify({
            'status': 'error',
            'message': 'Data Master Jam Kerja tidak ditemukan'
        }), 404

    shift_raw = str(payload.get('shift', '')).strip()
    hari_kerja_raw = str(payload.get('hari_kerja', '')).strip()
    tgl_mulai_raw = str(payload.get('tgl_mulai', '')).strip()
    jam_masuk_raw = str(payload.get('jam_masuk', '')).strip()
    jam_pulang_raw = str(payload.get('jam_pulang', '')).strip()
    penggantian_tlm1_raw = payload.get('penggantian_tlm1')

    if shift_raw not in ('1', '2', '3', '4'):
        return jsonify({
            'status': 'error',
            'message': 'Shift Kerja tidak valid. Pilih SK-1 sampai SK-4.'
        }), 400

    if hari_kerja_raw not in ('1', '2'):
        return jsonify({
            'status': 'error',
            'message': 'Hari Kerja tidak valid'
        }), 400

    if not tgl_mulai_raw or not jam_masuk_raw or not jam_pulang_raw:
        return jsonify({
            'status': 'error',
            'message': 'Tanggal Mulai, Jam Masuk, dan Jam Pulang wajib diisi'
        }), 400

    shift_definition = get_shift_kerja_definition(shift_raw)

    if shift_definition is None:
        return jsonify({
            'status': 'error',
            'message': 'Shift Kerja tidak valid'
        }), 400

    if shift_definition['hari_kerja'] != hari_kerja_raw:
        return jsonify({
            'status': 'error',
            'message': (
                f"Shift Kerja SK-{shift_raw} tidak sesuai dengan "
                f"Hari Kerja HK-{hari_kerja_raw}."
            )
        }), 400

    try:
        tgl_mulai = datetime.strptime(tgl_mulai_raw, '%Y-%m-%d')
        jam_masuk_time = datetime.strptime(jam_masuk_raw, '%H:%M').time()
        jam_pulang_time = datetime.strptime(jam_pulang_raw, '%H:%M').time()
    except ValueError:
        return jsonify({
            'status': 'error',
            'message': 'Format tanggal/jam tidak valid'
        }), 400

    shift_db = shift_definition['hari_kerja']
    shift_kerja_db = '1' if shift_raw in ('1', '3') else '2'

    jam_kerja.TGL_MULAI_BERLAKU = tgl_mulai
    jam_kerja.STD_JAM_IN = datetime.combine(
        tgl_mulai.date(),
        jam_masuk_time,
    )
    jam_kerja.STD_JAM_OUT = datetime.combine(
        tgl_mulai.date(),
        jam_pulang_time,
    )
    jam_kerja.SHIFT = shift_db
    jam_kerja.SHIFT_KERJA = shift_kerja_db
    jam_kerja.PENGGANTIAN_TLM1 = (
        'Y' if penggantian_tlm1_raw else 'N'
    )
    jam_kerja.UPDATE_BY = session.get('nip', 'system')
    jam_kerja.UPDATE_DATE = datetime.now()

    db.session.commit()

    return jsonify({
        'status': 'success',
        'message': f'Data Master Jam Kerja ID {id_jkerja} berhasil diperbarui',
        'id_jkerja': id_jkerja,
    })


def delete_jam_kerja():
    """Hapus Master Jam Kerja berdasarkan IDJKerja."""
    payload = request.get_json(silent=True) or {}
    id_raw = payload.get('id_jkerja')

    try:
        id_jkerja = int(id_raw)
    except (TypeError, ValueError):
        return jsonify({
            'status': 'error',
            'message': 'IDJKerja tidak valid'
        }), 400

    jam_kerja = MfJamKerja.query.filter(
        MfJamKerja.IDJKERJA == id_jkerja
    ).first()

    if jam_kerja is None:
        return jsonify({
            'status': 'error',
            'message': 'Data Master Jam Kerja tidak ditemukan'
        }), 404

    db.session.delete(jam_kerja)
    db.session.commit()

    return jsonify({
        'status': 'success',
        'message': f'Data Master Jam Kerja ID {id_jkerja} berhasil dihapus',
    })


def get_jam_kerja_list():
    """
    Ambil data Master Jam Kerja untuk tabel Cari Master Jam Kerja.

    Filter opsional (semua bisa kosong -> berlaku seperti klik Refresh biasa,
    menampilkan seluruh data):
      - periode        : filter TGL_MULAI_BERLAKU pada tanggal tertentu (format YYYY-MM-DD)
      - field1/keyword1 dan field2/keyword2 : dua dropdown "Filter"
        ("Hari Kerja Senin-Kamis(1)/Jumat(2)", "Tanggal(yyyy-mm-dd)"), digabung dengan AND
    """
    periode_raw = request.args.get('periode', '').strip()
    field1 = request.args.get('field1')
    keyword1 = request.args.get('keyword1', '').strip()
    field2 = request.args.get('field2')
    keyword2 = request.args.get('keyword2', '').strip()

    query = MfJamKerja.query

    # --- Filter Periode: cocokkan TGL_MULAI_BERLAKU pada tanggal yang dipilih ---
    if periode_raw:
        try:
            periode_date = datetime.strptime(periode_raw, '%Y-%m-%d')
        except ValueError:
            return jsonify({'status': 'error', 'message': 'Format periode harus YYYY-MM-DD'}), 400

        awal_hari = periode_date
        akhir_hari = periode_date + timedelta(days=1)
        query = query.filter(
            MfJamKerja.TGL_MULAI_BERLAKU >= awal_hari,
            MfJamKerja.TGL_MULAI_BERLAKU < akhir_hari
        )

    # --- Konversi label Hari Kerja: 1 -> "Senin-Kamis", 2 -> "Jum'at" ---
    hari_kerja_map = {'1': 'Senin-Kamis', '2': "Jum'at"}

    # --- Filter field1/field2 ---
    # "Hari Kerja Senin-Kamis(1)/Jumat(2)" -> kolom AGENDA, exact match dari mapping 1/2
    # "Tanggal(yyyy-mm-dd)"                -> kolom TGL_MULAI_BERLAKU, exact match tanggal
    for field, keyword in [(field1, keyword1), (field2, keyword2)]:
        if not field or not keyword:
            continue  # filter ini tidak dipakai -> skip, tidak wajib diisi

        if field == 'Hari Kerja Senin-Kamis(1)/Jumat(2)':
            shift = keyword.strip()
            if shift not in ('1', '2'):
                return jsonify({
                    'status': 'error',
                    'message': 'Hari Kerja harus diisi 1 (Senin-Kamis) atau 2 (Jumat)'
                }), 400
            query = query.filter(MfJamKerja.SHIFT == shift)

        elif field == 'Tanggal(yyyy-mm-dd)':
            try:
                tgl = datetime.strptime(keyword, '%Y-%m-%d')
            except ValueError:
                return jsonify({'status': 'error', 'message': 'Format Tanggal harus YYYY-MM-DD'}), 400
            awal_hari = tgl
            akhir_hari = tgl + timedelta(days=1)
            query = query.filter(
                MfJamKerja.TGL_MULAI_BERLAKU >= awal_hari,
                MfJamKerja.TGL_MULAI_BERLAKU < akhir_hari
            )

    jam_kerja_list = query.order_by(
        MfJamKerja.UPDATE_DATE.desc(),
        MfJamKerja.IDJKERJA.desc()
    ).all()

    # --- Label Ada Pengganti TLM1: 'Y' -> Ada, 'N' -> Tidak Ada ---
    tlm1_label = {'Y': 'Ada', 'N': 'Tidak Ada'}

    def _format_updated(value):
        if value is None:
            return '-'
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return '-'
            for fmt in (
                '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%d %H:%M',
                '%d-%m-%Y %H:%M:%S',
                '%d-%m-%Y %H:%M',
            ):
                try:
                    return datetime.strptime(value, fmt).strftime('%d-%m-%Y %H:%M')
                except ValueError:
                    continue
            return value
        try:
            return value.strftime('%d-%m-%Y %H:%M')
        except AttributeError:
            return str(value)

    data = [
        {
            'no': idx + 1,
            'id_jkerja': row.IDJKERJA,
            'tgl_mulai': row.TGL_MULAI_BERLAKU.strftime('%d-%m-%Y') if row.TGL_MULAI_BERLAKU else '-',
            'hari_kerja': hari_kerja_map.get(str(row.SHIFT).strip(), '-')
                if row.SHIFT is not None else '-',
            'shift': {
                ('1', '1'): 'SK-1',
                ('1', '2'): 'SK-2',
                ('2', '1'): 'SK-3',
                ('2', '2'): 'SK-4',
            }.get(
                (
                    str(row.SHIFT).strip() if row.SHIFT is not None else '',
                    str(row.SHIFT_KERJA).strip() if row.SHIFT_KERJA is not None else '',
                ),
                '-',
            ),
            'jam_masuk': _format_jam(row.STD_JAM_IN),
            'jam_pulang': _format_jam(row.STD_JAM_OUT),
            'penggantian_tlm1': tlm1_label.get(row.PENGGANTIAN_TLM1, '-'),
            'updated': _format_updated(row.UPDATE_DATE),
        }
        for idx, row in enumerate(jam_kerja_list)
    ]

    return jsonify({'status': 'success', 'data': data})

def cari_master_kalender():
    """Render halaman Cari Master Kalender."""
    return render_template('pages/dashboard_1/Cari Master Kalender.html')

def cari_master_potongan():
    """Render halaman Cari Master Potongan."""
    return render_template('pages/dashboard_1/Cari Master Potongan.html')

def cari_master_tunkin_class():
    """Render halaman Cari Master Tunkin Class."""
    return render_template('pages/dashboard_1/Cari Master Tunkin Class.html')

def get_tunkin_class_list():
    """
    Ambil data Master Tunkin/Class untuk tabel Cari Master Tunkin Class.

    Filter opsional (semua bisa kosong -> berlaku seperti klik Refresh biasa,
    menampilkan seluruh data):
      - periode        : filter TGL_MULAI pada tanggal tertentu (format YYYY-MM-DD)
      - field1/keyword1 dan field2/keyword2 : dua dropdown "Filter"
        (Class, Tunjangan, No Surat), digabung dengan AND
    """
    periode_raw = request.args.get('periode', '').strip()
    field1 = request.args.get('field1')
    keyword1 = request.args.get('keyword1', '').strip()
    field2 = request.args.get('field2')
    keyword2 = request.args.get('keyword2', '').strip()

    query = MfClass.query

    # --- Filter Periode: cocokkan TGL_MULAI pada tanggal yang dipilih ---
    if periode_raw:
        try:
            periode_date = datetime.strptime(periode_raw, '%Y-%m-%d')
        except ValueError:
            return jsonify({'status': 'error', 'message': 'Format periode harus YYYY-MM-DD'}), 400

        awal_hari = periode_date
        akhir_hari = periode_date + timedelta(days=1)
        query = query.filter(MfClass.TGL_MULAI >= awal_hari, MfClass.TGL_MULAI < akhir_hari)

    # --- Filter field1/field2 (Class, Tunjangan, No Surat) ---
    # Class     -> primary key Integer, exact match angka
    # Tunjangan -> kolom Float, exact match angka
    # No Surat  -> kolom teks (DOKREFF), partial match (ilike)
    for field, keyword in [(field1, keyword1), (field2, keyword2)]:
        if not field or not keyword:
            continue  # filter ini tidak dipakai -> skip, tidak wajib diisi

        if field == 'Class':
            try:
                nilai = int(keyword)
            except ValueError:
                return jsonify({'status': 'error', 'message': 'Class harus berupa angka'}), 400
            query = query.filter(MfClass.CLASS_ID == nilai)
        elif field == 'Tunjangan':
            try:
                nilai = float(keyword)
            except ValueError:
                return jsonify({'status': 'error', 'message': 'Tunjangan harus berupa angka'}), 400
            query = query.filter(MfClass.TUNJANGAN == nilai)
        elif field == 'No Surat':
            query = query.filter(MfClass.DOKREFF.ilike(f'%{keyword}%'))

    tunkin_class_list = query.order_by(MfClass.CLASS_ID.asc()).all()

    data = [
        {
            'no': idx + 1,
            'class_id': row.CLASS_ID,
            'tunjangan': row.TUNJANGAN if row.TUNJANGAN is not None else '-',
            'tgl_mulai': row.TGL_MULAI.strftime('%d-%m-%Y') if row.TGL_MULAI else '-',
            'dokreff': row.DOKREFF or '-',
            'updated': row.UPDATE_DATE.strftime('%d-%m-%Y %H:%M') if row.UPDATE_DATE else '-',
        }
        for idx, row in enumerate(tunkin_class_list)
    ]

    return jsonify({'status': 'success', 'data': data})

def cari_master_uang_makan():
    """Render halaman Cari Master Uang Makan."""
    return render_template('pages/dashboard_1/Cari Master Uang Makan.html')

def _query_tunjangan():
    """
    Query khusus halaman Master Uang Makan.

    Selalu membatasi:
      JenisTunjangan = U.Makan
      Activity       = Intern
    """
    query = (
        MfTunjangan.query
        .filter(
            MfTunjangan.JENIS_TUNJANGAN == "U.Makan",
            MfTunjangan.ACTIVITY == "Intern",
        )
    )

    field_map = {
        "Jenis Tunjangan": MfTunjangan.JENIS_TUNJANGAN,
        "Nominal": MfTunjangan.NOMINAL,
        "No Surat": MfTunjangan.DOKREFF,
        "Tanggal Mulai": MfTunjangan.TGL_MULAI,
    }

    for suffix in ("1", "2"):
        field = (request.args.get(f"field{suffix}") or "").strip()
        keyword = (request.args.get(f"keyword{suffix}") or "").strip()

        if not field or not keyword:
            continue

        column = field_map.get(field)
        if column is None:
            continue

        if field == "Tanggal Mulai":
            try:
                value = datetime.strptime(
                    keyword, "%Y-%m-%d"
                ).date()
                query = query.filter(column >= value)
            except ValueError:
                continue
        elif field == "Nominal":
            try:
                value = float(keyword)
                query = query.filter(column == value)
            except ValueError:
                continue
        else:
            query = query.filter(column.ilike(f"%{keyword}%"))

    return query.order_by(
        MfTunjangan.TGL_MULAI.desc(),
        MfTunjangan.IDTUNJANGAN.desc(),
    )


def export_tunjangan_excel():
    """
    Export khusus Master Uang Makan.
    """
    rows = _query_tunjangan().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Master Uang Makan"

    headers = [
        "No",
        "ID Tunjangan",
        "Jenis Tunjangan",
        "Tanggal Mulai",
        "Nominal",
        "Hari Kerja",
        "Fungsional",
        "No Surat",
        "Updated",
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for idx, row in enumerate(rows, start=1):
        ws.append([
            idx,
            row.IDTUNJANGAN,
            row.JENIS_TUNJANGAN,
            (
                row.TGL_MULAI.strftime("%d/%m/%Y")
                if row.TGL_MULAI else ""
            ),
            row.NOMINAL,
            "Hari Kerja" if row.HARI_KERJA == 1 else "Hari Libur",
            row.FUNGSIONAL or "",
            row.DOKREFF or "",
            (
                row.UPDATE_DATE.strftime("%d/%m/%Y %H:%M:%S")
                if row.UPDATE_DATE else ""
            ),
        ])

    for column_cells in ws.columns:
        length = max(
            len(str(cell.value or ""))
            for cell in column_cells
        )
        ws.column_dimensions[
            column_cells[0].column_letter
        ].width = min(length + 2, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="master_uang_makan.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )

def cari_master_unit_kerja():
    """Render halaman Cari Master Unit Kerja."""
    return render_template('pages/dashboard_1/Cari Master Unit Kerja.html')

def get_unit_kerja_list():
    """
    Ambil data Master Unit Kerja untuk tabel Cari Master Unit Kerja.

    Filter opsional (semua bisa kosong -> berlaku seperti klik Refresh biasa):
      - periode        : filter UPDATE_DATE pada tanggal tertentu (format YYYY-MM-DD)
      - field1/keyword1 dan field2/keyword2 : dua dropdown "Filter"
        (Nama Unit Kerja, Unit Kerja ID), digabung dengan AND
    """
    periode_raw = request.args.get('periode', '').strip()
    field1 = request.args.get('field1')
    keyword1 = request.args.get('keyword1', '').strip()
    field2 = request.args.get('field2')
    keyword2 = request.args.get('keyword2', '').strip()

    query = MfUnitKerja.query

    # --- Filter Periode: cocokkan UPDATE_DATE pada tanggal yang dipilih ---
    if periode_raw:
        try:
            periode_date = datetime.strptime(periode_raw, '%Y-%m-%d')
        except ValueError:
            return jsonify({'status': 'error', 'message': 'Format periode harus YYYY-MM-DD'}), 400

        awal_hari = periode_date
        akhir_hari = periode_date + timedelta(days=1)
        query = query.filter(MfUnitKerja.UPDATE_DATE >= awal_hari, MfUnitKerja.UPDATE_DATE < akhir_hari)

    # --- Filter field1/field2 ---
    # Nama Unit Kerja -> kolom teks, pakai partial match (ilike)
    # Unit Kerja ID   -> kolom Integer, exact match angka
    for field, keyword in [(field1, keyword1), (field2, keyword2)]:
        if not field or not keyword:
            continue  # filter tidak dipakai -> skip, tidak wajib diisi

        if field == 'Unit Kerja ID':
            try:
                nilai = int(keyword)
            except ValueError:
                return jsonify({'status': 'error', 'message': 'Unit Kerja ID harus berupa angka'}), 400
            query = query.filter(MfUnitKerja.UNIT_KERJA_ID == nilai)
        elif field == 'Nama Unit Kerja':
            query = query.filter(MfUnitKerja.NAMA_UNIT_KERJA.ilike(f'%{keyword}%'))

    unit_kerja_list = query.order_by(MfUnitKerja.URUT_REPORT.asc(), MfUnitKerja.NAMA_UNIT_KERJA.asc()).all()

    tipe_label = {1: 'Pusat', 2: 'Pos'}

    data = [
        {
            'no': idx + 1,
            'unit_kerja_id': row.UNIT_KERJA_ID,
            'nama_unit_kerja': row.NAMA_UNIT_KERJA or '-',
            'tipe': tipe_label.get(row.IS_PUSAT, '-'),
            'is_aktif': row.IS_AKTIF,
            'is_use': row.IS_USE,
            'urut_report': row.URUT_REPORT if row.URUT_REPORT is not None else '-',
            'updated': row.UPDATE_DATE.strftime('%d-%m-%Y %H:%M') if row.UPDATE_DATE else '-',
        }
        for idx, row in enumerate(unit_kerja_list)
    ]

    return jsonify({'status': 'success', 'data': data})

def cari_user_account():
    """Render halaman Cari User Account."""
    return render_template('pages/dashboard_1/Cari User Account.html')

def get_user_account_list():
    """
    Ambil data untuk tabel Cari User Account.
    Join PEGAWAI + USER_ACCOUNT (INNER JOIN) -- hanya pegawai yang
    SUDAH punya akun sistem yang ditampilkan, karena ini halaman
    pencarian akun, bukan daftar seluruh pegawai.

    Filter opsional (semua bisa kosong -> berlaku seperti klik Refresh
    biasa, menampilkan seluruh data):
      - field1/keyword1 dan field2/keyword2 : dua dropdown "Filter"
        (Gol, Jabatan, Jenis Kelamin, Nama Peg, NIP, Unit Kerja),
        digabung dengan AND.

    Catatan: "Gol" dan "No Finger" sengaja belum dimasukkan ke field_map
    -- sama seperti di get_pegawai_vip_list(), butuh model master
    tambahan (MF_GOLONGAN, master no-finger) yang belum tersedia.
    Kalau dipilih tapi belum didukung, akan dikembalikan error yang jelas.
    """
    field1 = request.args.get('field1')
    keyword1 = request.args.get('keyword1', '').strip()
    field2 = request.args.get('field2')
    keyword2 = request.args.get('keyword2', '').strip()

    field_map = {
        'Nama Peg': Pegawai.NAMA,
        'NIP': Pegawai.NIP,
        'Jabatan': MfJabatan.NAMA_JABATAN,
        'Jenis Kelamin': Pegawai.JENIS_KEL,
        'Unit Kerja': MfUnitKerja.NAMA_UNIT_KERJA,
    }
    not_yet_supported = ('Gol', 'No Finger')

    # ============================================================
    # HRIS REBORN BUSINESS RULE
    #
    # Jabatan resmi user berasal dari:
    #
    #   UserAccount.NIP
    #        ↓
    #   Pegawai.NIP
    #        ↓
    #   Pegawai.JABATAN_ID
    #        ↓
    #   MF_JABATAN.JABATAN_ID
    #        ↓
    #   MF_JABATAN.NAMA_JABATAN
    #
    # Pegawai.JABATAN adalah field legacy.
    # ============================================================

    query = db.session.query(
        UserAccount,
        Pegawai,
        MfJabatan
    ).join(
        Pegawai,
        UserAccount.NIP == Pegawai.NIP
    ).join(
        MfUnitKerja,
        Pegawai.UNIT_KERJA_ID == MfUnitKerja.UNIT_KERJA_ID
    ).outerjoin(
        MfJabatan,
        Pegawai.JABATAN_ID == MfJabatan.JABATAN_ID
    ).filter(
        UserAccount.MODUL == 'HRIS',
        MfUnitKerja.IS_AKTIF == 'Y'
    )

    for field, keyword in [(field1, keyword1), (field2, keyword2)]:
        if not field or not keyword:
            continue  # filter ini tidak dipakai -> skip, tidak wajib diisi

        if field in not_yet_supported:
            return jsonify({
                'status': 'error',
                'message': f'Filter "{field}" belum didukung (master data belum tersedia)'
            }), 400

        column = field_map.get(field)
        if column is not None:
            query = query.filter(column.ilike(f'%{keyword}%'))

    rows = query.order_by(Pegawai.NAMA.asc()).all()

    data = [
        {
            'no': idx + 1,
            'user_id': user_account.NIP,
            'nama': pegawai.NAMA if pegawai else '-',
            'jabatan': (
                jabatan.NAMA_JABATAN
                if jabatan and jabatan.NAMA_JABATAN
                else '-'
            ),
            'level': LEVEL_LABEL.get(user_account.INIT_LEVEL, '-'),
            'update_by': user_account.UPDATE_BY or '-',
        }
        for idx, (user_account, pegawai, jabatan) in enumerate(rows)
    ]

    return jsonify({'status': 'success', 'data': data})

# ---- Create ----
def create_kalender():
    """Render halaman Create Kalender."""
    return render_template('pages/dashboard_1/Create Kalender.html')
